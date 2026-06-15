"""Model postaci WFRP 4ed niezależny od interfejsu użytkownika.

Klasa :class:`DataManager` przechowuje pełny stan postaci (cechy, umiejętności,
talenty, doświadczenie, profesja/ścieżka kariery) oraz obsługuje wczytywanie i
zapisywanie do formatów Excel, PDF i natywnego JSON. Moduł nie importuje żadnych
bibliotek UI, dzięki czemu współdzieli logikę między aplikacją desktopową a
przyszłym środowiskiem web.
"""

from __future__ import annotations

import json
import os
import re
from typing import Dict, List, Optional

import openpyxl

import game_data
from core.rules import ATTRIBUTES
from pdf_character_io import extract_pdf_character_data, write_pdf_character_data

# --- Stałe formatów plików ---
FILE_TYPE_EXCEL = "excel"
FILE_TYPE_PDF = "pdf"
FILE_TYPE_JSON = "json"

CHARACTER_JSON_SCHEMA = "wfrp4e-character"
CHARACTER_JSON_VERSION = 1

EMPTY_PDF_TEMPLATE = game_data.resource_path(f"{game_data.DATA_DIR}/empty_card.pdf")

# --- Układ arkusza Excel ---
EXCEL_CHAR_COL_START = 2  # Kolumna B
EXCEL_ATTR_NAME_ROW = 11
EXCEL_ATTR_INITIAL_ROW = 12
EXCEL_ATTR_ADVANCED_ROW = 13
EXCEL_ATTR_CURRENT_ROW = 14
EXCEL_SKILL_ROWS = range(18, 31)
EXCEL_SKILL_BLOCKS = 3
EXCEL_SKILLS_PER_BLOCK = len(EXCEL_SKILL_ROWS)
EXCEL_EXP_COL = "P"  # Aktualne PD
EXCEL_SPENT_COL = "Q"  # Wydane PD
EXCEL_TOTAL_COL = "R"  # Suma PD


class DataManager:
    """Zarządza ładowaniem i zapisywaniem danych postaci."""

    def __init__(self):
        self.file_path: Optional[str] = None
        self.source_type: str = FILE_TYPE_EXCEL
        self.attributes: Dict = {}
        self.skills: Dict = {}
        self.talents: Dict = {}
        self.experience: Dict = {"available": 0, "spent": 0, "total": 0}
        self.character_name: str = "Nowa Postać"
        self.pdf_mapping: Dict = {}
        # Profesja / klasa / ścieżka kariery
        self.character_class: str = ""
        self.character_species: str = ""
        self.current_career: str = ""
        self.current_career_level: int = 1
        self.career_path: List[Dict] = []
        self.profession_raw: Dict = {}

    def load_from_excel(self, file_path: str) -> bool:
        """Ładuje dane z pliku Excel. Zwraca True jeśli sukces."""
        try:
            self.file_path = file_path
            self.source_type = FILE_TYPE_EXCEL
            self.pdf_mapping = {}
            self.character_class = ""
            self.character_species = ""
            self.current_career = ""
            self.current_career_level = 1
            self.career_path = []
            self.profession_raw = {}
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Odczyt charakteru postaci
            self.character_name = ws.cell(5, 2).value or "Brak Imienia"

            # Odczyt cech
            self.attributes = {}
            for col_idx, attr_name in enumerate(ATTRIBUTES, start=EXCEL_CHAR_COL_START):
                cell_name = ws.cell(EXCEL_ATTR_NAME_ROW, col_idx).value
                if not cell_name:
                    continue

                initial = ws.cell(EXCEL_ATTR_INITIAL_ROW, col_idx).value or 0
                advanced = ws.cell(EXCEL_ATTR_ADVANCED_ROW, col_idx).value or 0
                current = ws.cell(EXCEL_ATTR_CURRENT_ROW, col_idx).value or (initial + advanced)

                self.attributes[attr_name] = {
                    "initial": int(initial),
                    "advanced": int(advanced),
                    "current": int(current),
                    "base_advanced": int(advanced),  # Do resetu
                    "is_new": False,
                    "profession_available": False,
                }

            # Odczyt umiejętności
            self.skills = {}
            for block in range(EXCEL_SKILL_BLOCKS):
                col_offset = block * 5
                for row in EXCEL_SKILL_ROWS:
                    skill_name = ws.cell(row, 1 + col_offset).value
                    if not skill_name:
                        continue

                    attribute = ws.cell(row, 2 + col_offset).value or ""
                    initial = ws.cell(row, 3 + col_offset).value or 0
                    advanced = ws.cell(row, 4 + col_offset).value or 0
                    current = ws.cell(row, 5 + col_offset).value or (initial + advanced)

                    self.skills[skill_name] = {
                        "attribute": str(attribute).strip("()"),
                        "initial": int(initial),
                        "advanced": int(advanced),
                        "current": int(current),
                        "base_advanced": int(advanced),
                        "is_new": False,
                        "profession_available": False,
                    }

            # Odczyt doświadczenia
            self.experience["available"] = ws[f"{EXCEL_EXP_COL}11"].value or 0
            self.experience["spent"] = ws[f"{EXCEL_SPENT_COL}11"].value or 0
            self.experience["total"] = (
                self.experience["available"] + self.experience["spent"]
            )

            wb.close()
            return True

        except Exception as e:
            print(f"Błąd przy ładowaniu pliku: {e}")
            return False

    def load_from_pdf(self, file_path: str) -> bool:
        """Ładuje dane z formularza PDF. Zwraca True jeśli sukces."""
        try:
            payload = extract_pdf_character_data(file_path)
            self.file_path = file_path
            self.source_type = FILE_TYPE_PDF
            self.character_name = payload["character_name"]
            self.attributes = payload["attributes"]
            self.skills = payload["skills"]
            self.talents = payload["talents"]
            self.experience = payload["experience"]
            self.pdf_mapping = payload["pdf_mapping"]
            self._enrich_talents()
            self._load_profession(payload.get("profession_info", {}))
            return True
        except Exception as e:
            print(f"Błąd przy ładowaniu PDF: {e}")
            return False

    def save_to_excel(self, file_path: str = None) -> bool:
        """Zapisuje dane do Excel. Zwraca True jeśli sukces."""
        if not file_path and not self.file_path:
            return False

        file_path = file_path or self.file_path

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            if len(self.skills) > EXCEL_SKILL_BLOCKS * EXCEL_SKILLS_PER_BLOCK:
                raise ValueError(
                    "Za dużo umiejętności, aby zapisać je do aktualnego układu Excela."
                )

            ws.cell(5, 2, self.character_name)

            # Zapis cech
            for col_idx, attr_name in enumerate(ATTRIBUTES, start=EXCEL_CHAR_COL_START):
                if attr_name in self.attributes:
                    attr_data = self.attributes[attr_name]
                    ws.cell(EXCEL_ATTR_INITIAL_ROW, col_idx, attr_data["initial"])
                    ws.cell(EXCEL_ATTR_ADVANCED_ROW, col_idx, attr_data["advanced"])
                    ws.cell(
                        EXCEL_ATTR_CURRENT_ROW,
                        col_idx,
                        attr_data["initial"] + attr_data["advanced"],
                    )

            # Wyczyść wszystkie sloty umiejętności przed zapisem
            for block in range(EXCEL_SKILL_BLOCKS):
                col_offset = block * 5
                for row_idx in EXCEL_SKILL_ROWS:
                    for column_offset in range(5):
                        ws.cell(row_idx, 1 + col_offset + column_offset, None)

            # Zapis umiejętności
            for skill_index, (skill_name, skill_data) in enumerate(self.skills.items()):
                block = skill_index // EXCEL_SKILLS_PER_BLOCK
                row_offset = skill_index % EXCEL_SKILLS_PER_BLOCK
                col_offset = block * 5
                row_idx = EXCEL_SKILL_ROWS.start + row_offset

                ws.cell(row_idx, 1 + col_offset, skill_name)
                ws.cell(row_idx, 2 + col_offset, f"({skill_data['attribute']})")
                ws.cell(row_idx, 3 + col_offset, skill_data["initial"])
                ws.cell(row_idx, 4 + col_offset, skill_data["advanced"])
                ws.cell(
                    row_idx,
                    5 + col_offset,
                    skill_data["initial"] + skill_data["advanced"],
                )

            # Zapis doświadczenia
            ws[f"{EXCEL_EXP_COL}11"] = self.experience["available"]
            ws[f"{EXCEL_SPENT_COL}11"] = self.experience["spent"]
            ws[f"{EXCEL_TOTAL_COL}11"] = self.experience["total"]

            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            print(f"Błąd przy zapisie pliku: {e}")
            return False

    def save_to_pdf(self, file_path: str = None) -> bool:
        """Zapisuje dane do formularza PDF. Zwraca True jeśli sukces."""
        if not file_path and not self.file_path:
            return False

        file_path = file_path or self.file_path

        try:
            if not self.pdf_mapping:
                return False

            write_pdf_character_data(
                self.file_path,
                file_path,
                {
                    "character_name": self.character_name,
                    "attributes": self.attributes,
                    "skills": self.skills,
                    "talents": self.talents,
                    "experience": self.experience,
                    "profession": self.profession_payload(),
                },
                self.pdf_mapping,
            )
            return True
        except Exception as e:
            print(f"Błąd przy zapisie PDF: {e}")
            return False

    def export_to_pdf(self, target_path: str) -> bool:
        """Eksportuje postać do PDF. Dla postaci wczytanej z PDF używa jej jako
        wzorca; dla postaci utworzonej od zera używa pustego szablonu karty."""
        try:
            if (
                self.source_type == FILE_TYPE_PDF
                and self.file_path
                and self.pdf_mapping
                and os.path.exists(self.file_path)
            ):
                source_pdf = self.file_path
                mapping = self.pdf_mapping
            else:
                if not os.path.exists(EMPTY_PDF_TEMPLATE):
                    print("PDF export error: empty template missing")
                    return False
                source_pdf = str(EMPTY_PDF_TEMPLATE)
                mapping = extract_pdf_character_data(source_pdf)["pdf_mapping"]

            write_pdf_character_data(
                source_pdf,
                target_path,
                {
                    "character_name": self.character_name,
                    "attributes": self.attributes,
                    "skills": self.skills,
                    "talents": self.talents,
                    "experience": self.experience,
                    "profession": self.profession_payload(),
                },
                mapping,
            )
            return True
        except Exception as e:
            print(f"PDF export error: {e}")
            return False

    def add_skill(
        self,
        skill_name: str,
        attribute: str,
        initial: int = 0,
        advanced: int = 0,
        is_new: bool = False,
    ) -> bool:
        """Dodaje nową umiejętność."""
        if skill_name in self.skills:
            return False

        self.skills[skill_name] = {
            "attribute": attribute,
            "initial": initial,
            "advanced": advanced,
            "current": initial + advanced,
            "base_advanced": advanced,
            "is_new": is_new,
            "profession_available": False,
        }
        return True

    # ----- Talenty ---------------------------------------------------------
    def _enrich_talents(self) -> None:
        """Uzupełnia talenty wczytane z PDF o dane z bazy (Max, opis, źródło)."""
        database = game_data.load_talents()
        base_index = self._talent_base_index(database)
        enriched: Dict = {}
        for name, raw in self.talents.items():
            advances = self._parse_talent_advances(raw.get("advances"))
            db_entry = self._match_talent_db_entry(name, database, base_index)
            description = raw.get("description") or (
                db_entry.get("description") if db_entry else ""
            )
            enriched[name] = {
                "advances": advances,
                "base_advances": advances,
                "description": description,
                "max": (db_entry or {}).get("max"),
                "tests": (db_entry or {}).get("tests"),
                "source": (db_entry or {}).get("source", "Karta PDF"),
                "is_new": False,
                "is_custom": db_entry is None,
                "profession_available": False,
            }
        self.talents = enriched

    @staticmethod
    def _talent_base_index(database: Dict) -> Dict[str, str]:
        """Indeks {nazwa_bazowa_bez_specjalizacji -> klucz_bazy} do dopasowań."""
        index: Dict[str, str] = {}
        for key in database:
            base = game_data._normalize(str(key).split("(")[0])
            if base:
                index.setdefault(base, key)
        return index

    @staticmethod
    def _match_talent_db_entry(name, database: Dict, base_index: Dict[str, str]):
        """Dopasowuje talent z PDF do wpisu bazy (dokładnie lub po nazwie bazowej)."""
        entry = database.get(name)
        if entry is not None:
            return entry
        base = game_data._normalize(str(name).split("(")[0])
        matched_key = base_index.get(base)
        if matched_key:
            return database.get(matched_key)
        return None

    @staticmethod
    def _parse_talent_advances(value) -> int:
        """Zamienia wartość 'times taken' z PDF na liczbę (puste -> 1)."""
        if value is None:
            return 1
        text = str(value).strip()
        if not text:
            return 1
        match = re.search(r"\d+", text)
        return int(match.group()) if match else 1

    def talent_max_advances(self, name: str) -> Optional[int]:
        """Zwraca twardy limit wykupień talentu (None = brak limitu liczbowego)."""
        talent = self.talents.get(name)
        info = talent.get("max") if talent else None
        if not info:
            return None
        kind = info.get("type")
        if kind == "fixed":
            return info.get("value")
        if kind == "characteristic":
            attr = info.get("attr")
            attr_data = self.attributes.get(attr)
            if not attr_data:
                return None
            return max(1, game_data.attribute_bonus(attr_data.get("current", 0)))
        return None

    def add_talent(
        self,
        name: str,
        advances: int = 1,
        description: str = "",
        max_info: Optional[Dict] = None,
        tests: Optional[str] = None,
        source: str = "Lista",
        is_custom: bool = False,
        is_new: bool = True,
    ) -> bool:
        """Dodaje talent do postaci. Zwraca False, jeśli już istnieje."""
        if name in self.talents:
            return False
        self.talents[name] = {
            "advances": advances,
            "base_advances": 0 if is_new else advances,
            "description": description,
            "max": max_info,
            "tests": tests,
            "source": source,
            "is_new": is_new,
            "is_custom": is_custom,
            "profession_available": False,
        }
        return True

    # ----- Profesja / klasa / ścieżka kariery -----------------------------
    def _load_profession(self, info: Dict) -> None:
        """Wczytuje dane profesji z pól PDF i buduje ścieżkę kariery."""
        self.profession_raw = dict(info or {})
        self.character_class = (info or {}).get("class", "") or ""
        self.character_species = (info or {}).get("species", "") or ""

        profession_field = (info or {}).get("profession", "") or ""
        path_text = (info or {}).get("path_text", "") or ""
        level_text = (info or {}).get("level_text", "") or ""

        steps = game_data.resolve_career_path(path_text)
        # Jeśli ścieżka pusta, ale jest profesja, zbuduj jednoelementową ścieżkę
        if not steps and profession_field:
            steps = game_data.resolve_career_path(profession_field)

        parsed_level = game_data.parse_career_level(level_text)
        self.career_path = self._build_career_path(steps, profession_field, parsed_level)

        if self.career_path:
            last = self.career_path[-1]
            self.current_career = last.get("profession") or last.get("title") or profession_field
            self.current_career_level = last.get("level") or parsed_level or 1
        else:
            self.current_career = profession_field
            self.current_career_level = parsed_level or 1

        # Uzupełnij klasę z danych gry, jeśli pole PDF puste lub niespójne
        resolved_class = game_data.class_of_career(self.current_career)
        if resolved_class:
            self.character_class = resolved_class

    def _build_career_path(
        self, steps: List[Dict], current_profession: str, current_level: Optional[int]
    ) -> List[Dict]:
        """Składa listę kroków kariery z oznaczeniem poziomu i kompletowania."""
        path: List[Dict] = []
        for index, step in enumerate(steps):
            is_last = index == len(steps) - 1
            level = step.get("level") or 1
            if is_last and current_level:
                level = current_level
            entry = {
                "title": step.get("title", ""),
                "profession": step.get("profession"),
                "level": level,
                "resolved": step.get("resolved", False),
                "completed": not is_last,
            }
            path.append(entry)
        return path

    def profession_payload(self) -> Dict:
        """Buduje słownik pól profesji do zapisu w PDF."""
        path_titles = [
            (step.get("title") or step.get("profession") or "")
            for step in self.career_path
            if (step.get("title") or step.get("profession"))
        ]
        level_text = self.profession_raw.get("level_text", "")
        if self.current_career:
            level_text = f"{self.current_career} ({self.current_career_level})"
        return {
            "class": self.character_class,
            "profession": self.current_career,
            "level_text": level_text,
            "path_text": ", ".join(path_titles),
            "species": self.character_species,
        }

    def set_current_career(self, profession: str, level: int) -> None:
        """Ustawia/poprawia bieżącą profesję i poziom (bez kosztu)."""
        self.current_career = profession
        self.current_career_level = max(1, min(4, int(level)))
        resolved_class = game_data.class_of_career(profession)
        if resolved_class:
            self.character_class = resolved_class
        if not self.career_path:
            self.career_path = [
                {
                    "title": profession,
                    "profession": profession if game_data.get_profession(profession) else None,
                    "level": self.current_career_level,
                    "resolved": bool(game_data.get_profession(profession)),
                    "completed": False,
                }
            ]
        else:
            last = self.career_path[-1]
            last["title"] = profession
            last["profession"] = profession if game_data.get_profession(profession) else None
            last["level"] = self.current_career_level
            last["resolved"] = bool(game_data.get_profession(profession))
            last["completed"] = False

    def advance_to_career(self, profession: str, level: int) -> None:
        """Awansuje do nowej profesji: oznacza obecną jako ukończoną i dopisuje krok."""
        level = max(1, min(4, int(level)))
        if self.career_path:
            self.career_path[-1]["completed"] = True
        resolved = bool(game_data.get_profession(profession))
        self.career_path.append(
            {
                "title": profession,
                "profession": profession if resolved else None,
                "level": level,
                "resolved": resolved,
                "completed": False,
            }
        )
        self.current_career = profession
        self.current_career_level = level
        resolved_class = game_data.class_of_career(profession)
        if resolved_class:
            self.character_class = resolved_class

    def current_career_completion(self) -> Dict:
        """Zwraca status kompletowania bieżącej profesji."""
        profession = self.current_career
        if not game_data.get_profession(profession):
            return {
                "completed": False,
                "skills_ok": False,
                "talents_ok": False,
                "characteristics_ok": False,
                "skills_done": 0,
                "talents_done": 0,
                "characteristics_pending": True,
                "unknown_profession": True,
            }
        result = game_data.is_career_completed(
            profession,
            self.current_career_level,
            self.skills,
            self.talents,
            self.attributes,
        )
        result["unknown_profession"] = False
        return result

    def reset_character(self) -> None:
        """Resetuje postać do wartości z pliku."""
        for attr in self.attributes.values():
            attr["advanced"] = attr["base_advanced"]

        for skill in self.skills.values():
            skill["advanced"] = skill["base_advanced"]

    def create_new_character(self, name: str = "Nowa Postać") -> None:
        """Tworzy nową postać z wartościami domyślnymi."""
        self.character_name = name
        self.file_path = None
        self.source_type = FILE_TYPE_EXCEL
        self.pdf_mapping = {}
        self.attributes = {
            attr: {
                "initial": 30,
                "advanced": 0,
                "current": 30,
                "base_advanced": 0,
                "is_new": False,
                "profession_available": False,
            }
            for attr in ATTRIBUTES
        }
        self.skills = {}
        self.experience = {"available": 0, "spent": 0, "total": 0}
        self.talents = {}
        self.character_class = ""
        self.character_species = ""
        self.current_career = ""
        self.current_career_level = 1
        self.career_path = []
        self.profession_raw = {}

    def to_dict(self) -> Dict:
        """Serializuje pełny stan postaci do słownika (kanoniczny format JSON)."""
        return {
            "schema": CHARACTER_JSON_SCHEMA,
            "version": CHARACTER_JSON_VERSION,
            "character_name": self.character_name,
            "character_class": self.character_class,
            "character_species": self.character_species,
            "current_career": self.current_career,
            "current_career_level": self.current_career_level,
            "career_path": [dict(step) for step in self.career_path],
            "profession_raw": dict(self.profession_raw),
            "source_type": self.source_type,
            "attributes": {k: dict(v) for k, v in self.attributes.items()},
            "skills": {k: dict(v) for k, v in self.skills.items()},
            "talents": {k: dict(v) for k, v in self.talents.items()},
            "experience": dict(self.experience),
        }

    def from_dict(self, data: Dict) -> bool:
        """Wczytuje stan postaci ze słownika (kanoniczny format JSON)."""
        try:
            self.character_name = data.get("character_name", "Nowa Postać")
            self.character_class = data.get("character_class", "") or ""
            self.character_species = data.get("character_species", "") or ""
            self.current_career = data.get("current_career", "") or ""
            self.current_career_level = int(data.get("current_career_level", 1) or 1)
            self.career_path = [dict(s) for s in data.get("career_path", [])]
            self.profession_raw = dict(data.get("profession_raw", {}) or {})
            self.source_type = data.get("source_type", FILE_TYPE_EXCEL) or FILE_TYPE_EXCEL
            self.attributes = {
                k: dict(v) for k, v in (data.get("attributes", {}) or {}).items()
            }
            self.skills = {k: dict(v) for k, v in (data.get("skills", {}) or {}).items()}
            self.talents = {
                k: dict(v) for k, v in (data.get("talents", {}) or {}).items()
            }
            exp = data.get("experience", {}) or {}
            self.experience = {
                "available": int(exp.get("available", 0) or 0),
                "spent": int(exp.get("spent", 0) or 0),
                "total": int(exp.get("total", 0) or 0),
            }
            self.pdf_mapping = {}
            return True
        except Exception as e:
            print(f"JSON decode error: {e}")
            return False

    def save_to_json(self, file_path: str) -> bool:
        """Zapisuje postać do natywnego pliku JSON. Zwraca True przy sukcesie."""
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(self.to_dict(), f, ensure_ascii=False, indent=2)
            self.file_path = file_path
            return True
        except Exception as e:
            print(f"JSON save error: {e}")
            return False

    def load_from_json(self, file_path: str) -> bool:
        """Wczytuje postać z natywnego pliku JSON. Zwraca True przy sukcesie."""
        try:
            with open(file_path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"JSON load error: {e}")
            return False
        if not self.from_dict(data):
            return False
        self.file_path = file_path
        return True
