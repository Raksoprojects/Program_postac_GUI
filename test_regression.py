#!/usr/bin/env python3
"""Testy regresyjne dla kluczowych scenariuszy aplikacji."""

import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from Postac_program_gui import (
    CharacterSheetApp,
    DataManager,
    HistoryManager,
    calculate_advancement_cost,
    calculate_talent_cost,
)


WORKSPACE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = WORKSPACE_DIR / "karta_postaci.xlsx"
PDF_FILE = WORKSPACE_DIR / "Rein_Nuhr_lepsza_4ed.pdf"


class CharacterSheetAppRegressionTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_excel = Path(self.temp_dir.name) / "karta_postaci_test.xlsx"
        shutil.copyfile(EXCEL_FILE, self.temp_excel)
        self.temp_history = Path(self.temp_dir.name) / "history_test.json"
        self.app = CharacterSheetApp()
        self.app.withdraw()
        self.app.history_manager.json_file = str(self.temp_history)
        self.app.history_manager.history = []

    def tearDown(self):
        self.app.destroy()
        self.temp_dir.cleanup()

    def load_test_excel(self):
        self.assertTrue(self.app.data_manager.load_from_excel(str(self.temp_excel)))

    def test_confirm_keeps_reserved_xp_without_double_charge(self):
        self.load_test_excel()

        start_available = self.app.data_manager.experience["available"]
        start_spent = self.app.data_manager.experience["spent"]
        cost = calculate_advancement_cost(
            "cecha",
            self.app.data_manager.attributes["WW"]["advanced"],
            1,
        )

        self.app.on_increase_attribute("WW", 1)
        self.assertEqual(
            self.app.data_manager.experience["available"],
            start_available - cost,
        )

        with patch("Postac_program_gui.messagebox.askyesno", return_value=True), patch(
            "Postac_program_gui.messagebox.showinfo"
        ):
            self.app.on_confirm_changes()

        self.assertEqual(
            self.app.data_manager.experience["available"],
            start_available - cost,
        )
        self.assertEqual(self.app.data_manager.experience["spent"], start_spent + cost)

    def test_cost_preview_uses_current_advancements(self):
        self.load_test_excel()

        selection = "WW"
        desired_advancements = 7
        current_adv = self.app.data_manager.attributes[selection]["advanced"]

        preview = self.app.get_cost_preview(selection, desired_advancements)

        self.assertIsNotNone(preview)
        self.assertEqual(preview["type"], "Cecha")
        self.assertEqual(preview["current_adv"], current_adv)
        self.assertEqual(
            preview["cost"],
            calculate_advancement_cost("cecha", current_adv, desired_advancements),
        )

    def test_new_skill_is_finalized_after_confirm(self):
        self.app.data_manager.create_new_character("Tester")
        self.app.data_manager.add_skill(
            "Nowa Umiejętność",
            "WW",
            initial=30,
            advanced=2,
            is_new=True,
        )
        self.app.pending_changes["new_skills"].add("Nowa Umiejętność")

        with patch("Postac_program_gui.messagebox.askyesno", return_value=True), patch(
            "Postac_program_gui.messagebox.showinfo"
        ):
            self.app.on_confirm_changes()

        self.assertIn("Nowa Umiejętność", self.app.data_manager.skills)
        self.assertFalse(self.app.data_manager.skills["Nowa Umiejętność"]["is_new"])
        self.assertEqual(self.app.data_manager.skills["Nowa Umiejętność"]["base_advanced"], 2)

    def test_exact_attribute_filter_shows_matching_skills(self):
        self.load_test_excel()
        self.app.initialize_skills_display()

        self.app.skill_attribute_filter_var.set("Inteligencja (Int)")
        self.app.apply_skill_filter()

        visible_skills = {
            skill_name
            for skill_name, row_data in self.app.skill_rows.items()
            if row_data["row"].winfo_manager() == "pack"
        }

        self.assertTrue(visible_skills)
        self.assertTrue(
            all(
                self.app.data_manager.skills[skill_name]["attribute"] == "Int"
                for skill_name in visible_skills
            )
        )

    def test_skill_order_returns_after_clearing_filter(self):
        self.load_test_excel()
        self.app.initialize_skills_display()
        self.app.update_idletasks()

        basic_content = self.app.skills_group_frames["Podstawowa"]["content"]
        row_to_name = {
            row_data["row"]: skill_name
            for skill_name, row_data in self.app.skill_rows.items()
        }
        original_order = [row_to_name[row] for row in basic_content.pack_slaves()]

        self.app.skill_filter_var.set("Plotkowanie")
        self.app.apply_skill_filter()
        self.app.clear_skill_filter()
        self.app.update_idletasks()

        restored_order = [row_to_name[row] for row in basic_content.pack_slaves()]
        self.assertEqual(restored_order, original_order)

    def test_profession_only_filter_shows_only_plus_skills(self):
        self.assertTrue(self.app.data_manager.load_from_pdf(str(PDF_FILE)))
        self.app.initialize_skills_display()

        self.app.skill_profession_filter_var.set(True)
        self.app.apply_skill_filter()

        visible_skills = {
            skill_name
            for skill_name, row_data in self.app.skill_rows.items()
            if row_data["row"].winfo_manager() == "pack"
        }

        self.assertTrue(visible_skills)
        self.assertTrue(
            all(
                self.app.data_manager.skills[skill_name].get("profession_available")
                for skill_name in visible_skills
            )
        )


class DataManagerRegressionTests(unittest.TestCase):
    def test_save_to_excel_distributes_skills_between_blocks(self):
        data_manager = DataManager()
        self.assertTrue(data_manager.load_from_excel(str(EXCEL_FILE)))

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_excel = Path(temp_dir) / "karta_postaci_copy.xlsx"
            shutil.copyfile(EXCEL_FILE, temp_excel)

            self.assertTrue(data_manager.save_to_excel(str(temp_excel)))

            workbook = load_workbook(temp_excel, data_only=True)
            worksheet = workbook.active
            block_lengths = []

            for block in range(3):
                col_offset = block * 5
                block_rows = []
                for row in range(18, 31):
                    skill_name = worksheet.cell(row, 1 + col_offset).value
                    if skill_name:
                        block_rows.append(skill_name)
                block_lengths.append(len(block_rows))

            workbook.close()

        self.assertEqual(sum(block_lengths), len(data_manager.skills))
        self.assertEqual(block_lengths, [13, 13, 10])

    def test_load_from_pdf_reads_rein_and_experience(self):
        data_manager = DataManager()

        self.assertTrue(data_manager.load_from_pdf(str(PDF_FILE)))

        self.assertEqual(data_manager.character_name, "Rein Nuhr")
        self.assertEqual(data_manager.experience["available"], 185)
        self.assertEqual(data_manager.experience["spent"], 2260)
        self.assertEqual(data_manager.experience["total"], 2445)
        self.assertIn("Splatanie (Aqshy)", data_manager.skills)
        self.assertIn("Broń Biała (Drzewcowa)", data_manager.skills)

    def test_load_from_pdf_marks_profession_advancements(self):
        data_manager = DataManager()

        self.assertTrue(data_manager.load_from_pdf(str(PDF_FILE)))

        self.assertTrue(data_manager.attributes["WW"]["profession_available"])
        self.assertFalse(data_manager.attributes["US"]["profession_available"])
        self.assertTrue(data_manager.skills["Broń Biała (Podstawowa)"]["profession_available"])
        self.assertFalse(data_manager.skills["Atletyka"]["profession_available"])

    def test_save_to_pdf_updates_form_fields(self):
        data_manager = DataManager()
        self.assertTrue(data_manager.load_from_pdf(str(PDF_FILE)))
        talent_name = next(iter(data_manager.talents))
        basic_skill_mapping = data_manager.pdf_mapping["skills"]["Broń Biała (Drzewcowa)"]
        advanced_skill_mapping = data_manager.pdf_mapping["skills"]["Splatanie (Aqshy)"]
        talent_mapping = data_manager.pdf_mapping["talents"][talent_name]

        data_manager.character_name = "Rein Test"
        data_manager.experience["available"] = 321
        data_manager.attributes["WW"]["advanced"] = 9
        data_manager.skills["Broń Biała (Drzewcowa)"]["advanced"] = 11
        data_manager.skills["Splatanie (Aqshy)"]["advanced"] = 7
        data_manager.skills["Splatanie (Aqshy)"]["initial"] = 48
        data_manager.talents[talent_name] = {
            "advances": "2",
            "description": "Test opisu talentu",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf = Path(temp_dir) / "rein_test.pdf"
            self.assertTrue(data_manager.save_to_pdf(str(temp_pdf)))

            from pypdf import PdfReader

            reader = PdfReader(str(temp_pdf))
            fields = reader.get_fields() or {}

            self.assertEqual(fields["Imię"].get("/V"), "Rein Test")
            self.assertEqual(str(fields["Aktualne_doświadczenie"].get("/V")), "321")
            self.assertEqual(str(fields["WW_rozwieniecie"].get("/V")), "9")
            self.assertEqual(
                str(fields[basic_skill_mapping["specialization_field"]].get("/V")),
                "Drzewcowa",
            )
            self.assertEqual(
                str(fields[basic_skill_mapping["advanced_field"]].get("/V")),
                "11",
            )
            self.assertEqual(
                str(fields[advanced_skill_mapping["advanced_field"]].get("/V")),
                "7",
            )
            self.assertEqual(
                str(fields[advanced_skill_mapping["initial_field"]].get("/V")),
                "48",
            )
            self.assertEqual(
                str(fields[advanced_skill_mapping["current_field"]].get("/V")),
                "55",
            )
            self.assertEqual(
                str(fields[talent_mapping["name_field"]].get("/V")),
                talent_name,
            )
            self.assertEqual(
                str(fields[talent_mapping["advances_field"]].get("/V")),
                "2",
            )
            self.assertEqual(
                str(fields[talent_mapping["description_field"]].get("/V")),
                "Test opisu talentu",
            )



class GameDataRegressionTests(unittest.TestCase):
    def test_data_loads_with_expected_counts(self):
        import game_data

        self.assertEqual(len(game_data.load_classes()), 8)
        self.assertEqual(len(game_data.load_professions()), 64)
        self.assertGreaterEqual(len(game_data.load_talents()), 160)

    def test_validation_has_no_errors(self):
        import game_data

        result = game_data.validate_data()
        self.assertEqual(result["errors"], [])

    def test_every_profession_has_four_levels(self):
        import game_data

        for name, prof in game_data.load_professions().items():
            self.assertEqual(len(prof["levels"]), 4, f"Profesja {name}")

    def test_class_and_profession_cross_reference(self):
        import game_data

        for class_name in game_data.all_class_names():
            for career in game_data.careers_for_class(class_name):
                self.assertEqual(game_data.class_of_career(career), class_name)

    def test_talent_max_fixed_and_none(self):
        import game_data

        # "Bardzo Silny" ma Maksimum: 1
        self.assertEqual(game_data.talent_max("Bardzo Silny"), 1)
        # "Arcydzieło" ma Maksimum: brak -> bez limitu
        self.assertIsNone(game_data.talent_max("Arcydzieło"))

    def test_talent_max_from_characteristic(self):
        import game_data

        # "Aptekarz": Maksimum = Bonus z Inteligencji
        attrs = {"Int": 37}
        self.assertEqual(game_data.talent_max("Aptekarz", attrs), 3)
        # bez podanych cech -> nieznany limit
        self.assertIsNone(game_data.talent_max("Aptekarz"))


class TalentModelRegressionTests(unittest.TestCase):
    def test_calculate_talent_cost_flat_and_multipliers(self):
        self.assertEqual(calculate_talent_cost(1), 100)
        self.assertEqual(calculate_talent_cost(3), 300)
        # rozwój spoza profesji -> podwójny koszt
        self.assertEqual(calculate_talent_cost(2, out_of_profession=True), 400)
        # zgoda MG znosi mnożnik
        self.assertEqual(
            calculate_talent_cost(2, out_of_profession=True, gm_approved=True), 200
        )

    def test_enrich_talents_after_pdf_load(self):
        data_manager = DataManager()
        self.assertTrue(data_manager.load_from_pdf(str(PDF_FILE)))

        self.assertTrue(data_manager.talents)
        for name, talent in data_manager.talents.items():
            self.assertIsInstance(talent["advances"], int)
            self.assertEqual(talent["base_advances"], talent["advances"])
            self.assertIn("max", talent)
            self.assertFalse(talent["is_new"])

    def test_talent_max_advances_from_characteristic(self):
        data_manager = DataManager()
        data_manager.attributes["Int"] = {"current": 37}
        data_manager.talents["Próbny"] = {
            "advances": 1,
            "base_advances": 1,
            "max": {"type": "characteristic", "attr": "Int"},
        }
        # Bonus z Int 37 = 3
        self.assertEqual(data_manager.talent_max_advances("Próbny"), 3)

    def test_add_talent_dedupes(self):
        data_manager = DataManager()
        self.assertTrue(
            data_manager.add_talent("Własny", advances=1, is_custom=True, is_new=True)
        )
        self.assertFalse(data_manager.add_talent("Własny"))
        self.assertEqual(data_manager.talents["Własny"]["base_advances"], 0)

    def test_save_to_pdf_persists_new_talent_in_free_row(self):
        data_manager = DataManager()
        self.assertTrue(data_manager.load_from_pdf(str(PDF_FILE)))

        free_rows = data_manager.pdf_mapping.get("talents_free", [])
        self.assertTrue(free_rows, "Karta powinna mieć wolne wiersze talentów")
        target_name_field = free_rows[0]["name_field"]

        data_manager.add_talent(
            "Nowy Testowy Talent",
            advances=1,
            description="Opis nowego talentu",
            is_new=True,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf = Path(temp_dir) / "rein_new_talent.pdf"
            self.assertTrue(data_manager.save_to_pdf(str(temp_pdf)))

            from pypdf import PdfReader

            reader = PdfReader(str(temp_pdf))
            fields = reader.get_fields() or {}
            self.assertEqual(
                str(fields[target_name_field].get("/V")),
                "Nowy Testowy Talent",
            )


class ProfessionModelRegressionTests(unittest.TestCase):
    def test_load_profession_from_pdf(self):
        dm = DataManager()
        self.assertTrue(dm.load_from_pdf(str(PDF_FILE)))
        self.assertEqual(dm.current_career, "Piromanta")
        self.assertEqual(dm.current_career_level, 2)
        self.assertEqual(dm.character_species, "Człowiek")
        # Ścieżka: rozpoznany 'Uczeń czarodzieja' -> Czarodziej, oraz 'Piromanta' (spoza podstawki)
        self.assertEqual(len(dm.career_path), 2)
        self.assertTrue(dm.career_path[0]["resolved"])
        self.assertFalse(dm.career_path[1]["resolved"])

    def test_profession_payload_roundtrip(self):
        dm = DataManager()
        self.assertTrue(dm.load_from_pdf(str(PDF_FILE)))
        payload = dm.profession_payload()
        self.assertEqual(payload["profession"], "Piromanta")
        self.assertEqual(payload["level_text"], "Piromanta (2)")
        self.assertIn("Piromanta", payload["path_text"])

    def test_save_to_pdf_writes_profession_fields(self):
        dm = DataManager()
        self.assertTrue(dm.load_from_pdf(str(PDF_FILE)))
        dm.set_current_career("Czarodziej", 3)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf = Path(temp_dir) / "prof_test.pdf"
            self.assertTrue(dm.save_to_pdf(str(temp_pdf)))

            from pypdf import PdfReader

            reader = PdfReader(str(temp_pdf))
            fields = reader.get_fields() or {}
            self.assertEqual(str(fields["Profesja"].get("/V")), "Czarodziej")
            self.assertEqual(str(fields["Poziom Profesji"].get("/V")), "Czarodziej (3)")

    def test_set_current_career_updates_class(self):
        dm = DataManager()
        dm.set_current_career("Czarodziej", 2)
        self.assertEqual(dm.current_career, "Czarodziej")
        self.assertEqual(dm.current_career_level, 2)
        self.assertEqual(dm.character_class, "Uczeni")

    def test_advance_to_career_marks_previous_completed(self):
        dm = DataManager()
        dm.set_current_career("Czarodziej", 1)
        dm.advance_to_career("Czarodziej", 2)
        self.assertEqual(len(dm.career_path), 2)
        self.assertTrue(dm.career_path[0]["completed"])
        self.assertFalse(dm.career_path[1]["completed"])
        self.assertEqual(dm.current_career_level, 2)

    def test_career_transition_cost_rules(self):
        import game_data

        self.assertEqual(game_data.career_transition_cost(True, True), 100)
        self.assertEqual(game_data.career_transition_cost(False, True), 200)
        self.assertEqual(game_data.career_transition_cost(True, False), 200)
        self.assertEqual(game_data.career_transition_cost(False, False), 300)


class HistoryManagerRegressionTests(unittest.TestCase):
    def test_migrates_old_flat_list(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            history_file = Path(temp_dir) / "history.json"
            history_file.write_text(
                '[{"timestamp": "t", "action": "Stary wpis", "details": "x"}]',
                encoding="utf-8",
            )
            manager = HistoryManager(str(history_file))
            self.assertEqual(len(manager.data["global_log"]), 1)
            self.assertEqual(manager.data["characters"], {})

    def test_per_character_entries_and_path(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            history_file = Path(temp_dir) / "history.json"
            manager = HistoryManager(str(history_file))
            manager.set_current_character("Bohater")
            manager.add_entry("Test akcji", "szczegóły")
            self.assertEqual(len(manager.data["characters"]["Bohater"]["changes"]), 1)
            self.assertEqual(len(manager.data["global_log"]), 1)

            path = [{"title": "Czarodziej", "profession": "Czarodziej", "level": 1}]
            manager.set_career_path("Bohater", path)
            self.assertEqual(manager.get_career_path("Bohater"), path)

            # Po ponownym wczytaniu dane przetrwają
            reloaded = HistoryManager(str(history_file))
            self.assertEqual(reloaded.get_career_path("Bohater"), path)


class DevelopableRegressionTests(unittest.TestCase):
    """Faza 3: rozwijalność (gating) i koszt rozwoju spoza profesji."""

    def test_developable_skills_cumulative_up_to_level(self):
        import game_data
        dev = game_data.get_career_developable("Czarodziej", 2)
        self.assertTrue(dev["resolved"])
        # Umiejętności z poziomu 1 i 2 są rozwijalne
        self.assertIn("Intuicja", dev["skills"])          # poziom 1
        self.assertIn("Charyzma", dev["skills"])           # poziom 2
        # Umiejętność z poziomu 3 NIE jest jeszcze rozwijalna
        self.assertNotIn("Wycena", dev["skills"])

    def test_developable_talents_current_level_plus_purchased(self):
        import game_data
        # Bez wykupień: tylko talenty obecnego poziomu są rozwijalne
        dev = game_data.get_career_developable("Czarodziej", 2)
        self.assertIn("Szósty Zmysł", dev["talents"])         # poziom 2 (świeży)
        self.assertNotIn("Magia Prosta", dev["talents"])       # poziom 1, niewykupiony
        self.assertNotIn("Groźny", dev["talents"])             # poziom 3

        # Wykupiony talent z niższego poziomu staje się rozwijalny (rośnie do Max)
        owned = {"Magia Prosta": {"advances": 1}}
        dev2 = game_data.get_career_developable("Czarodziej", 2, owned)
        self.assertIn("Magia Prosta", dev2["talents"])

    def test_unknown_profession_not_resolved(self):
        import game_data
        dev = game_data.get_career_developable("Piromanta", 2)
        self.assertFalse(dev["resolved"])
        self.assertEqual(dev["skills"], set())
        self.assertEqual(dev["talents"], set())

    def test_is_skill_developable_handles_group_specialisation(self):
        import game_data
        dev = game_data.get_career_developable("Czarodziej", 4)
        # "Wiedza (Dowolna)" w schemacie poziomu 4 pasuje do konkretnej specjalizacji
        self.assertTrue(game_data.is_skill_developable("Wiedza (Reikland)", dev))
        # Konkretna umiejętność spoza schematu nie jest rozwijalna
        self.assertFalse(game_data.is_skill_developable("Hazard", dev))

    def test_out_of_profession_cost_doubles_unless_gm_approved(self):
        # Koszt rozwoju spoza profesji jest podwojony, zgoda MG przywraca x1
        base = calculate_advancement_cost("cecha", 0, 5)
        doubled = calculate_advancement_cost("cecha", 0, 5, out_of_profession=True)
        gm = calculate_advancement_cost("cecha", 0, 5, out_of_profession=True, gm_approved=True)
        self.assertEqual(doubled, base * 2)
        self.assertEqual(gm, base)

    def test_skill_developable_falls_back_to_pdf_flag_when_unresolved(self):
        manager = DataManager()
        manager.skills = {"Hazard": {"advanced": 0, "profession_available": True}}
        app = None
        try:
            app = CharacterSheetApp()
            app.withdraw()
            app.data_manager = manager
            app.data_manager.current_career = "Piromanta"  # spoza podstawki -> nierozpoznane
            app._recompute_developable()
            # Brak schematu -> fallback do flagi z PDF
            self.assertTrue(app._skill_developable("Hazard"))
            # Brak schematu -> rozwój nie jest karany jako spoza profesji
            self.assertFalse(app._is_out_of_profession("umiejetnosc", "Hazard"))
        finally:
            if app is not None:
                app.destroy()

    def test_characteristic_name_to_code_mapping_in_gating(self):
        # 4B1: pełne nazwy cech z professions.json mapowane na kody używane w GUI.
        import game_data
        dev = game_data.get_career_developable("Aptekarka", 2)
        # L1: Wytrzymałość/Zręczność/Inteligencja, L2: Ogłada -> kody
        self.assertEqual(dev["characteristics"], {"Wt", "Zr", "Int", "Ogd"})
        self.assertTrue(game_data.is_characteristic_developable("Wt", dev))
        self.assertTrue(game_data.is_characteristic_developable("Ogd", dev))
        # Inicjatywa pojawia się dopiero na L3 -> nierozwijalna na L2
        self.assertFalse(game_data.is_characteristic_developable("I", dev))

    def test_characteristic_to_code_accepts_names_and_codes(self):
        import game_data
        self.assertEqual(game_data.characteristic_to_code("Wytrzymałość"), "Wt")
        self.assertEqual(game_data.characteristic_to_code("Siła Woli"), "SW")
        self.assertEqual(game_data.characteristic_to_code("Wt"), "Wt")
        self.assertEqual(game_data.characteristic_to_code("nieznana"), "")


class TalentMatchingRegressionTests(unittest.TestCase):
    """Faza 4B3: dopasowanie talentów ze specjalizacją do wpisów bazy."""

    def test_specialised_talent_matches_database_by_base_name(self):
        import game_data
        database = game_data.load_talents()
        index = DataManager._talent_base_index(database)
        # Talenty z konkretną specjalizacją dopasowują się do wpisu bazowego.
        for pdf_name, expected_attr in [
            ("Etykieta (Uczeni)", "Ogd"),
            ("Wytwórca (Aptekarz)", "Zr"),
            ("Odporny na (Trucizny)", "Wt"),
            ("Znawca (Aptekarstwo)", "Int"),
        ]:
            entry = DataManager._match_talent_db_entry(pdf_name, database, index)
            self.assertIsNotNone(entry, pdf_name)
            self.assertEqual(entry.get("max", {}).get("attr"), expected_attr, pdf_name)
            self.assertNotEqual(entry.get("source"), "Karta PDF", pdf_name)

    def test_enriched_specialised_talent_is_not_marked_custom(self):
        manager = DataManager()
        manager.talents = {
            "Etykieta (Uczeni)": {"advances": "1"},
            "Wytwórca (Aptekarz)": {"advances": "1"},
        }
        manager._enrich_talents()
        for name in ("Etykieta (Uczeni)", "Wytwórca (Aptekarz)"):
            self.assertFalse(manager.talents[name]["is_custom"], name)
            self.assertEqual(manager.talents[name]["source"], "Podręcznik podstawowy", name)
            self.assertIsNotNone(manager.talents[name]["max"], name)


if __name__ == "__main__":
    unittest.main()

