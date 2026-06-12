"""
Warhammer Fantasy Roleplay 4ed - Symulator Karty Postaci
Program do zarządzania postacią RPG z obsługą cech, umiejętności i talentów.
"""

import customtkinter as ctk
import openpyxl
import json
import os
import re
import unicodedata
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import winsound  # type: ignore
except ImportError:  # pragma: no cover - tylko poza Windows
    winsound = None

from pdf_character_io import extract_pdf_character_data, write_pdf_character_data
import game_data


def play_blocked_sound() -> None:
    """Odtwarza krótki dźwięk systemowy zamiast okna błędu przy niedozwolonej akcji."""
    if winsound is not None:
        try:
            winsound.MessageBeep(winsound.MB_ICONHAND)
            return
        except Exception:
            pass
    try:
        tk._default_root.bell()  # type: ignore[attr-defined]
    except Exception:
        pass


# ============================================================================
# CONSTANTS
# ============================================================================

COST_TABLE = {
    5: (25, 10),
    10: (30, 15),
    15: (40, 20),
    20: (50, 30),
    25: (70, 40),
    30: (90, 60),
    35: (120, 80),
    40: (150, 110),
    45: (190, 140),
    50: (230, 180),
    55: (280, 220),
    60: (330, 270),
    65: (390, 320),
    70: (450, 380),
    float("inf"): (520, 440),
}

ATTRIBUTES = ["WW", "US", "S", "Wt", "I", "Zw", "Zr", "Int", "SW", "Ogd"]
# Koszt jednego rozwinięcia talentu (WFRP 4ed: 100 PD za każde wykupienie).
TALENT_COST_PER_ADVANCE = 100
EXCEL_CHAR_COL_START = 2  # Kolumna B
EXCEL_ATTR_NAME_ROW = 11
EXCEL_ATTR_INITIAL_ROW = 12
EXCEL_ATTR_ADVANCED_ROW = 13
EXCEL_ATTR_CURRENT_ROW = 14
EXCEL_SKILL_ROWS = range(18, 31)
EXCEL_SKILL_BLOCKS = 3
EXCEL_SKILLS_PER_BLOCK = len(EXCEL_SKILL_ROWS)
EXCEL_EXP_COL = "P"  # Aktualne PD
EXCEL_SPENT_COL = "Q"  # Wydane PD
EXCEL_TOTAL_COL = "R"  # Suma PD

# Kolory do customtkinter
COLOR_BG = "#1a1a1a"
COLOR_FG_LIGHT = "#e0e0e0"
COLOR_ACCENT = "#0084d1"
COLOR_SUCCESS = "#26b552"
COLOR_ERROR = "#d1214e"
COLOR_WARNING = "#ff9500"
COLOR_SURFACE = "#22252b"
COLOR_SURFACE_ALT = "#2c3138"
COLOR_SURFACE_SOFT = "#343a43"
COLOR_TEXT_MUTED = "#aab3c0"
COLOR_BORDER = "#404856"
COLOR_HIGHLIGHT = "#f0c44c"
# Lekkie podświetlenie wierszy rozwijalnych w profesji (cechy/umiejętności/talenty)
COLOR_DEVELOPABLE_BG = "#22323d"
COLOR_DEVELOPABLE_BORDER = "#2f6f8f"


class Tooltip:
    """Lekki dymek podpowiedzi pokazywany po najechaniu kursorem na widget."""

    def __init__(self, widget, text: str = "", delay: int = 350, wraplength: int = 480):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.wraplength = wraplength
        self._after_id = None
        self._tip = None
        self._label = None
        self.attach(widget)

    def attach(self, widget) -> None:
        """Podpina dymek pod dodatkowy widget (np. etykietę w wierszu)."""
        widget.bind("<Enter>", self._schedule, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")

    def set_text(self, text: str) -> None:
        self.text = text or ""
        if self._tip is not None and self._label is not None:
            self._label.configure(text=self.text)

    def _schedule(self, _event=None) -> None:
        self._cancel()
        if not self.text:
            return
        self._after_id = self.widget.after(self.delay, self._show)

    def _cancel(self) -> None:
        if self._after_id is not None:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def _show(self) -> None:
        if self._tip is not None or not self.text:
            return
        x = self.widget.winfo_pointerx() + 16
        y = self.widget.winfo_pointery() + 18
        self._tip = tk.Toplevel(self.widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        self._tip.configure(bg=COLOR_DEVELOPABLE_BORDER)
        self._label = tk.Label(
            self._tip,
            text=self.text,
            justify="left",
            bg=COLOR_SURFACE_SOFT,
            fg=COLOR_FG_LIGHT,
            wraplength=self.wraplength,
            padx=10,
            pady=8,
            font=("Segoe UI", 10),
            bd=0,
        )
        self._label.pack(padx=1, pady=1)

    def _hide(self, _event=None) -> None:
        self._cancel()
        if self._tip is not None:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None
            self._label = None


class AutocompletePopup:
    """Lista podpowiedzi pod polem filtra, pokazywana po wpisaniu znaków."""

    def __init__(self, entry, var, candidates_provider, on_select, max_items: int = 10):
        self.entry = entry
        self.var = var
        self.candidates_provider = candidates_provider  # callable -> list[str]
        self.on_select = on_select  # callable(value)
        self.max_items = max_items
        self._popup = None
        self._listbox = None
        entry.bind("<KeyRelease>", self._on_key, add="+")
        entry.bind("<FocusOut>", lambda _e: entry.after(160, self._hide), add="+")
        entry.bind("<Escape>", lambda _e: self._hide(), add="+")

    def _on_key(self, event) -> None:
        if event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            if event.keysym == "Return":
                self._choose_first()
            return
        text = self.var.get().strip()
        if len(text) < 1:
            self._hide()
            return
        norm = normalize_search_text(text)
        matches = [
            c for c in self.candidates_provider()
            if norm in normalize_search_text(c) and normalize_search_text(c) != norm
        ]
        # Zachowaj kolejność, usuń duplikaty
        seen = set()
        unique = []
        for m in matches:
            if m not in seen:
                seen.add(m)
                unique.append(m)
        unique = unique[: self.max_items]
        if not unique:
            self._hide()
            return
        self._show(unique)

    def _choose_first(self) -> None:
        if self._listbox is not None and self._listbox.size() > 0:
            value = self._listbox.get(0)
            self._select(value)

    def _show(self, matches) -> None:
        x = self.entry.winfo_rootx()
        y = self.entry.winfo_rooty() + self.entry.winfo_height()
        width = self.entry.winfo_width()
        if self._popup is None:
            self._popup = tk.Toplevel(self.entry)
            self._popup.wm_overrideredirect(True)
            self._popup.configure(bg=COLOR_DEVELOPABLE_BORDER)
            self._listbox = tk.Listbox(
                self._popup,
                bg=COLOR_SURFACE_SOFT,
                fg=COLOR_FG_LIGHT,
                selectbackground=COLOR_ACCENT,
                selectforeground="#ffffff",
                font=("Segoe UI", 10),
                bd=0,
                highlightthickness=0,
                activestyle="none",
            )
            self._listbox.pack(fill="both", expand=True, padx=1, pady=1)
            self._listbox.bind("<ButtonRelease-1>", self._on_click)
        self._listbox.delete(0, "end")
        for m in matches:
            self._listbox.insert("end", m)
        height = min(len(matches), self.max_items) * 20 + 4
        self._popup.wm_geometry(f"{max(width, 200)}x{height}+{x}+{y}")
        self._popup.deiconify()
        self._popup.lift()

    def _on_click(self, _event=None) -> None:
        if self._listbox is None:
            return
        selection = self._listbox.curselection()
        if not selection:
            return
        self._select(self._listbox.get(selection[0]))

    def _select(self, value: str) -> None:
        self.var.set(value)
        self._hide()
        self.entry.focus_set()
        self.on_select(value)

    def _hide(self, _event=None) -> None:
        if self._popup is not None:
            try:
                self._popup.destroy()
            except Exception:
                pass
            self._popup = None
            self._listbox = None

FONT_DISPLAY = ("Bahnschrift SemiBold", 26)
FONT_TITLE = ("Bahnschrift SemiBold", 18)
FONT_SECTION = ("Bahnschrift SemiBold", 15)
FONT_BODY = ("Segoe UI", 13)
FONT_BODY_BOLD = ("Segoe UI Semibold", 13)
FONT_SMALL = ("Segoe UI", 11)

FILE_TYPE_EXCEL = "excel"
FILE_TYPE_PDF = "pdf"

ATTRIBUTE_DETAILS = {
    "WW": "Walka Wręcz",
    "US": "Umiejętności Strzeleckie",
    "S": "Siła",
    "Wt": "Wytrzymałość",
    "I": "Inicjatywa",
    "Zw": "Zwinność",
    "Zr": "Zręczność",
    "Int": "Inteligencja",
    "SW": "Siła Woli",
    "Ogd": "Ogłada",
}

BASIC_SKILL_PATTERNS = [
    "Atletyka",
    "Broń Biała",
    "Charyzma",
    "Dowodzenie",
    "Hazard",
    "Intuicja",
    "Jeździectwo",
    "Mocna głowa",
    "Nawigacja",
    "Odporność",
    "Opanowanie",
    "Oswajanie",
    "Percepcja",
    "Plotkowanie",
    "Powożenie",
    "Przekupstwo",
    "Skradanie",
    "Sztuka",
    "Sztuka Przetrwania",
    "Targowanie",
    "Unik",
    "Wioślarstwo",
    "Wspinaczka",
    "Występy",
    "Zastraszanie",
]

ATTRIBUTE_FILTER_ALL = "Wszystkie cechy"
SKILL_CATEGORY_BASIC = "Podstawowa"
SKILL_CATEGORY_ADVANCED = "Zaawansowana / grupowa"
SKILL_CATEGORY_SHORT = {
    SKILL_CATEGORY_BASIC: "Podst",
    SKILL_CATEGORY_ADVANCED: "Zw/grup",
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def calculate_advancement_cost(
    advancement_type: str,
    current_advancements: int,
    desired_advancements: int,
    out_of_profession: bool = False,
    gm_approved: bool = False,
) -> int:
    """Oblicza całkowity koszt PD dla rozwinięć.

    Rozwój SPOZA profesji podwaja koszt, chyba że MG wyraził zgodę (gm_approved).
    """
    total_cost = 0
    remaining = desired_advancements
    current_threshold = current_advancements

    while remaining > 0:
        for threshold in sorted(COST_TABLE.keys()):
            if current_threshold < threshold:
                char_cost, skill_cost = COST_TABLE[threshold]
                to_threshold = min(remaining, threshold - current_threshold)
                
                if advancement_type == "cecha":
                    total_cost += char_cost * to_threshold
                elif advancement_type == "umiejetnosc":
                    total_cost += skill_cost * to_threshold
                elif advancement_type == "talent":
                    # Koszt talentu: 100 * poziom
                    total_cost = desired_advancements * 100
                    remaining = 0
                    break

                remaining -= to_threshold
                current_threshold += to_threshold
                if remaining == 0:
                    break

    if out_of_profession and not gm_approved:
        total_cost *= 2
    return total_cost


def calculate_talent_cost(
    amount: int, out_of_profession: bool = False, gm_approved: bool = False
) -> int:
    """Koszt PD za 'amount' wykupień talentu (100 PD każde).

    Rozwój spoza profesji podwaja koszt, chyba że MG wyraził zgodę.
    """
    cost = max(0, amount) * TALENT_COST_PER_ADVANCE
    if out_of_profession and not gm_approved:
        cost *= 2
    return cost


def normalize_search_text(text: str) -> str:
    """Normalizuje tekst do wyszukiwania niezależnie od wielkości liter i polskich znaków."""
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip().casefold()


def normalize_skill_name(text: str) -> str:
    """Upraszcza nazwę umiejętności do porównań antyduplikacyjnych."""
    normalized = normalize_search_text(text)
    normalized = re.sub(r"\s*\(\s*", "(", normalized)
    normalized = re.sub(r"\s*\)\s*", ")", normalized)
    return normalized


def get_skill_category(skill_name: str) -> str:
    """Klasyfikuje umiejętność jako podstawową albo zaawansowaną/grupową."""
    normalized = normalize_search_text(skill_name)
    for pattern in BASIC_SKILL_PATTERNS:
        if normalized.startswith(normalize_search_text(pattern)):
            return SKILL_CATEGORY_BASIC
    return SKILL_CATEGORY_ADVANCED


def get_skill_category_short(skill_name: str) -> str:
    """Zwraca skróconą etykietę kategorii umiejętności."""
    return SKILL_CATEGORY_SHORT[get_skill_category(skill_name)]


def get_attribute_filter_label(attribute_code: str) -> str:
    """Buduje czytelną etykietę filtra cechy."""
    return f"{ATTRIBUTE_DETAILS[attribute_code]} ({attribute_code})"


def get_attribute_code_from_filter_label(filter_label: str) -> Optional[str]:
    """Zwraca kod cechy na podstawie etykiety z comboboxa."""
    for attribute_code in ATTRIBUTES:
        if filter_label == get_attribute_filter_label(attribute_code):
            return attribute_code
    return None


# ============================================================================
# DATA MANAGEMENT
# ============================================================================

class DataManager:
    """Zarządza ładowaniem i zapisywaniem danych postaci."""

    def __init__(self):
        self.file_path: Optional[str] = None
        self.source_type: str = FILE_TYPE_EXCEL
        self.attributes: Dict = {}
        self.skills: Dict = {}
        self.talents: Dict = {}
        self.experience: Dict = {"available": 0, "spent": 0, "total": 0}
        self.character_name: str = "Nowa Postać"
        self.pdf_mapping: Dict = {}
        # Profesja / klasa / ścieżka kariery
        self.character_class: str = ""
        self.character_species: str = ""
        self.current_career: str = ""
        self.current_career_level: int = 1
        self.career_path: List[Dict] = []
        self.profession_raw: Dict = {}

    def load_from_excel(self, file_path: str) -> bool:
        """Ładuje dane z pliku Excel. Zwraca True jeśli sukces."""
        try:
            self.file_path = file_path
            self.source_type = FILE_TYPE_EXCEL
            self.pdf_mapping = {}
            self.character_class = ""
            self.character_species = ""
            self.current_career = ""
            self.current_career_level = 1
            self.career_path = []
            self.profession_raw = {}
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Odczyt charakteru postaci
            self.character_name = ws.cell(5, 2).value or "Brak Imienia"

            # Odczyt cech
            self.attributes = {}
            for col_idx, attr_name in enumerate(ATTRIBUTES, start=EXCEL_CHAR_COL_START):
                cell_name = ws.cell(EXCEL_ATTR_NAME_ROW, col_idx).value
                if not cell_name:
                    continue

                initial = ws.cell(EXCEL_ATTR_INITIAL_ROW, col_idx).value or 0
                advanced = ws.cell(EXCEL_ATTR_ADVANCED_ROW, col_idx).value or 0
                current = ws.cell(EXCEL_ATTR_CURRENT_ROW, col_idx).value or (initial + advanced)

                self.attributes[attr_name] = {
                    "initial": int(initial),
                    "advanced": int(advanced),
                    "current": int(current),
                    "base_advanced": int(advanced),  # Do resetu
                    "is_new": False,
                    "profession_available": False,
                }

            # Odczyt umiejętności
            self.skills = {}
            for block in range(EXCEL_SKILL_BLOCKS):
                col_offset = block * 5
                for row in EXCEL_SKILL_ROWS:
                    skill_name = ws.cell(row, 1 + col_offset).value
                    if not skill_name:
                        continue

                    attribute = ws.cell(row, 2 + col_offset).value or ""
                    initial = ws.cell(row, 3 + col_offset).value or 0
                    advanced = ws.cell(row, 4 + col_offset).value or 0
                    current = ws.cell(row, 5 + col_offset).value or (initial + advanced)

                    self.skills[skill_name] = {
                        "attribute": str(attribute).strip("()"),
                        "initial": int(initial),
                        "advanced": int(advanced),
                        "current": int(current),
                        "base_advanced": int(advanced),
                        "is_new": False,
                        "profession_available": False,
                    }

            # Odczyt doświadczenia
            self.experience["available"] = ws[f"{EXCEL_EXP_COL}11"].value or 0
            self.experience["spent"] = ws[f"{EXCEL_SPENT_COL}11"].value or 0
            self.experience["total"] = (
                self.experience["available"] + self.experience["spent"]
            )

            wb.close()
            return True

        except Exception as e:
            print(f"Błąd przy ładowaniu pliku: {e}")
            return False

    def load_from_pdf(self, file_path: str) -> bool:
        """Ładuje dane z formularza PDF. Zwraca True jeśli sukces."""
        try:
            payload = extract_pdf_character_data(file_path)
            self.file_path = file_path
            self.source_type = FILE_TYPE_PDF
            self.character_name = payload["character_name"]
            self.attributes = payload["attributes"]
            self.skills = payload["skills"]
            self.talents = payload["talents"]
            self.experience = payload["experience"]
            self.pdf_mapping = payload["pdf_mapping"]
            self._enrich_talents()
            self._load_profession(payload.get("profession_info", {}))
            return True
        except Exception as e:
            print(f"Błąd przy ładowaniu PDF: {e}")
            return False

    def save_to_excel(self, file_path: str = None) -> bool:
        """Zapisuje dane do Excel. Zwraca True jeśli sukces."""
        if not file_path and not self.file_path:
            return False

        file_path = file_path or self.file_path

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            if len(self.skills) > EXCEL_SKILL_BLOCKS * EXCEL_SKILLS_PER_BLOCK:
                raise ValueError(
                    "Za dużo umiejętności, aby zapisać je do aktualnego układu Excela."
                )

            ws.cell(5, 2, self.character_name)

            # Zapis cech
            for col_idx, attr_name in enumerate(ATTRIBUTES, start=EXCEL_CHAR_COL_START):
                if attr_name in self.attributes:
                    attr_data = self.attributes[attr_name]
                    ws.cell(EXCEL_ATTR_INITIAL_ROW, col_idx, attr_data["initial"])
                    ws.cell(EXCEL_ATTR_ADVANCED_ROW, col_idx, attr_data["advanced"])
                    ws.cell(
                        EXCEL_ATTR_CURRENT_ROW,
                        col_idx,
                        attr_data["initial"] + attr_data["advanced"],
                    )

            # Wyczyść wszystkie sloty umiejętności przed zapisem
            for block in range(EXCEL_SKILL_BLOCKS):
                col_offset = block * 5
                for row_idx in EXCEL_SKILL_ROWS:
                    for column_offset in range(5):
                        ws.cell(row_idx, 1 + col_offset + column_offset, None)

            # Zapis umiejętności
            for skill_index, (skill_name, skill_data) in enumerate(self.skills.items()):
                block = skill_index // EXCEL_SKILLS_PER_BLOCK
                row_offset = skill_index % EXCEL_SKILLS_PER_BLOCK
                col_offset = block * 5
                row_idx = EXCEL_SKILL_ROWS.start + row_offset

                ws.cell(row_idx, 1 + col_offset, skill_name)
                ws.cell(row_idx, 2 + col_offset, f"({skill_data['attribute']})")
                ws.cell(row_idx, 3 + col_offset, skill_data["initial"])
                ws.cell(row_idx, 4 + col_offset, skill_data["advanced"])
                ws.cell(
                    row_idx,
                    5 + col_offset,
                    skill_data["initial"] + skill_data["advanced"],
                )

            # Zapis doświadczenia
            ws[f"{EXCEL_EXP_COL}11"] = self.experience["available"]
            ws[f"{EXCEL_SPENT_COL}11"] = self.experience["spent"]
            ws[f"{EXCEL_TOTAL_COL}11"] = self.experience["total"]

            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            print(f"Błąd przy zapisie pliku: {e}")
            return False

    def save_to_pdf(self, file_path: str = None) -> bool:
        """Zapisuje dane do formularza PDF. Zwraca True jeśli sukces."""
        if not file_path and not self.file_path:
            return False

        file_path = file_path or self.file_path

        try:
            if not self.pdf_mapping:
                return False

            write_pdf_character_data(
                self.file_path,
                file_path,
                {
                    "character_name": self.character_name,
                    "attributes": self.attributes,
                    "skills": self.skills,
                    "talents": self.talents,
                    "experience": self.experience,
                    "profession": self.profession_payload(),
                },
                self.pdf_mapping,
            )
            return True
        except Exception as e:
            print(f"Błąd przy zapisie PDF: {e}")
            return False

    def add_skill(
        self,
        skill_name: str,
        attribute: str,
        initial: int = 0,
        advanced: int = 0,
        is_new: bool = False,
    ) -> bool:
        """Dodaje nową umiejętność."""
        if skill_name in self.skills:
            return False

        self.skills[skill_name] = {
            "attribute": attribute,
            "initial": initial,
            "advanced": advanced,
            "current": initial + advanced,
            "base_advanced": advanced,
            "is_new": is_new,
            "profession_available": False,
        }
        return True

    # ----- Talenty ---------------------------------------------------------
    def _enrich_talents(self) -> None:
        """Uzupełnia talenty wczytane z PDF o dane z bazy (Max, opis, źródło)."""
        database = game_data.load_talents()
        base_index = self._talent_base_index(database)
        enriched: Dict = {}
        for name, raw in self.talents.items():
            advances = self._parse_talent_advances(raw.get("advances"))
            db_entry = self._match_talent_db_entry(name, database, base_index)
            description = raw.get("description") or (
                db_entry.get("description") if db_entry else ""
            )
            enriched[name] = {
                "advances": advances,
                "base_advances": advances,
                "description": description,
                "max": (db_entry or {}).get("max"),
                "tests": (db_entry or {}).get("tests"),
                "source": (db_entry or {}).get("source", "Karta PDF"),
                "is_new": False,
                "is_custom": db_entry is None,
                "profession_available": False,
            }
        self.talents = enriched

    @staticmethod
    def _talent_base_index(database: Dict) -> Dict[str, str]:
        """Indeks {nazwa_bazowa_bez_specjalizacji -> klucz_bazy} do dopasowań."""
        index: Dict[str, str] = {}
        for key in database:
            base = game_data._normalize(str(key).split("(")[0])
            if base:
                index.setdefault(base, key)
        return index

    @staticmethod
    def _match_talent_db_entry(name, database: Dict, base_index: Dict[str, str]):
        """Dopasowuje talent z PDF do wpisu bazy (dokładnie lub po nazwie bazowej)."""
        entry = database.get(name)
        if entry is not None:
            return entry
        base = game_data._normalize(str(name).split("(")[0])
        matched_key = base_index.get(base)
        if matched_key:
            return database.get(matched_key)
        return None

    @staticmethod
    def _parse_talent_advances(value) -> int:
        """Zamienia wartość 'times taken' z PDF na liczbę (puste -> 1)."""
        if value is None:
            return 1
        text = str(value).strip()
        if not text:
            return 1
        match = re.search(r"\d+", text)
        return int(match.group()) if match else 1

    def talent_max_advances(self, name: str) -> Optional[int]:
        """Zwraca twardy limit wykupień talentu (None = brak limitu liczbowego)."""
        talent = self.talents.get(name)
        info = talent.get("max") if talent else None
        if not info:
            return None
        kind = info.get("type")
        if kind == "fixed":
            return info.get("value")
        if kind == "characteristic":
            attr = info.get("attr")
            attr_data = self.attributes.get(attr)
            if not attr_data:
                return None
            return max(1, game_data.attribute_bonus(attr_data.get("current", 0)))
        return None

    def add_talent(
        self,
        name: str,
        advances: int = 1,
        description: str = "",
        max_info: Optional[Dict] = None,
        tests: Optional[str] = None,
        source: str = "Lista",
        is_custom: bool = False,
        is_new: bool = True,
    ) -> bool:
        """Dodaje talent do postaci. Zwraca False, jeśli już istnieje."""
        if name in self.talents:
            return False
        self.talents[name] = {
            "advances": advances,
            "base_advances": 0 if is_new else advances,
            "description": description,
            "max": max_info,
            "tests": tests,
            "source": source,
            "is_new": is_new,
            "is_custom": is_custom,
            "profession_available": False,
        }
        return True

    # ----- Profesja / klasa / ścieżka kariery -----------------------------
    def _load_profession(self, info: Dict) -> None:
        """Wczytuje dane profesji z pól PDF i buduje ścieżkę kariery."""
        self.profession_raw = dict(info or {})
        self.character_class = (info or {}).get("class", "") or ""
        self.character_species = (info or {}).get("species", "") or ""

        profession_field = (info or {}).get("profession", "") or ""
        path_text = (info or {}).get("path_text", "") or ""
        level_text = (info or {}).get("level_text", "") or ""

        steps = game_data.resolve_career_path(path_text)
        # Jeśli ścieżka pusta, ale jest profesja, zbuduj jednoelementową ścieżkę
        if not steps and profession_field:
            steps = game_data.resolve_career_path(profession_field)

        parsed_level = game_data.parse_career_level(level_text)
        self.career_path = self._build_career_path(steps, profession_field, parsed_level)

        if self.career_path:
            last = self.career_path[-1]
            self.current_career = last.get("profession") or last.get("title") or profession_field
            self.current_career_level = last.get("level") or parsed_level or 1
        else:
            self.current_career = profession_field
            self.current_career_level = parsed_level or 1

        # Uzupełnij klasę z danych gry, jeśli pole PDF puste lub niespójne
        resolved_class = game_data.class_of_career(self.current_career)
        if resolved_class:
            self.character_class = resolved_class

    def _build_career_path(
        self, steps: List[Dict], current_profession: str, current_level: Optional[int]
    ) -> List[Dict]:
        """Składa listę kroków kariery z oznaczeniem poziomu i kompletowania."""
        path: List[Dict] = []
        for index, step in enumerate(steps):
            is_last = index == len(steps) - 1
            level = step.get("level") or 1
            if is_last and current_level:
                level = current_level
            entry = {
                "title": step.get("title", ""),
                "profession": step.get("profession"),
                "level": level,
                "resolved": step.get("resolved", False),
                "completed": not is_last,
            }
            path.append(entry)
        return path

    def profession_payload(self) -> Dict:
        """Buduje słownik pól profesji do zapisu w PDF."""
        path_titles = [
            (step.get("title") or step.get("profession") or "")
            for step in self.career_path
            if (step.get("title") or step.get("profession"))
        ]
        level_text = self.profession_raw.get("level_text", "")
        if self.current_career:
            level_text = f"{self.current_career} ({self.current_career_level})"
        return {
            "class": self.character_class,
            "profession": self.current_career,
            "level_text": level_text,
            "path_text": ", ".join(path_titles),
            "species": self.character_species,
        }

    def set_current_career(self, profession: str, level: int) -> None:
        """Ustawia/poprawia bieżącą profesję i poziom (bez kosztu)."""
        self.current_career = profession
        self.current_career_level = max(1, min(4, int(level)))
        resolved_class = game_data.class_of_career(profession)
        if resolved_class:
            self.character_class = resolved_class
        if not self.career_path:
            self.career_path = [
                {
                    "title": profession,
                    "profession": profession if game_data.get_profession(profession) else None,
                    "level": self.current_career_level,
                    "resolved": bool(game_data.get_profession(profession)),
                    "completed": False,
                }
            ]
        else:
            last = self.career_path[-1]
            last["title"] = profession
            last["profession"] = profession if game_data.get_profession(profession) else None
            last["level"] = self.current_career_level
            last["resolved"] = bool(game_data.get_profession(profession))
            last["completed"] = False

    def advance_to_career(self, profession: str, level: int) -> None:
        """Awansuje do nowej profesji: oznacza obecną jako ukończoną i dopisuje krok."""
        level = max(1, min(4, int(level)))
        if self.career_path:
            self.career_path[-1]["completed"] = True
        resolved = bool(game_data.get_profession(profession))
        self.career_path.append(
            {
                "title": profession,
                "profession": profession if resolved else None,
                "level": level,
                "resolved": resolved,
                "completed": False,
            }
        )
        self.current_career = profession
        self.current_career_level = level
        resolved_class = game_data.class_of_career(profession)
        if resolved_class:
            self.character_class = resolved_class

    def current_career_completion(self) -> Dict:
        """Zwraca status kompletowania bieżącej profesji."""
        profession = self.current_career
        if not game_data.get_profession(profession):
            return {
                "completed": False,
                "skills_ok": False,
                "talents_ok": False,
                "characteristics_ok": False,
                "skills_done": 0,
                "talents_done": 0,
                "characteristics_pending": True,
                "unknown_profession": True,
            }
        result = game_data.is_career_completed(
            profession,
            self.current_career_level,
            self.skills,
            self.talents,
            self.attributes,
        )
        result["unknown_profession"] = False
        return result

    def reset_character(self) -> None:
        """Resetuje postać do wartości z pliku."""
        for attr in self.attributes.values():
            attr["advanced"] = attr["base_advanced"]

        for skill in self.skills.values():
            skill["advanced"] = skill["base_advanced"]

    def create_new_character(self, name: str = "Nowa Postać") -> None:
        """Tworzy nową postać z wartościami domyślnymi."""
        self.character_name = name
        self.file_path = None
        self.source_type = FILE_TYPE_EXCEL
        self.pdf_mapping = {}
        self.attributes = {
            attr: {
                "initial": 30,
                "advanced": 0,
                "current": 30,
                "base_advanced": 0,
                "is_new": False,
                "profession_available": False,
            }
            for attr in ATTRIBUTES
        }
        self.skills = {}
        self.experience = {"available": 0, "spent": 0, "total": 0}
        self.talents = {}
        self.character_class = ""
        self.character_species = ""
        self.current_career = ""
        self.current_career_level = 1
        self.career_path = []
        self.profession_raw = {}



# ============================================================================
# HISTORY MANAGEMENT
# ============================================================================

class HistoryManager:
    """Zarządza historią działań per-postać w formacie JSON.

    Struktura pliku:
        {
            "characters": { "<imię>": { "career_path": [...], "changes": [...] } },
            "global_log": [ {timestamp, action, details, character}, ... ]
        }
    Stary format (płaska lista wpisów) jest automatycznie migrowany do
    ``global_log`` przy pierwszym wczytaniu.
    """

    def __init__(self, json_file: str = "history.json"):
        self.json_file = json_file
        self.data: Dict = {"characters": {}, "global_log": []}
        self.current_character: Optional[str] = None
        self.load_history()

    @property
    def history(self) -> List[Dict]:
        """Zgodność wstecz: globalny dziennik zdarzeń."""
        return self.data.get("global_log", [])

    @history.setter
    def history(self, value: List[Dict]) -> None:
        """Zgodność wstecz: ustawienie globalnego dziennika."""
        self.data["global_log"] = list(value) if value else []

    def load_history(self) -> None:
        """Ładuje historię z pliku JSON (z migracją starego formatu)."""
        try:
            if os.path.exists(self.json_file):
                with open(self.json_file, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                self.data = self._migrate(loaded)
            else:
                self.data = {"characters": {}, "global_log": []}
        except Exception as e:
            print(f"Błąd przy ładowaniu historii: {e}")
            self.data = {"characters": {}, "global_log": []}

    @staticmethod
    def _migrate(loaded) -> Dict:
        """Konwertuje wczytane dane do formatu per-postać."""
        if isinstance(loaded, list):
            return {"characters": {}, "global_log": loaded}
        if isinstance(loaded, dict):
            characters = loaded.get("characters", {})
            global_log = loaded.get("global_log", [])
            if not isinstance(characters, dict):
                characters = {}
            if not isinstance(global_log, list):
                global_log = []
            # Uzupełnij brakujące pola w postaciach
            for name, record in list(characters.items()):
                if not isinstance(record, dict):
                    characters[name] = {"career_path": [], "changes": []}
                    continue
                record.setdefault("career_path", [])
                record.setdefault("changes", [])
            return {"characters": characters, "global_log": global_log}
        return {"characters": {}, "global_log": []}

    def save_history(self) -> None:
        """Zapisuje historię do pliku JSON."""
        try:
            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Błąd przy zapisie historii: {e}")

    def ensure_character(self, name: str) -> Dict:
        """Zwraca (tworząc w razie potrzeby) rekord postaci."""
        characters = self.data.setdefault("characters", {})
        return characters.setdefault(name, {"career_path": [], "changes": []})

    def set_current_character(self, name: Optional[str]) -> None:
        """Ustawia bieżącą postać dla kolejnych wpisów."""
        self.current_character = name or None
        if self.current_character:
            self.ensure_character(self.current_character)
            self.save_history()

    def add_entry(self, action: str, details: str = "") -> None:
        """Dodaje wpis do globalnego dziennika oraz do bieżącej postaci."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "action": action,
            "details": details,
            "character": self.current_character,
        }
        self.data.setdefault("global_log", []).append(entry)
        if self.current_character:
            record = self.ensure_character(self.current_character)
            record.setdefault("changes", []).append(entry)
        self.save_history()

    def get_career_path(self, name: str) -> List[Dict]:
        """Zwraca zapisaną ścieżkę kariery postaci."""
        record = self.data.get("characters", {}).get(name)
        return list(record.get("career_path", [])) if record else []

    def set_career_path(self, name: str, career_path: List[Dict]) -> None:
        """Zapisuje ścieżkę kariery dla postaci."""
        record = self.ensure_character(name)
        record["career_path"] = list(career_path)
        self.save_history()

    def get_history_text(self) -> str:
        """Zwraca historię jako tekst (bieżąca postać albo dziennik globalny)."""
        if self.current_character:
            record = self.data.get("characters", {}).get(self.current_character, {})
            entries = record.get("changes", [])
        else:
            entries = self.data.get("global_log", [])

        text = ""
        for entry in entries[-50:]:  # Ostatnie 50 wpisów
            timestamp = entry.get("timestamp", "")
            action = entry.get("action", "")
            details = entry.get("details", "")
            text += f"[{timestamp}] {action}"
            if details:
                text += f" - {details}"
            text += "\n"
        return text


# ============================================================================
# GUI - MAIN APPLICATION
# ============================================================================

class CharacterSheetApp(ctk.CTk):
    """Główna aplikacja GUI."""

    def __init__(self):
        super().__init__()

        self.title("Warhammer Fantasy Roleplay 4ed - Karta Postaci")
        self.geometry("1400x900")
        self.minsize(1180, 760)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        self.configure(fg_color=COLOR_BG)

        self.data_manager = DataManager()
        self.history_manager = HistoryManager()

        self.pending_changes: Dict = self._create_empty_pending_changes()
        
        # Cache widgetów do szybkiego updatu bez destroy/recreate
        self.attribute_rows: Dict = {}  # {attr_name: {"row": Frame, "labels": [...], "buttons": [...]}}
        self.skill_rows: Dict = {}      # {skill_name: {"row": Frame, "labels": [...], "buttons": [...]}}
        self.talent_rows: Dict = {}     # {talent_name: {"row": Frame, "labels": [...], "buttons": [...]}}
        self.initialized_attrs = False
        self.initialized_skills = False
        self.initialized_talents = False
        self.skill_filter_var = tk.StringVar(value="")
        self.skill_summary_var = tk.StringVar(value="Wyświetlane umiejętności: 0 / 0")
        self.cost_filter_var = tk.StringVar(value="")
        self.skill_attribute_filter_var = tk.StringVar(value=ATTRIBUTE_FILTER_ALL)
        self.skill_profession_filter_var = tk.BooleanVar(value=False)
        self.talent_filter_var = tk.StringVar(value="")
        self.talent_summary_var = tk.StringVar(value="Wyświetlane talenty: 0 / 0")
        self.talent_out_of_profession_var = tk.BooleanVar(value=False)
        self.talent_gm_approval_var = tk.BooleanVar(value=False)
        self.talent_profession_filter_var = tk.BooleanVar(value=False)
        self.attr_profession_filter_var = tk.BooleanVar(value=False)
        self.dev_gm_approval_var = tk.BooleanVar(value=False)
        self._developable = {
            "characteristics": set(), "skills": set(), "talents": set(), "resolved": False,
        }
        self.profession_class_var = tk.StringVar(value="")
        self.profession_career_var = tk.StringVar(value="")
        self.profession_level_var = tk.StringVar(value="1")
        self.profession_summary_var = tk.StringVar(value="Brak wczytanej profesji.")
        self.profession_completion_var = tk.StringVar(value="")
        self.profession_status_var = tk.StringVar(value="")
        self.resize_refresh_after_id: Optional[str] = None
        self.last_resize_width: Optional[int] = None
        self.cost_layout_mode: Optional[bool] = None
        self.cost_options_dirty = True
        self.skills_group_frames: Dict[str, Dict] = {}

        self.setup_ui()
        self.bind("<Configure>", self.on_window_resize)

    def _create_empty_pending_changes(self) -> Dict:
        """Tworzy nowy kontener oczekujących zmian."""
        return {
            "attribute_changes": {},
            "skill_changes": {},
            "talent_changes": {},
            "new_skills": set(),
            "new_talents": set(),
            "experience_delta": 0,
        }

    def has_pending_changes(self) -> bool:
        """Sprawdza, czy istnieją niezakończone zmiany."""
        return any(bool(value) for value in self.pending_changes.values())

    def on_window_resize(self, event) -> None:
        """Odświeża zależne od szerokości układy po zmianie rozmiaru okna."""
        if event.widget is not self or not self._is_costs_tab_active():
            return
        if self.last_resize_width is not None and abs(event.width - self.last_resize_width) < 24:
            return
        self.last_resize_width = event.width
        if self.resize_refresh_after_id is not None:
            self.after_cancel(self.resize_refresh_after_id)
        self.resize_refresh_after_id = self.after(220, self._refresh_responsive_sections)

    def _refresh_responsive_sections(self) -> None:
        """Debounced refresh dla układów zależnych od szerokości."""
        self.resize_refresh_after_id = None
        if hasattr(self, "costs_frame") and self._should_refresh_costs_on_resize():
            self.refresh_costs_display()

    def _has_active_skill_filters(self) -> bool:
        """Sprawdza, czy lista umiejętności ma aktywny filtr."""
        selected_attribute = self.skill_attribute_filter_var.get()
        return any(
            (
                normalize_search_text(self.skill_filter_var.get()),
                selected_attribute and selected_attribute != ATTRIBUTE_FILTER_ALL,
                self.skill_profession_filter_var.get(),
            )
        )

    def _is_costs_tab_active(self) -> bool:
        """Sprawdza, czy aktywna jest zakładka kosztów."""
        return hasattr(self, "notebook") and self.notebook.get() == "Koszty Rozwinięć"

    def _should_refresh_costs_on_resize(self) -> bool:
        """Ogranicza kosztowne przebudowy tabeli tylko do istotnych zmian szerokości."""
        if not self._is_costs_tab_active():
            return False
        current_layout_mode = max(self.winfo_width(), self.costs_frame.winfo_width()) >= 1500
        if self.cost_layout_mode is None:
            return True
        return current_layout_mode != self.cost_layout_mode

    def refresh_costs_if_visible(self, force: bool = False) -> None:
        """Odświeża tabelę kosztów tylko wtedy, gdy ma to sens dla użytkownika."""
        if force or self._is_costs_tab_active():
            self.refresh_costs_display()

    def setup_ui(self) -> None:
        """Konfiguruje interfejs użytkownika."""
        self.top_frame = ctk.CTkFrame(
            self,
            fg_color=COLOR_SURFACE,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.top_frame.pack(side="top", fill="x", padx=14, pady=(14, 10))

        header_row = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        header_row.pack(fill="x", padx=18, pady=(16, 10))

        title_block = ctk.CTkFrame(header_row, fg_color="transparent")
        title_block.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(
            title_block,
            text="Warhammer Fantasy Roleplay 4ed",
            font=FONT_DISPLAY,
            text_color=COLOR_FG_LIGHT,
        ).pack(anchor="w")
        ctk.CTkLabel(
            title_block,
            text="Kalkulator rozwinięć, doświadczenia i planowania postaci",
            font=FONT_BODY,
            text_color=COLOR_TEXT_MUTED,
        ).pack(anchor="w", pady=(2, 0))

        button_frame = ctk.CTkFrame(header_row, fg_color="transparent")
        button_frame.pack(side="right", padx=(12, 0))

        ctk.CTkButton(
            button_frame,
            text="Wczytaj",
            command=self.on_load_character,
            width=110,
            height=38,
            fg_color=COLOR_ACCENT,
            hover_color="#0f96ea",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            button_frame,
            text="Nowa",
            command=self.on_new_character,
            width=110,
            height=38,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            button_frame,
            text="Zapisz",
            command=self.on_save_character,
            width=110,
            height=38,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            button_frame,
            text="Zatwierdź zmiany",
            command=self.on_confirm_changes,
            width=150,
            height=38,
            fg_color=COLOR_SUCCESS,
            hover_color="#2dcc62",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            button_frame,
            text="Cofnij zmiany",
            command=self.on_revert_changes,
            width=140,
            height=38,
            fg_color=COLOR_ERROR,
            hover_color="#ee3b6c",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=4)

        stats_row = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        stats_row.pack(fill="x", padx=18, pady=(0, 18))

        self.character_card = self._create_summary_card(
            stats_row, "Postać", "Nowa Postać", COLOR_ACCENT
        )
        self.character_card["frame"].pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.available_card = self._create_summary_card(
            stats_row, "Dostępne PD", "0", COLOR_SUCCESS
        )
        self.available_card["frame"].pack(side="left", fill="x", expand=True, padx=8)

        self.spent_card = self._create_summary_card(
            stats_row, "Wydane PD", "0", COLOR_HIGHLIGHT
        )
        self.spent_card["frame"].pack(side="left", fill="x", expand=True, padx=8)

        self.pending_card = self._create_summary_card(
            stats_row, "Oczekujące zmiany", "Brak", COLOR_WARNING
        )
        self.pending_card["frame"].pack(side="left", fill="x", expand=True, padx=(8, 0))

        self.char_name_label = self.character_card["value_label"]
        self.exp_label = self.available_card["value_label"]

        # Notebook z zakładkami
        self.notebook = ctk.CTkTabview(
            self,
            fg_color=COLOR_SURFACE,
            segmented_button_fg_color=COLOR_SURFACE_SOFT,
            segmented_button_selected_color=COLOR_ACCENT,
            segmented_button_selected_hover_color="#0f96ea",
            segmented_button_unselected_color=COLOR_SURFACE_SOFT,
            segmented_button_unselected_hover_color="#48505e",
            text_color=COLOR_FG_LIGHT,
        )
        self.notebook.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        self.create_character_tab()
        self.create_profession_tab()
        self.create_skills_tab()
        self.create_talents_tab()
        self.create_costs_tab()
        self.create_experience_tab()
        self.create_history_tab()
        self.refresh_header_summary()

    def _create_summary_card(self, parent, title: str, value: str, accent_color: str) -> Dict:
        """Buduje kartę podsumowania w górnym nagłówku."""
        card = ctk.CTkFrame(
            parent,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        ctk.CTkLabel(
            card,
            text=title,
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(12, 2))
        value_label = ctk.CTkLabel(
            card,
            text=value,
            font=FONT_TITLE,
            text_color=COLOR_FG_LIGHT,
        )
        value_label.pack(anchor="w", padx=14)
        badge = ctk.CTkLabel(
            card,
            text="",
            width=18,
            height=18,
            corner_radius=9,
            fg_color=accent_color,
        )
        badge.pack(anchor="w", padx=14, pady=(8, 12))
        return {"frame": card, "value_label": value_label, "badge": badge}

    def _create_tab_shell(self, tab_name: str, title: str, subtitle: str):
        """Tworzy spójny kontener sekcji dla zakładek."""
        tab = self.notebook.add(tab_name)
        wrapper = ctk.CTkFrame(tab, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=14, pady=14)

        hero = ctk.CTkFrame(
            wrapper,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        hero.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(
            hero,
            text=title,
            font=FONT_TITLE,
            text_color=COLOR_FG_LIGHT,
        ).pack(anchor="w", padx=18, pady=(16, 4))
        ctk.CTkLabel(
            hero,
            text=subtitle,
            font=FONT_BODY,
            text_color=COLOR_TEXT_MUTED,
            wraplength=1080,
            justify="left",
        ).pack(anchor="w", padx=18, pady=(0, 16))
        return tab, wrapper

    def _create_info_chip(self, parent, text: str, fg_color: str, text_color: str = COLOR_FG_LIGHT):
        """Tworzy krótki znacznik informacyjny."""
        chip = ctk.CTkLabel(
            parent,
            text=text,
            font=FONT_SMALL,
            text_color=text_color,
            fg_color=fg_color,
            corner_radius=14,
            padx=10,
            pady=5,
        )
        chip.pack(side="left", padx=(0, 8))
        return chip

    def _create_collapsible_section(
        self,
        parent,
        title: str,
        subtitle: str = "",
        default_open: bool = True,
        expand: bool = False,
    ) -> Dict:
        """Tworzy sekcję, którą można zwinąć i rozwinąć."""
        section = ctk.CTkFrame(
            parent,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        section.pack(fill="both" if expand else "x", expand=expand, pady=(0, 10))

        header = ctk.CTkFrame(section, fg_color="transparent")
        header.pack(fill="x", padx=14, pady=(12, 8))

        icon_label = ctk.CTkLabel(
            header,
            text="▼" if default_open else "▶",
            font=FONT_BODY_BOLD,
            width=18,
        )
        icon_label.pack(side="left", padx=(0, 8))

        title_block = ctk.CTkFrame(header, fg_color="transparent")
        title_block.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(
            title_block,
            text=title,
            font=FONT_SECTION,
            text_color=COLOR_FG_LIGHT,
        ).pack(anchor="w")
        if subtitle:
            ctk.CTkLabel(
                title_block,
                text=subtitle,
                font=FONT_SMALL,
                text_color=COLOR_TEXT_MUTED,
            ).pack(anchor="w")

        content = ctk.CTkFrame(section, fg_color="transparent")
        if default_open:
            content.pack(
                fill="both" if expand else "x",
                expand=expand,
                padx=14,
                pady=(0, 12),
            )

        state = {"open": default_open}

        def toggle_section() -> None:
            state["open"] = not state["open"]
            icon_label.configure(text="▼" if state["open"] else "▶")
            if state["open"]:
                content.pack(
                    fill="both" if expand else "x",
                    expand=expand,
                    padx=14,
                    pady=(0, 12),
                )
            else:
                content.pack_forget()

        toggle_button = ctk.CTkButton(
            header,
            text="Pokaż / ukryj",
            command=toggle_section,
            width=120,
            height=32,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_SMALL,
        )
        toggle_button.pack(side="right")

        return {
            "frame": section,
            "content": content,
            "toggle": toggle_section,
            "state": state,
        }

    def _get_pending_change_count(self) -> int:
        """Zlicza wszystkie oczekujące zmiany do pokazania w UI."""
        count = 0
        for pending_key in ("attribute_changes", "skill_changes", "talent_changes"):
            count += sum(abs(change) for change in self.pending_changes[pending_key].values())
        count += len(self.pending_changes["new_skills"])
        count += len(self.pending_changes["new_talents"])
        if self.pending_changes["experience_delta"]:
            count += 1
        return count

    def refresh_header_summary(self) -> None:
        """Odświeża karty stanu w nagłówku."""
        available = self.data_manager.experience["available"]
        spent = self.data_manager.experience["spent"]
        total = self.data_manager.experience["total"]
        pending_count = self._get_pending_change_count()

        self.char_name_label.configure(text=self.data_manager.character_name)
        self.exp_label.configure(text=str(available))
        self.spent_card["value_label"].configure(text=str(spent))
        pending_text = f"{pending_count} pozycji" if pending_count else "Brak"
        self.pending_card["value_label"].configure(text=pending_text)

        self.available_card["badge"].configure(fg_color=COLOR_SUCCESS if available > 0 else COLOR_BORDER)
        self.spent_card["badge"].configure(fg_color=COLOR_HIGHLIGHT if spent > 0 else COLOR_BORDER)
        self.pending_card["badge"].configure(
            fg_color=COLOR_WARNING if pending_count else COLOR_BORDER
        )

        if hasattr(self, "exp_total_label"):
            self.exp_total_label.configure(text=f"Łączna pula postaci: {total} PD")
        if hasattr(self, "status_chip"):
            status_text = "Masz oczekujące zmiany" if pending_count else "Brak oczekujących zmian"
            status_color = COLOR_WARNING if pending_count else COLOR_SUCCESS
            self.status_chip.configure(text=status_text, fg_color=status_color)
        if hasattr(self, "source_chip"):
            source_label = "Źródło: PDF" if self.data_manager.source_type == FILE_TYPE_PDF else "Źródło: Excel"
            source_color = COLOR_ACCENT if self.data_manager.source_type == FILE_TYPE_PDF else COLOR_SURFACE_SOFT
            self.source_chip.configure(text=source_label, fg_color=source_color)

    def _recompute_developable(self) -> None:
        """Przelicza zbiory elementów rozwijalnych w bieżącej profesji (gating)."""
        dm = self.data_manager
        self._developable = game_data.get_career_developable(
            dm.current_career,
            dm.current_career_level,
            dm.talents,
        )

    def _developable_sets(self) -> Dict:
        """Zwraca cache rozwijalności (przelicza, jeśli brak)."""
        if not hasattr(self, "_developable") or self._developable is None:
            self._recompute_developable()
        return self._developable

    def _attr_developable(self, attr_name: str) -> bool:
        """Czy cecha jest rozwijalna w profesji (JSON nadrzędny, fallback flaga PDF)."""
        dev = self._developable_sets()
        chars = dev.get("characteristics") or set()
        if dev.get("resolved") and chars:
            return game_data.is_characteristic_developable(attr_name, dev)
        # cechy schematu nieuzupełnione lub brak profesji -> fallback do flagi z PDF
        return bool(self.data_manager.attributes.get(attr_name, {}).get("profession_available"))

    def _skill_developable(self, skill_name: str) -> bool:
        """Czy umiejętność jest rozwijalna w profesji (JSON nadrzędny, fallback flaga PDF)."""
        dev = self._developable_sets()
        if dev.get("resolved") and (dev.get("skills") or set()):
            return game_data.is_skill_developable(skill_name, dev)
        return bool(self.data_manager.skills.get(skill_name, {}).get("profession_available"))

    def _talent_developable(self, talent_name: str) -> bool:
        """Czy talent jest rozwijalny w profesji (JSON nadrzędny, fallback flaga PDF)."""
        dev = self._developable_sets()
        if dev.get("resolved") and (dev.get("talents") or set()):
            return game_data.is_talent_developable(talent_name, dev)
        return bool(self.data_manager.talents.get(talent_name, {}).get("profession_available"))

    def _is_out_of_profession(self, advancement_type: str, item_name: str) -> bool:
        """Czy rozwój danego elementu jest SPOZA profesji (koszt ×2).

        Karę nakładamy tylko, gdy mamy schemat profesji do oceny. Brak schematu
        (profesja spoza podstawki / nieustawiona) lub nieuzupełnione cechy -> brak kary.
        """
        dev = self._developable_sets()
        if not dev.get("resolved"):
            return False
        if advancement_type == "cecha":
            if not (dev.get("characteristics") or set()):
                return False
            return not game_data.is_characteristic_developable(item_name, dev)
        return not self._skill_developable(item_name)

    def _advancement_cost_mode(self, advancement_type: str, item_name: str) -> tuple:
        """Zwraca (out_of_profession, gm_approved) dla cechy/umiejętności."""
        out_of_profession = self._is_out_of_profession(advancement_type, item_name)
        gm_approved = out_of_profession and bool(self.dev_gm_approval_var.get())
        return out_of_profession, gm_approved

    def _on_dev_gm_change(self) -> None:
        """Reaguje na zmianę zgody MG dla rozwoju spoza profesji (cechy/umiejętności)."""
        if getattr(self, "initialized_attrs", False):
            self._update_all_attribute_labels()
        if getattr(self, "initialized_skills", False):
            self._update_all_skill_labels()
            self.apply_skill_filter()
        self.refresh_costs_if_visible()

    def _format_attribute_display_name(self, attr_name: str) -> str:
        """Buduje etykietę cechy z markerem rozwijalności w profesji."""
        suffix = " +" if self._attr_developable(attr_name) else ""
        return f"{attr_name}{suffix}"

    def _format_skill_display_name(self, skill_name: str) -> str:
        """Buduje etykietę umiejętności z markerem rozwijalności w profesji."""
        suffix = " +" if self._skill_developable(skill_name) else ""
        return f"{skill_name}{suffix}"

    def _format_talent_display_name(self, talent_name: str) -> str:
        """Buduje etykietę talentu z markerem rozwijalności w profesji."""
        suffix = " +" if self._talent_developable(talent_name) else ""
        return f"{talent_name}{suffix}"

    def _apply_row_developable_style(self, row, developable: bool) -> None:
        """Podświetla tło wiersza, gdy element jest rozwijalny w profesji."""
        if developable:
            row.configure(fg_color=COLOR_DEVELOPABLE_BG, border_color=COLOR_DEVELOPABLE_BORDER)
        else:
            row.configure(fg_color=COLOR_SURFACE_ALT, border_color=COLOR_BORDER)

    def create_character_tab(self) -> None:
        """Zakładka: Cechy postaci."""
        _, wrapper = self._create_tab_shell(
            "Cechy",
            "Cechy postaci",
            "Podgląd bieżących wartości, rezerwacji PD i maksymalnych możliwych rozwinięć dla każdej cechy.",
        )

        legend = ctk.CTkFrame(wrapper, fg_color="transparent")
        legend.pack(fill="x", pady=(0, 10))
        self._create_info_chip(legend, "Niebieski: aktualne rozwinięcia", COLOR_ACCENT)
        self._create_info_chip(legend, "Pomarańczowy: maksymalny zakup", COLOR_WARNING)
        self._create_info_chip(legend, "Zielony: możesz kupić", COLOR_SUCCESS)

        controls = ctk.CTkFrame(
            wrapper, fg_color=COLOR_SURFACE_ALT, corner_radius=16,
            border_width=1, border_color=COLOR_BORDER,
        )
        controls.pack(fill="x", pady=(0, 10))
        self.attr_profession_filter_checkbox = ctk.CTkCheckBox(
            controls,
            text="Tylko rozwijalne (+)",
            variable=self.attr_profession_filter_var,
            command=self.apply_attribute_filter,
            checkbox_width=20, checkbox_height=20, border_width=2, corner_radius=6,
            font=FONT_BODY,
        )
        self.attr_profession_filter_checkbox.pack(side="left", padx=(16, 12), pady=14)
        self.attr_gm_checkbox = ctk.CTkCheckBox(
            controls,
            text="Zgoda MG (spoza profesji ×1)",
            variable=self.dev_gm_approval_var,
            command=self._on_dev_gm_change,
            checkbox_width=20, checkbox_height=20, border_width=2, corner_radius=6,
            font=FONT_BODY,
        )
        self.attr_gm_checkbox.pack(side="left", padx=(0, 12), pady=14)

        self.attributes_frame = ctk.CTkScrollableFrame(
            wrapper,
            fg_color=COLOR_SURFACE,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.attributes_frame.pack(fill="both", expand=True)

    def create_profession_tab(self) -> None:
        """Zakładka: Profesja, klasa i ścieżka kariery."""
        _, wrapper = self._create_tab_shell(
            "Profesja",
            "Profesja i klasa",
            "Ustaw lub popraw bieżącą profesję (bez kosztu) albo awansuj do nowej profesji (z kosztem przejścia). "
            "Poniżej zobaczysz ścieżkę kariery, schemat 4 poziomów i status kompletowania obecnej profesji.",
        )

        # Jeden wspólny obszar przewijania dla całej zakładki
        outer = ctk.CTkScrollableFrame(wrapper, fg_color="transparent")
        outer.pack(fill="both", expand=True)

        # Podsumowanie bieżącej profesji
        summary = ctk.CTkFrame(
            outer, fg_color=COLOR_SURFACE_ALT, corner_radius=16,
            border_width=1, border_color=COLOR_BORDER,
        )
        summary.pack(fill="x", pady=(0, 10))
        header_row = ctk.CTkFrame(summary, fg_color="transparent")
        header_row.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(
            header_row, textvariable=self.profession_summary_var,
            font=FONT_BODY_BOLD, justify="left", anchor="w",
        ).pack(side="left", anchor="w")
        self.profession_status_label = ctk.CTkLabel(
            header_row, textvariable=self.profession_status_var,
            font=FONT_SMALL, text_color=COLOR_FG_LIGHT,
            fg_color=COLOR_SURFACE_SOFT, corner_radius=12, padx=12, pady=4,
        )
        self.profession_status_label.pack(side="right", anchor="e")
        ctk.CTkLabel(
            summary, textvariable=self.profession_completion_var,
            font=FONT_SMALL, text_color=COLOR_TEXT_MUTED, justify="left", anchor="w",
            wraplength=1060,
        ).pack(anchor="w", padx=16, pady=(0, 12))

        # Edytor wyboru profesji
        editor = ctk.CTkFrame(
            outer, fg_color=COLOR_SURFACE_ALT, corner_radius=16,
            border_width=1, border_color=COLOR_BORDER,
        )
        editor.pack(fill="x", pady=(0, 10))

        row1 = ctk.CTkFrame(editor, fg_color="transparent")
        row1.pack(fill="x", padx=16, pady=(14, 6))

        ctk.CTkLabel(row1, text="Klasa:", font=FONT_BODY_BOLD, width=70, anchor="w").pack(side="left")
        self.profession_class_combo = ctk.CTkComboBox(
            row1, values=game_data.all_class_names(), variable=self.profession_class_var,
            width=180, command=lambda _v: self._on_profession_class_change(),
        )
        self.profession_class_combo.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="Profesja:", font=FONT_BODY_BOLD, width=80, anchor="w").pack(side="left")
        self.profession_career_combo = ctk.CTkComboBox(
            row1, values=[], variable=self.profession_career_var,
            width=240, command=lambda _v: self._on_profession_career_change(),
        )
        self.profession_career_combo.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="Poziom:", font=FONT_BODY_BOLD, width=70, anchor="w").pack(side="left")
        self.profession_level_combo = ctk.CTkComboBox(
            row1, values=["1", "2", "3", "4"], variable=self.profession_level_var, width=80,
        )
        self.profession_level_combo.pack(side="left")

        row2 = ctk.CTkFrame(editor, fg_color="transparent")
        row2.pack(fill="x", padx=16, pady=(6, 14))

        ctk.CTkButton(
            row2, text="Ustaw / popraw (bez kosztu)", command=self.on_set_profession,
            width=240, height=36, fg_color=COLOR_SURFACE_SOFT, hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(0, 12))
        ctk.CTkButton(
            row2, text="Awansuj profesję (koszt PD)", command=self.on_advance_profession,
            width=240, height=36, fg_color=COLOR_ACCENT, hover_color="#0f96ea",
            font=FONT_BODY_BOLD,
        ).pack(side="left")

        # Ścieżka kariery
        path_section = ctk.CTkFrame(
            outer, fg_color=COLOR_SURFACE_ALT, corner_radius=16,
            border_width=1, border_color=COLOR_BORDER,
        )
        path_section.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(
            path_section, text="Ścieżka kariery", font=FONT_SECTION,
        ).pack(anchor="w", padx=16, pady=(12, 4))
        self.career_path_frame = ctk.CTkFrame(path_section, fg_color="transparent")
        self.career_path_frame.pack(fill="x", padx=12, pady=(0, 12))

        # Schemat 4 poziomów (wewnątrz wspólnego scrolla)
        self.profession_levels_frame = ctk.CTkFrame(
            outer, fg_color=COLOR_SURFACE, corner_radius=16,
            border_width=1, border_color=COLOR_BORDER,
        )
        self.profession_levels_frame.pack(fill="both", expand=True, pady=(0, 4))

        self.refresh_profession_display()

    def _on_profession_class_change(self) -> None:
        """Aktualizuje listę profesji po zmianie klasy."""
        class_name = self.profession_class_var.get()
        careers = game_data.careers_for_class(class_name)
        self.profession_career_combo.configure(values=careers)
        if careers and self.profession_career_var.get() not in careers:
            self.profession_career_var.set(careers[0])
            self._on_profession_career_change()

    def _on_profession_career_change(self) -> None:
        """Reaguje na zmianę wyboru profesji (np. odświeżenie poziomów)."""
        # Brak akcji natychmiastowej; podgląd po Ustaw/Awansuj.
        return

    def refresh_profession_display(self) -> None:
        """Odświeża podsumowanie, ścieżkę kariery i schemat poziomów."""
        if not hasattr(self, "profession_levels_frame"):
            return

        self._recompute_developable()

        dm = self.data_manager
        career = dm.current_career or "—"
        class_name = dm.character_class or game_data.class_of_career(dm.current_career) or "—"
        species = dm.character_species or "—"
        self.profession_summary_var.set(
            f"Profesja: {career}  (poziom {dm.current_career_level})    "
            f"Klasa: {class_name}    Rasa: {species}"
        )

        self._update_profession_completion_label()
        self._sync_profession_inputs()
        self._render_career_path()
        self._render_profession_levels()
        self._refresh_gating_displays()

    def _refresh_gating_displays(self) -> None:
        """Odświeża markery/filtry rozwijalności w zakładkach Cechy/Umiejętności/Talenty."""
        if getattr(self, "initialized_attrs", False):
            self._update_all_attribute_labels()
            self.apply_attribute_filter()
        if getattr(self, "initialized_skills", False):
            self._update_all_skill_labels()
            self.apply_skill_filter()
        if getattr(self, "initialized_talents", False):
            self._update_all_talent_labels()
            self.apply_talent_filter()

    def _update_profession_completion_label(self) -> None:
        """Buduje opis statusu kompletowania bieżącej profesji."""
        status = self.data_manager.current_career_completion()
        if status.get("unknown_profession"):
            self.profession_completion_var.set(
                "Profesja spoza podstawki (brak danych schematu) — kompletowanie liczone ręcznie."
            )
            self._set_profession_status_chip("SPOZA PODSTAWKI", COLOR_SURFACE_SOFT)
            return
        parts = []
        parts.append(
            f"Umiejętności zawodowe: {status['skills_done']}/8 "
            + ("✓" if status["skills_ok"] else "✗")
        )
        parts.append(
            f"Talenty: {status['talents_done']} "
            + ("✓" if status["talents_ok"] else "✗")
        )
        if status["characteristics_pending"]:
            parts.append("Cechy: schemat nieuzupełniony (pominięto)")
        else:
            parts.append("Cechy: " + ("✓" if status["characteristics_ok"] else "✗"))
        header = "PROFESJA SKOMPLETOWANA" if status["completed"] else "Profesja w trakcie"
        self.profession_completion_var.set(header + "  —  " + "   |   ".join(parts))
        if status["completed"]:
            self._set_profession_status_chip("✓ SKOMPLETOWANA", COLOR_SUCCESS)
        else:
            self._set_profession_status_chip("● NIESKOMPLETOWANA", COLOR_WARNING)

    def _set_profession_status_chip(self, text: str, color: str) -> None:
        """Ustawia tekst i kolor odznaki statusu profesji."""
        self.profession_status_var.set(text)
        if hasattr(self, "profession_status_label"):
            self.profession_status_label.configure(fg_color=color)

    def _sync_profession_inputs(self) -> None:
        """Ustawia wartości pól wyboru zgodnie z bieżącą profesją."""
        dm = self.data_manager
        class_name = dm.character_class or game_data.class_of_career(dm.current_career)
        if class_name and class_name in game_data.all_class_names():
            self.profession_class_var.set(class_name)
            self.profession_career_combo.configure(
                values=game_data.careers_for_class(class_name)
            )
        if dm.current_career:
            self.profession_career_var.set(dm.current_career)
        self.profession_level_var.set(str(dm.current_career_level or 1))

    def _render_career_path(self) -> None:
        """Rysuje listę kroków ścieżki kariery z przyciskami usuwania."""
        for widget in self.career_path_frame.winfo_children():
            widget.destroy()

        path = self.data_manager.career_path
        if not path:
            ctk.CTkLabel(
                self.career_path_frame,
                text="Brak zapisanej ścieżki kariery.",
                font=FONT_SMALL, text_color=COLOR_TEXT_MUTED,
            ).pack(anchor="w", padx=4, pady=4)
            return

        for index, step in enumerate(path):
            is_current = index == len(path) - 1
            chip = ctk.CTkFrame(
                self.career_path_frame,
                fg_color=COLOR_ACCENT if is_current else COLOR_SURFACE_SOFT,
                corner_radius=10,
            )
            chip.pack(side="left", padx=4, pady=4)
            title = step.get("title") or step.get("profession") or "?"
            level = step.get("level") or "?"
            badge = "" if step.get("resolved", True) else "  (spoza podstawki)"
            done = "✓" if step.get("completed") else "•"
            ctk.CTkLabel(
                chip, text=f"{done} {title} ({level}){badge}",
                font=FONT_SMALL, text_color=COLOR_FG_LIGHT,
            ).pack(side="left", padx=(10, 6), pady=6)
            ctk.CTkButton(
                chip, text="✕", width=24, height=24,
                fg_color="transparent", hover_color=COLOR_ERROR,
                command=lambda i=index: self.on_remove_career_step(i),
            ).pack(side="left", padx=(0, 6), pady=4)

    def _render_profession_levels(self) -> None:
        """Rysuje schemat 4 poziomów bieżącej profesji."""
        for widget in self.profession_levels_frame.winfo_children():
            widget.destroy()

        prof = game_data.get_profession(self.data_manager.current_career)
        if not prof:
            ctk.CTkLabel(
                self.profession_levels_frame,
                text="Brak danych schematu dla tej profesji (profesja spoza podstawki "
                "albo nieustawiona). Możesz nadal rozwijać postać ręcznie.",
                font=FONT_BODY, text_color=COLOR_TEXT_MUTED, wraplength=1040, justify="left",
            ).pack(anchor="w", padx=16, pady=16)
            return

        current_level = self.data_manager.current_career_level
        for lvl in prof.get("levels", []):
            level_num = lvl.get("level")
            is_current = level_num == current_level
            is_past = level_num < current_level
            border = COLOR_ACCENT if is_current else COLOR_BORDER
            card = ctk.CTkFrame(
                self.profession_levels_frame,
                fg_color=COLOR_SURFACE_ALT if is_current else COLOR_SURFACE_SOFT,
                corner_radius=14, border_width=2 if is_current else 1, border_color=border,
            )
            card.pack(fill="x", padx=10, pady=6)

            tag = "OBECNY POZIOM" if is_current else ("ukończony" if is_past else "do rozwoju później")
            ctk.CTkLabel(
                card,
                text=f"Poziom {level_num}: {lvl.get('title', '')}   ({lvl.get('status', '')})   [{tag}]",
                font=FONT_BODY_BOLD,
                text_color=COLOR_HIGHLIGHT if is_current else COLOR_FG_LIGHT,
            ).pack(anchor="w", padx=14, pady=(10, 4))

            self._add_scheme_line(card, "Cechy", lvl.get("characteristics", []))
            self._add_scheme_line(card, "Umiejętności", lvl.get("skills", []))
            self._add_scheme_line(card, "Talenty", lvl.get("talents", []))
            self._add_scheme_line(card, "Wyposażenie", lvl.get("trappings", []))

    def _add_scheme_line(self, parent, label: str, items: List[str]) -> None:
        """Dodaje wiersz 'Etykieta: lista' do karty poziomu."""
        if not items:
            text = f"{label}: —"
        else:
            text = f"{label}: " + ", ".join(items)
        ctk.CTkLabel(
            parent, text=text, font=FONT_SMALL, text_color=COLOR_TEXT_MUTED,
            wraplength=1020, justify="left", anchor="w",
        ).pack(fill="x", anchor="w", padx=14, pady=(0, 2))

    # ----- Operacje na profesji -------------------------------------------
    def _selected_profession_inputs(self):
        """Zwraca (klasa, profesja, poziom) z pól wyboru."""
        class_name = self.profession_class_var.get().strip()
        career = self.profession_career_var.get().strip()
        try:
            level = int(self.profession_level_var.get())
        except (ValueError, TypeError):
            level = 1
        return class_name, career, max(1, min(4, level))

    def on_set_profession(self) -> None:
        """Ustawia/poprawia bieżącą profesję bez naliczania kosztu."""
        class_name, career, level = self._selected_profession_inputs()
        if not career:
            messagebox.showwarning("Brak profesji", "Wybierz profesję z listy.")
            return

        self.data_manager.set_current_career(career, level)
        if class_name and class_name in game_data.all_class_names():
            self.data_manager.character_class = class_name

        self._persist_career_path()
        self.history_manager.add_entry(
            "Ustawiono/poprawiono profesję", f"{career} (poziom {level})"
        )
        self.refresh_profession_display()
        self.update_character_info()
        self.refresh_history_display()
        messagebox.showinfo("Profesja", f"Ustawiono profesję: {career} (poziom {level}).")

    def on_advance_profession(self) -> None:
        """Awansuje do nowej profesji z naliczeniem kosztu przejścia."""
        class_name, career, level = self._selected_profession_inputs()
        if not career:
            messagebox.showwarning("Brak profesji", "Wybierz profesję z listy.")
            return

        dm = self.data_manager
        # Pierwsza profesja postaci = ustawienie bez kosztu.
        if not dm.current_career:
            messagebox.showinfo(
                "Pierwsza profesja",
                "Brak bieżącej profesji — użyj 'Ustaw / popraw', aby przypisać pierwszą profesję bez kosztu.",
            )
            return

        if career == dm.current_career and level == dm.current_career_level:
            messagebox.showinfo(
                "Brak zmiany", "Wybrana profesja i poziom są takie same jak obecne."
            )
            return

        status = dm.current_career_completion()
        completed = bool(status.get("completed"))
        current_class = dm.character_class or game_data.class_of_career(dm.current_career) or ""
        target_class = (
            game_data.class_of_career(career)
            or (class_name if class_name in game_data.all_class_names() else "")
        )
        same_class = bool(current_class) and current_class == target_class

        cost = game_data.career_transition_cost(completed, same_class)

        if dm.experience["available"] < cost:
            messagebox.showerror(
                "Za mało PD",
                f"Awans kosztuje {cost} PD, masz {dm.experience['available']} PD.",
            )
            return

        completion_note = "skompletowana" if completed else "NIESKOMPLETOWANA"
        class_note = "ta sama klasa" if same_class else "ZMIANA KLASY"
        confirm = messagebox.askyesno(
            "Awans profesji",
            f"Awans: {dm.current_career} ({dm.current_career_level}) -> {career} ({level})\n"
            f"Obecna profesja: {completion_note}\n"
            f"Klasa: {class_note}\n\n"
            f"Koszt przejścia: {cost} PD\n\nKontynuować?",
        )
        if not confirm:
            return
        dm.experience["available"] -= cost
        dm.experience["spent"] += cost
        dm.experience["total"] = dm.experience["available"] + dm.experience["spent"]
        dm.advance_to_career(career, level)
        if class_name and class_name in game_data.all_class_names() and not game_data.class_of_career(career):
            dm.character_class = class_name

        self._persist_career_path()
        self.history_manager.add_entry(
            "Awans profesji", f"{career} (poziom {level}), koszt {cost} PD"
        )

        if dm.file_path:
            if dm.source_type == FILE_TYPE_PDF:
                dm.save_to_pdf(dm.file_path)
            else:
                dm.save_to_excel(dm.file_path)

        self.refresh_profession_display()
        self.update_character_info()
        self.update_experience_display()
        self.refresh_history_display()
        messagebox.showinfo("Awans", f"Awansowano do: {career} (poziom {level}).")

    def on_remove_career_step(self, index: int) -> None:
        """Usuwa krok ze ścieżki kariery (korekta)."""
        path = self.data_manager.career_path
        if index < 0 or index >= len(path):
            return
        step = path[index]
        title = step.get("title") or step.get("profession") or "?"
        if not messagebox.askyesno(
            "Usuń krok kariery", f"Usunąć krok '{title}' ze ścieżki kariery?"
        ):
            return

        del path[index]
        if path:
            last = path[-1]
            last["completed"] = False
            self.data_manager.current_career = last.get("profession") or last.get("title") or ""
            self.data_manager.current_career_level = last.get("level") or 1
            resolved_class = game_data.class_of_career(self.data_manager.current_career)
            if resolved_class:
                self.data_manager.character_class = resolved_class
        else:
            self.data_manager.current_career = ""
            self.data_manager.current_career_level = 1

        self._persist_career_path()
        self.history_manager.add_entry("Korekta ścieżki kariery", f"Usunięto: {title}")
        self.refresh_profession_display()
        self.update_character_info()
        self.refresh_history_display()

    def _persist_career_path(self) -> None:
        """Zapisuje bieżącą ścieżkę kariery do historii postaci."""
        if self.data_manager.character_name:
            self.history_manager.set_career_path(
                self.data_manager.character_name, self.data_manager.career_path
            )

    def create_skills_tab(self) -> None:
        """Zakładka: Umiejętności."""
        _, wrapper = self._create_tab_shell(
            "Umiejętności",
            "Umiejętności postaci",
            "Filtruj listę po nazwie albo atrybucie i szybko sprawdzaj, ile rozwinięć da się jeszcze kupić.",
        )

        controls = ctk.CTkFrame(
            wrapper,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        controls.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            controls,
            text="Filtr umiejętności:",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(16, 8), pady=14)
        self.skill_filter_entry = ctk.CTkEntry(
            controls,
            textvariable=self.skill_filter_var,
            placeholder_text="Wpisz fragment nazwy lub atrybut, np. Int albo Plotkowanie",
            width=300,
            height=36,
        )
        self.skill_filter_entry.pack(side="left", padx=(0, 12), pady=14)
        self.skill_filter_entry.bind("<KeyRelease>", lambda _event: self.apply_skill_filter())
        AutocompletePopup(
            self.skill_filter_entry,
            self.skill_filter_var,
            lambda: sorted(self.data_manager.skills.keys(), key=str.casefold),
            lambda _v: self.apply_skill_filter(),
        )

        self.skill_attribute_filter_combo = ctk.CTkComboBox(
            controls,
            values=[ATTRIBUTE_FILTER_ALL] + [get_attribute_filter_label(attr) for attr in ATTRIBUTES],
            variable=self.skill_attribute_filter_var,
            width=220,
            height=36,
            command=lambda _value: self.apply_skill_filter(),
        )
        self.skill_attribute_filter_combo.pack(side="left", padx=(0, 12), pady=14)

        self.skill_profession_filter_checkbox = ctk.CTkCheckBox(
            controls,
            text="Tylko rozwijalne (+)",
            variable=self.skill_profession_filter_var,
            command=self.apply_skill_filter,
            checkbox_width=20,
            checkbox_height=20,
            border_width=2,
            corner_radius=6,
            font=FONT_BODY,
        )
        self.skill_profession_filter_checkbox.pack(side="left", padx=(0, 12), pady=14)

        self.skill_gm_checkbox = ctk.CTkCheckBox(
            controls,
            text="Zgoda MG (spoza profesji ×1)",
            variable=self.dev_gm_approval_var,
            command=self._on_dev_gm_change,
            checkbox_width=20,
            checkbox_height=20,
            border_width=2,
            corner_radius=6,
            font=FONT_BODY,
        )
        self.skill_gm_checkbox.pack(side="left", padx=(0, 12), pady=14)

        ctk.CTkButton(
            controls,
            text="Wyczyść filtr",
            command=self.clear_skill_filter,
            width=120,
            height=36,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(0, 12), pady=14)

        ctk.CTkButton(
            controls,
            text="Dodaj umiejętność",
            command=self.on_add_skill_dialog,
            width=160,
            height=36,
            fg_color=COLOR_ACCENT,
            hover_color="#0f96ea",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(0, 12), pady=14)

        self.skill_filter_hint_button = ctk.CTkButton(
            controls,
            text="",
            command=self.apply_suggested_attribute_filter,
            width=200,
            height=36,
            fg_color=COLOR_ACCENT,
            hover_color="#0f96ea",
            font=FONT_SMALL,
        )

        self.skills_summary_label = ctk.CTkLabel(
            controls,
            textvariable=self.skill_summary_var,
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
        )
        self.skills_summary_label.pack(side="right", padx=16, pady=14)

        self.skills_frame = ctk.CTkScrollableFrame(
            wrapper,
            fg_color=COLOR_SURFACE,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.skills_frame.pack(fill="both", expand=True)
        self._build_skill_sections()

    def _build_skill_sections(self) -> None:
        """Tworzy stałe sekcje dla umiejętności podstawowych i zaawansowanych."""
        for widget in self.skills_frame.winfo_children():
            widget.destroy()

        self.skills_group_frames = {}
        for category_key, title in (
            (SKILL_CATEGORY_BASIC, "Podstawowe"),
            (SKILL_CATEGORY_ADVANCED, "Zaawansowane / grupowe"),
        ):
            section = ctk.CTkFrame(
                self.skills_frame,
                fg_color=COLOR_SURFACE_SOFT,
                corner_radius=16,
                border_width=1,
                border_color=COLOR_BORDER,
            )
            section.pack(fill="x", padx=10, pady=(8, 6))
            header = ctk.CTkLabel(
                section,
                text=title,
                font=FONT_SECTION,
                text_color=COLOR_FG_LIGHT,
            )
            header.pack(anchor="w", padx=14, pady=(10, 6))
            content = ctk.CTkFrame(section, fg_color="transparent")
            content.pack(fill="x", padx=8, pady=(0, 8))

            self.skills_group_frames[category_key] = {
                "section": section,
                "header": header,
                "content": content,
            }

    def create_talents_tab(self) -> None:
        """Zakładka: Talenty (zakup, własne talenty, twardy Max, rozwój spoza profesji)."""
        _, wrapper = self._create_tab_shell(
            "Talenty",
            "Talenty",
            "Dodawaj talenty z listy lub własne, kupuj kolejne wykupienia (100 PD każde) i pilnuj limitu (Maksimum). Rozwój spoza profesji kosztuje podwójnie, chyba że MG wyrazi zgodę.",
        )

        controls = ctk.CTkFrame(
            wrapper,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        controls.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            controls,
            text="Filtr talentów:",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(16, 8), pady=14)
        self.talent_filter_entry = ctk.CTkEntry(
            controls,
            textvariable=self.talent_filter_var,
            placeholder_text="Wpisz fragment nazwy lub opisu talentu",
            width=300,
            height=36,
        )
        self.talent_filter_entry.pack(side="left", padx=(0, 12), pady=14)
        self.talent_filter_entry.bind("<KeyRelease>", lambda _event: self.apply_talent_filter())
        AutocompletePopup(
            self.talent_filter_entry,
            self.talent_filter_var,
            lambda: sorted(self.data_manager.talents.keys(), key=str.casefold),
            lambda _v: self.apply_talent_filter(),
        )

        self.talent_profession_filter_checkbox = ctk.CTkCheckBox(
            controls,
            text="Tylko rozwijalne (+)",
            variable=self.talent_profession_filter_var,
            command=self.apply_talent_filter,
            checkbox_width=20, checkbox_height=20, border_width=2, corner_radius=6,
            font=FONT_BODY,
        )
        self.talent_profession_filter_checkbox.pack(side="left", padx=(0, 12), pady=14)

        ctk.CTkButton(
            controls,
            text="Wyczyść filtr",
            command=self.clear_talent_filter,
            width=120,
            height=36,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(0, 12), pady=14)

        self.talents_summary_label = ctk.CTkLabel(
            controls,
            textvariable=self.talent_summary_var,
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
        )
        self.talents_summary_label.pack(side="right", padx=16, pady=14)

        # Pasek dodawania i opcji rozwoju spoza profesji
        actions = ctk.CTkFrame(
            wrapper,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        actions.pack(fill="x", pady=(0, 10))

        ctk.CTkButton(
            actions,
            text="Dodaj talent z listy",
            command=self.on_add_talent_from_list,
            width=180,
            height=36,
            fg_color=COLOR_ACCENT,
            hover_color="#0f96ea",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(16, 8), pady=14)

        ctk.CTkButton(
            actions,
            text="Dodaj własny talent",
            command=self.on_add_custom_talent,
            width=180,
            height=36,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left", padx=(0, 16), pady=14)

        ctk.CTkCheckBox(
            actions,
            text="Rozwój spoza profesji (×2)",
            variable=self.talent_out_of_profession_var,
            command=self._on_talent_cost_mode_change,
            checkbox_width=20,
            checkbox_height=20,
            border_width=2,
            corner_radius=6,
            font=FONT_BODY,
        ).pack(side="left", padx=(0, 12), pady=14)

        ctk.CTkCheckBox(
            actions,
            text="Zgoda MG (×1)",
            variable=self.talent_gm_approval_var,
            command=self._on_talent_cost_mode_change,
            checkbox_width=20,
            checkbox_height=20,
            border_width=2,
            corner_radius=6,
            font=FONT_BODY,
        ).pack(side="left", padx=(0, 12), pady=14)

        self.talents_frame = ctk.CTkScrollableFrame(
            wrapper,
            fg_color=COLOR_SURFACE,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.talents_frame.pack(fill="both", expand=True)

        self.talents_empty_label = ctk.CTkLabel(
            self.talents_frame,
            text="Brak talentów. Dodaj talent z listy albo własny powyżej.",
            font=FONT_BODY,
            text_color=COLOR_TEXT_MUTED,
        )

        self.initialized_talents = False
        self.initialize_talents_display()

    def _on_talent_cost_mode_change(self) -> None:
        """Reaguje na zmianę trybu rozwoju spoza profesji / zgody MG."""
        if self.talent_out_of_profession_var.get():
            self.talent_gm_approval_var.set(self.talent_gm_approval_var.get())
        else:
            self.talent_gm_approval_var.set(False)
        self.refresh_talents_display()

    def _talent_cost_mode(self) -> tuple:
        """Zwraca (out_of_profession, gm_approved) z bieżących przełączników."""
        out_of_profession = bool(self.talent_out_of_profession_var.get())
        gm_approved = out_of_profession and bool(self.talent_gm_approval_var.get())
        return out_of_profession, gm_approved

    def initialize_talents_display(self) -> None:
        """Tworzy wiersze talentów jeden raz (bez destroy)."""
        if self.initialized_talents:
            self._update_all_talent_labels()
            self.apply_talent_filter()
            return

        for widget in self.talents_frame.winfo_children():
            if widget is not self.talents_empty_label:
                widget.destroy()
        self.talent_rows.clear()

        for talent_name, talent_data in self.data_manager.talents.items():
            self._create_talent_row_cached(talent_name, talent_data)

        self.initialized_talents = True
        self.apply_talent_filter()

    def _format_talent_max(self, name: str) -> str:
        """Buduje czytelny opis limitu wykupień talentu."""
        talent = self.data_manager.talents.get(name, {})
        info = talent.get("max")
        limit = self.data_manager.talent_max_advances(name)
        if not info:
            return "brak"
        kind = info.get("type")
        if kind == "fixed":
            return str(info.get("value"))
        if kind == "characteristic":
            attr = info.get("attr")
            if limit is not None:
                return f"{limit} (bonus {attr})"
            return f"bonus {attr}"
        if kind == "special":
            return "specjalne"
        return "brak"

    def _create_talent_row_cached(self, talent_name: str, talent_data: Dict) -> None:
        """Tworzy wiersz talentu i zapamiętuje referencje do widgetów."""
        row = ctk.CTkFrame(
            self.talents_frame,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=14,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        row.pack(fill="x", pady=5, padx=10)

        labels = []
        buttons = []
        top = ctk.CTkFrame(row, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=(8, 2))

        lbl_name = ctk.CTkLabel(
            top, text=self._format_talent_display_name(talent_name), font=FONT_BODY_BOLD, width=240, anchor="w"
        )
        lbl_name.pack(side="left", padx=(0, 6))
        labels.append(("name", lbl_name))

        source_text = "własny" if talent_data.get("is_custom") else talent_data.get("source", "")
        lbl_source = ctk.CTkLabel(
            top,
            text=source_text,
            width=120,
            wraplength=120,
            justify="left",
            anchor="w",
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
        )
        lbl_source.pack(side="left", padx=(0, 6))
        labels.append(("source", lbl_source))

        lbl_adv = ctk.CTkLabel(
            top, text=f"Wykupienia: {talent_data['advances']}", width=130,
            font=FONT_BODY_BOLD, text_color=COLOR_ACCENT,
        )
        lbl_adv.pack(side="left", padx=2)
        labels.append(("adv", lbl_adv))

        lbl_max = ctk.CTkLabel(
            top, text=f"Maks: {self._format_talent_max(talent_name)}", width=150,
            font=FONT_BODY_BOLD, text_color=COLOR_WARNING,
        )
        lbl_max.pack(side="left", padx=2)
        labels.append(("max", lbl_max))

        btn_plus1 = ctk.CTkButton(
            top, text="+1", width=46, height=34,
            command=lambda: self.on_increase_talent(talent_name, 1),
            fg_color=COLOR_SUCCESS if self.can_buy_talent(talent_name) else COLOR_ERROR,
        )
        btn_plus1.pack(side="left", padx=1)
        buttons.append(("plus1", btn_plus1))

        btn_minus1 = ctk.CTkButton(
            top, text="-1", width=46, height=34,
            command=lambda: self.on_decrease_talent(talent_name, 1),
            fg_color=COLOR_WARNING if self._can_decrease_talent(talent_name) else "#555555",
        )
        btn_minus1.pack(side="left", padx=1)
        buttons.append(("minus1", btn_minus1))

        if talent_data.get("is_new", False):
            btn_delete = ctk.CTkButton(
                top, text="Usuń", width=62, height=34,
                command=lambda: self.on_delete_talent(talent_name),
                fg_color=COLOR_ERROR,
            )
            btn_delete.pack(side="left", padx=2)
            buttons.append(("delete", btn_delete))

        description = talent_data.get("description") or ""
        # Pełny opis pokazujemy w dymku po najechaniu (4A3) zamiast stałej etykiety.
        hint_text = "ⓘ Najedź, aby zobaczyć opis" if description else "Brak opisu"
        lbl_desc = ctk.CTkLabel(
            row,
            text=hint_text,
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
            justify="left",
            anchor="w",
        )
        lbl_desc.pack(fill="x", anchor="w", padx=10, pady=(0, 8))
        labels.append(("desc", lbl_desc))

        tooltip = Tooltip(row, text=description)
        tooltip.attach(lbl_name)
        tooltip.attach(lbl_desc)

        self.talent_rows[talent_name] = {
            "row": row,
            "labels": labels,
            "buttons": buttons,
            "tooltip": tooltip,
        }
        self._apply_row_developable_style(row, self._talent_developable(talent_name))

    def _update_all_talent_labels(self) -> None:
        """Aktualizuje teksty i kolory wierszy talentów bez tworzenia nowych widgetów."""
        for talent_name, talent_data in self.data_manager.talents.items():
            row_data = self.talent_rows.get(talent_name)
            if not row_data:
                continue
            labels = dict(row_data["labels"])
            buttons = dict(row_data["buttons"])
            if "name" in labels:
                labels["name"].configure(text=self._format_talent_display_name(talent_name))
            labels["adv"].configure(text=f"Wykupienia: {talent_data['advances']}")
            labels["max"].configure(text=f"Maks: {self._format_talent_max(talent_name)}")
            if "desc" in labels:
                description = talent_data.get("description") or ""
                labels["desc"].configure(
                    text="ⓘ Najedź, aby zobaczyć opis" if description else "Brak opisu"
                )
            if "tooltip" in row_data:
                row_data["tooltip"].set_text(talent_data.get("description") or "")
            buttons["plus1"].configure(
                fg_color=COLOR_SUCCESS if self.can_buy_talent(talent_name) else COLOR_ERROR
            )
            buttons["minus1"].configure(
                fg_color=COLOR_WARNING if self._can_decrease_talent(talent_name) else "#555555"
            )
            self._apply_row_developable_style(row_data["row"], self._talent_developable(talent_name))

    def refresh_talents_display(self) -> None:
        """Aktualizuje listę talentów (lekko) albo inicjalizuje przy pierwszym użyciu."""
        if not hasattr(self, "talents_frame"):
            return
        if self.initialized_talents:
            self._update_all_talent_labels()
            self.apply_talent_filter()
        else:
            self.initialize_talents_display()

    def clear_talent_filter(self) -> None:
        """Czyści filtr listy talentów."""
        self.talent_filter_var.set("")
        self.apply_talent_filter()

    def apply_talent_filter(self) -> None:
        """Filtruje widoczne talenty po nazwie lub opisie."""
        if not hasattr(self, "talents_frame"):
            return

        query = normalize_search_text(self.talent_filter_var.get())
        profession_only = bool(self.talent_profession_filter_var.get())
        visible_count = 0
        total_count = len(self.talent_rows)

        self.talents_empty_label.pack_forget()
        if total_count == 0:
            self.talents_empty_label.pack(anchor="w", padx=16, pady=16)
            self.talent_summary_var.set("Wyświetlane talenty: 0 / 0")
            return

        for talent_name, row_data in self.talent_rows.items():
            talent_data = self.data_manager.talents.get(talent_name, {})
            searchable = normalize_search_text(
                f"{talent_name} {talent_data.get('description', '')} {talent_data.get('source', '')}"
            )
            matches_query = not query or query in searchable
            matches_profession = not profession_only or self._talent_developable(talent_name)
            should_show = matches_query and matches_profession
            row_data["row"].pack_forget()
            if should_show:
                visible_count += 1
                row_data["row"].pack(fill="x", pady=5, padx=10)

        self.talent_summary_var.set(
            f"Wyświetlane talenty: {visible_count} / {total_count}"
        )

    # ----- Operacje na talentach ------------------------------------------
    def can_buy_talent(self, talent_name: str) -> bool:
        """Czy stać postać na kolejne wykupienie i czy nie przekroczono Max."""
        if not self._talent_below_max(talent_name, 1):
            return False
        out_of_profession, gm_approved = self._talent_cost_mode()
        cost = calculate_talent_cost(1, out_of_profession, gm_approved)
        return self.data_manager.experience["available"] >= cost

    def _talent_below_max(self, talent_name: str, amount: int) -> bool:
        """Sprawdza, czy po dokupieniu 'amount' nie przekroczymy twardego Max."""
        talent = self.data_manager.talents.get(talent_name)
        if not talent:
            return False
        limit = self.data_manager.talent_max_advances(talent_name)
        if limit is None:
            return True
        return talent["advances"] + amount <= limit

    def _can_decrease_talent(self, talent_name: str) -> bool:
        """Czy można cofnąć wykupienie (nie poniżej zatwierdzonego poziomu)."""
        talent = self.data_manager.talents.get(talent_name)
        if not talent:
            return False
        return talent["advances"] > talent.get("base_advances", 0)

    def on_increase_talent(self, talent_name: str, amount: int = 1) -> None:
        """Kupuje kolejne wykupienie talentu (do zatwierdzenia)."""
        talent = self.data_manager.talents.get(talent_name)
        if not talent:
            return

        if not self._talent_below_max(talent_name, amount):
            limit = self.data_manager.talent_max_advances(talent_name)
            messagebox.showwarning(
                "Osiągnięto Maksimum",
                f"Talent '{talent_name}' ma limit {limit} wykupień i nie można go przekroczyć.",
            )
            return

        out_of_profession, gm_approved = self._talent_cost_mode()
        cost = calculate_talent_cost(amount, out_of_profession, gm_approved)
        if self.data_manager.experience["available"] < cost:
            messagebox.showerror(
                "Za mało PD",
                f"Potrzebujesz {cost} PD, masz {self.data_manager.experience['available']} PD.",
            )
            return

        self.pending_changes["talent_changes"][talent_name] = (
            self.pending_changes["talent_changes"].get(talent_name, 0) + amount
        )
        talent["advances"] += amount
        self.data_manager.experience["available"] -= cost

        self.refresh_talents_display()
        self.update_experience_display()

    def on_decrease_talent(self, talent_name: str, amount: int = 1) -> None:
        """Cofa wykupienie talentu (zwrot PD wg trybu kosztu)."""
        talent = self.data_manager.talents.get(talent_name)
        if not talent:
            return

        minimum = talent.get("base_advances", 0)
        if talent["advances"] - amount < minimum:
            play_blocked_sound()
            return

        out_of_profession, gm_approved = self._talent_cost_mode()
        refund = calculate_talent_cost(amount, out_of_profession, gm_approved)

        self.pending_changes["talent_changes"][talent_name] = (
            self.pending_changes["talent_changes"].get(talent_name, 0) - amount
        )
        talent["advances"] -= amount
        self.data_manager.experience["available"] += refund

        self.refresh_talents_display()
        self.update_experience_display()

    def on_add_talent_from_list(self) -> None:
        """Otwiera okno wyboru talentu z bazy i dodaje go do postaci."""
        available = [
            name
            for name in game_data.all_talent_names()
            if name not in self.data_manager.talents
        ]
        if not available:
            messagebox.showinfo(
                "Brak talentów",
                "Wszystkie talenty z listy zostały już dodane do postaci.",
            )
            return
        self._open_talent_picker(available)

    def _open_talent_picker(self, available: List[str]) -> None:
        """Modal z wyszukiwarką talentów do dodania."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Dodaj talent z listy")
        dialog.geometry("520x520")
        dialog.transient(self)
        dialog.grab_set()

        search_var = tk.StringVar(value="")
        ctk.CTkLabel(dialog, text="Wyszukaj talent:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(16, 4)
        )
        search_entry = ctk.CTkEntry(dialog, textvariable=search_var, width=480, height=34)
        search_entry.pack(padx=16, pady=(0, 8))

        desc_label = ctk.CTkLabel(
            dialog,
            text="",
            font=FONT_SMALL,
            text_color=COLOR_TEXT_MUTED,
            wraplength=480,
            justify="left",
            anchor="w",
        )
        desc_label.pack(fill="x", padx=16, pady=(0, 6))

        list_frame = ctk.CTkScrollableFrame(dialog, width=480, height=300)
        list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        selected = {"name": None}
        row_buttons: Dict[str, ctk.CTkButton] = {}

        def select(name: str) -> None:
            selected["name"] = name
            talent = game_data.get_talent(name) or {}
            max_raw = talent.get("max_raw", "")
            desc_label.configure(
                text=f"Maksimum: {max_raw}\n{talent.get('description', '')}"
            )
            for other, btn in row_buttons.items():
                btn.configure(fg_color=COLOR_ACCENT if other == name else COLOR_SURFACE_SOFT)

        def render(filter_text: str = "") -> None:
            for child in list_frame.winfo_children():
                child.destroy()
            row_buttons.clear()
            query = normalize_search_text(filter_text)
            for name in available:
                if query and query not in normalize_search_text(name):
                    continue
                btn = ctk.CTkButton(
                    list_frame,
                    text=name,
                    anchor="w",
                    height=30,
                    fg_color=COLOR_SURFACE_SOFT,
                    hover_color="#48505e",
                    command=lambda n=name: select(n),
                )
                btn.pack(fill="x", pady=2)
                row_buttons[name] = btn

        def confirm() -> None:
            name = selected["name"]
            if not name:
                messagebox.showwarning("Brak wyboru", "Wybierz talent z listy.", parent=dialog)
                return
            dialog.destroy()
            self._add_talent_and_register(name, from_database=True)

        search_entry.bind("<KeyRelease>", lambda _e: render(search_var.get()))
        button_row = ctk.CTkFrame(dialog, fg_color="transparent")
        button_row.pack(fill="x", padx=16, pady=(0, 16))
        ctk.CTkButton(
            button_row, text="Dodaj", command=confirm, fg_color=COLOR_SUCCESS, width=120
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            button_row, text="Anuluj", command=dialog.destroy,
            fg_color=COLOR_SURFACE_SOFT, hover_color="#48505e", width=120,
        ).pack(side="right")

        render()
        search_entry.focus_set()

    def _add_talent_and_register(
        self, name: str, from_database: bool, custom_info: Optional[Dict] = None
    ) -> None:
        """Dodaje talent (z bazy lub własny), rezerwuje koszt pierwszego wykupienia."""
        out_of_profession, gm_approved = self._talent_cost_mode()
        cost = calculate_talent_cost(1, out_of_profession, gm_approved)
        if self.data_manager.experience["available"] < cost:
            messagebox.showerror(
                "Za mało PD",
                f"Dodanie talentu kosztuje {cost} PD, masz {self.data_manager.experience['available']} PD.",
            )
            return

        if from_database:
            db_entry = game_data.get_talent(name) or {}
            added = self.data_manager.add_talent(
                name,
                advances=1,
                description=db_entry.get("description", ""),
                max_info=db_entry.get("max"),
                tests=db_entry.get("tests"),
                source=db_entry.get("source", "Lista"),
                is_custom=False,
                is_new=True,
            )
        else:
            info = custom_info or {}
            added = self.data_manager.add_talent(
                name,
                advances=1,
                description=info.get("description", ""),
                max_info=info.get("max"),
                tests=None,
                source="Własny",
                is_custom=True,
                is_new=True,
            )

        if not added:
            messagebox.showwarning("Już istnieje", f"Talent '{name}' już jest na liście.")
            return

        self.pending_changes["new_talents"].add(name)
        self.pending_changes["talent_changes"][name] = (
            self.pending_changes["talent_changes"].get(name, 0) + 1
        )
        self.data_manager.experience["available"] -= cost

        self.initialized_talents = False
        self.initialize_talents_display()
        self.update_experience_display()
        self.history_manager.add_entry("Dodano talent", name)
        self.refresh_history_display()

    def on_add_custom_talent(self) -> None:
        """Okno tworzenia własnego talentu (nazwa, opis, limit Max)."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Dodaj własny talent")
        dialog.geometry("520x560")
        dialog.transient(self)
        dialog.grab_set()

        ctk.CTkLabel(dialog, text="Nazwa talentu:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(16, 2)
        )
        name_entry = ctk.CTkEntry(dialog, width=480, height=34)
        name_entry.pack(padx=16, pady=(0, 8))

        ctk.CTkLabel(dialog, text="Opis:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(0, 2)
        )
        desc_box = ctk.CTkTextbox(dialog, width=480, height=160)
        desc_box.pack(padx=16, pady=(0, 8))

        ctk.CTkLabel(dialog, text="Rodzaj Maksimum:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(0, 2)
        )
        max_type_var = tk.StringVar(value="brak")
        max_type_combo = ctk.CTkComboBox(
            dialog,
            values=["brak", "liczba", "bonus z cechy"],
            variable=max_type_var,
            width=480,
        )
        max_type_combo.pack(padx=16, pady=(0, 8))

        value_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        value_frame.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(value_frame, text="Wartość liczbowa:").pack(side="left", padx=(0, 6))
        number_entry = ctk.CTkEntry(value_frame, width=80)
        number_entry.insert(0, "1")
        number_entry.pack(side="left", padx=(0, 12))
        ctk.CTkLabel(value_frame, text="Cecha:").pack(side="left", padx=(0, 6))
        attr_combo = ctk.CTkComboBox(value_frame, values=ATTRIBUTES, width=90)
        attr_combo.set(ATTRIBUTES[0])
        attr_combo.pack(side="left")

        def confirm() -> None:
            name = name_entry.get().strip()
            if not name:
                messagebox.showwarning("Brak nazwy", "Podaj nazwę talentu.", parent=dialog)
                return
            if name in self.data_manager.talents:
                messagebox.showwarning(
                    "Już istnieje", f"Talent '{name}' już jest na liście.", parent=dialog
                )
                return

            max_kind = max_type_var.get()
            if max_kind == "liczba":
                try:
                    value = int(number_entry.get())
                except ValueError:
                    messagebox.showwarning(
                        "Błędna wartość", "Podaj liczbę dla limitu.", parent=dialog
                    )
                    return
                max_info = {"type": "fixed", "value": value}
            elif max_kind == "bonus z cechy":
                max_info = {"type": "characteristic", "attr": attr_combo.get()}
            else:
                max_info = None

            description = desc_box.get("1.0", "end").strip()
            dialog.destroy()
            self._add_talent_and_register(
                name,
                from_database=False,
                custom_info={"description": description, "max": max_info},
            )

        button_row = ctk.CTkFrame(dialog, fg_color="transparent")
        button_row.pack(fill="x", padx=16, pady=(8, 16))
        ctk.CTkButton(
            button_row, text="Dodaj", command=confirm, fg_color=COLOR_SUCCESS, width=120
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            button_row, text="Anuluj", command=dialog.destroy,
            fg_color=COLOR_SURFACE_SOFT, hover_color="#48505e", width=120,
        ).pack(side="right")

        name_entry.focus_set()

    def on_delete_talent(self, talent_name: str) -> None:
        """Usuwa talent dodany w bieżącej sesji i zwraca zarezerwowane PD."""
        talent = self.data_manager.talents.get(talent_name)
        if not talent:
            return
        if not talent.get("is_new", False):
            messagebox.showerror(
                "Błąd", "Można usuwać tylko talenty dodane w tej sesji."
            )
            return

        result = messagebox.askyesno(
            "Potwierdzenie usunięcia",
            f"Czy usunąć talent '{talent_name}'?",
        )
        if not result:
            return

        out_of_profession, gm_approved = self._talent_cost_mode()
        refund = calculate_talent_cost(talent["advances"], out_of_profession, gm_approved)
        self.data_manager.experience["available"] += refund

        self.pending_changes["new_talents"].discard(talent_name)
        self.pending_changes["talent_changes"].pop(talent_name, None)
        del self.data_manager.talents[talent_name]

        if talent_name in self.talent_rows:
            self.talent_rows[talent_name]["row"].destroy()
            del self.talent_rows[talent_name]

        self.apply_talent_filter()
        self.update_experience_display()
        self.history_manager.add_entry("Usunięto talent", talent_name)
        self.refresh_history_display()

    def create_costs_tab(self) -> None:
        """Zakładka: Tabela kosztów rozwinięć."""
        _, wrapper = self._create_tab_shell(
            "Koszty Rozwinięć",
            "Koszty rozwinięć",
            "Na górze możesz policzyć dowolną liczbę rozwinięć dla wybranej cechy lub umiejętności. Poniżej masz pełną tabelę szybkiego porównania dla progów 5, 10, 15 i 20.",
        )

        calculator_section = self._create_collapsible_section(
            wrapper,
            "Kalkulator dowolnych rozwinięć",
            "Koszt aktualizuje się na żywo i uwzględnia obecny poziom rozwinięć.",
            default_open=True,
        )
        calculator_frame = calculator_section["content"]

        ctk.CTkLabel(
            calculator_frame,
            text="KALKULATOR DOWOLNYCH ROZWINIĘĆ",
            font=FONT_TITLE,
            text_color=COLOR_ACCENT,
        ).pack(anchor="w", padx=12, pady=(10, 6))

        ctk.CTkLabel(
            calculator_frame,
            text="Wybierz cechę albo umiejętność i wpisz liczbę rozwinięć. Koszt aktualizuje się na żywo.",
            font=FONT_BODY,
        ).pack(anchor="w", padx=12, pady=(0, 8))

        calculator_inputs = ctk.CTkFrame(calculator_frame)
        calculator_inputs.pack(fill="x", padx=12, pady=(0, 10))

        ctk.CTkLabel(calculator_inputs, text="Cecha / Umiejętność:").pack(
            side="left", padx=(8, 6), pady=8
        )
        self.cost_combo = ctk.CTkComboBox(
            calculator_inputs,
            values=[],
            width=260,
            command=lambda _choice: self.on_calculate_cost(),
        )
        self.cost_combo.pack(side="left", padx=(0, 10), pady=8)

        ctk.CTkLabel(calculator_inputs, text="Liczba rozwinięć:").pack(
            side="left", padx=(8, 6), pady=8
        )
        self.cost_spinbox = ctk.CTkEntry(calculator_inputs, width=120)
        self.cost_spinbox.insert(0, "1")
        self.cost_spinbox.pack(side="left", padx=(0, 10), pady=8)
        self.cost_spinbox.bind("<KeyRelease>", lambda _event: self.on_calculate_cost())

        calculator_status = ctk.CTkFrame(calculator_frame)
        calculator_status.pack(fill="x", padx=12, pady=(0, 12))

        self.cost_type_label = ctk.CTkLabel(calculator_status, text="Typ: -")
        self.cost_type_label.pack(side="left", padx=(8, 12), pady=8)

        self.cost_current_label = ctk.CTkLabel(
            calculator_status, text="Aktualne rozwinięcia: -"
        )
        self.cost_current_label.pack(side="left", padx=(0, 12), pady=8)

        self.cost_result_label = ctk.CTkLabel(
            calculator_status,
            text="Koszt: -",
            font=("Arial", 13, "bold"),
            text_color=COLOR_SUCCESS,
        )
        self.cost_result_label.pack(side="right", padx=(12, 8), pady=8)

        table_section = self._create_collapsible_section(
            wrapper,
            "Szybkie tabele kosztów",
            "Filtrowanie działa jednocześnie dla cech i umiejętności w tabeli porównawczej.",
            default_open=True,
            expand=True,
        )
        table_content = table_section["content"]

        filter_row = ctk.CTkFrame(table_content, fg_color="transparent")
        filter_row.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(filter_row, text="Filtr tabeli:", font=FONT_BODY_BOLD).pack(side="left", padx=(0, 8))
        self.cost_filter_entry = ctk.CTkEntry(
            filter_row,
            textvariable=self.cost_filter_var,
            placeholder_text="Wpisz fragment nazwy cechy lub umiejętności",
            width=340,
            height=36,
        )
        self.cost_filter_entry.pack(side="left", padx=(0, 8))
        self.cost_filter_entry.bind("<KeyRelease>", lambda _event: self.refresh_costs_display())
        ctk.CTkButton(
            filter_row,
            text="Wyczyść filtr",
            command=self.clear_cost_filter,
            width=120,
            height=36,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left")

        self.costs_header_frame = ctk.CTkFrame(
            table_content,
            fg_color=COLOR_SURFACE_SOFT,
            corner_radius=12,
        )
        self.costs_header_frame.pack(fill="x", pady=(0, 8))
        self._build_costs_header(self.costs_header_frame)

        self.costs_frame = ctk.CTkScrollableFrame(
            table_content,
            fg_color=COLOR_SURFACE,
            corner_radius=16,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.costs_frame.pack(fill="both", expand=True)

        self.refresh_cost_options()

    def refresh_cost_options(self) -> None:
        """Odświeża listę pozycji dostępnych w kalkulatorze kosztów."""
        if not self.cost_options_dirty and hasattr(self, "cost_combo"):
            self.on_calculate_cost()
            return

        options = list(self.data_manager.attributes.keys()) + sorted(
            self.data_manager.skills.keys(),
            key=str.casefold,
        )

        self.cost_combo.configure(values=options or [""])
        self.cost_options_dirty = False

        current_selection = self.cost_combo.get().strip()
        if current_selection not in options:
            self.cost_combo.set(options[0] if options else "")

        self.on_calculate_cost()

    def get_cost_preview(self, selection: str, advancements: int) -> Optional[Dict[str, int | str]]:
        """Zwraca dane do podglądu kosztu dla wybranej cechy lub umiejętności."""
        if advancements <= 0:
            return None

        if selection in self.data_manager.attributes:
            current_adv = self.data_manager.attributes[selection]["advanced"]
            return {
                "type": "Cecha",
                "current_adv": current_adv,
                "cost": calculate_advancement_cost("cecha", current_adv, advancements),
            }

        if selection in self.data_manager.skills:
            current_adv = self.data_manager.skills[selection]["advanced"]
            return {
                "type": "Umiejętność",
                "current_adv": current_adv,
                "cost": calculate_advancement_cost(
                    "umiejetnosc", current_adv, advancements
                ),
            }

        return None

    def _build_cost_row(
        self,
        parent,
        label: str,
        current_adv: int,
        advancement_type: str,
        category_short: str = "",
        current_value: Optional[int] = None,
    ) -> None:
        """Tworzy pojedynczy wiersz tabeli kosztów."""
        row = ctk.CTkFrame(parent, fg_color=COLOR_SURFACE_ALT, corner_radius=12)
        row.pack(fill="x", pady=3)
        ctk.CTkLabel(
            row,
            text=category_short,
            font=FONT_BODY_BOLD,
            width=72,
            text_color=COLOR_TEXT_MUTED,
        ).pack(side="left", padx=(10, 4), pady=6)
        ctk.CTkLabel(
            row,
            text=label,
            font=FONT_BODY_BOLD,
            width=250,
            anchor="w",
            justify="left",
        ).pack(side="left", padx=(0, 6), pady=6)
        ctk.CTkLabel(
            row,
            text=str(current_value) if current_value is not None else "-",
            font=FONT_BODY_BOLD,
            width=96,
            text_color=COLOR_HIGHLIGHT,
        ).pack(side="left", padx=(0, 6), pady=6)

        for num_adv in [5, 10, 15, 20]:
            cost = calculate_advancement_cost(advancement_type, current_adv, num_adv)
            ctk.CTkLabel(
                row,
                text=str(cost),
                width=78,
                font=FONT_BODY,
                text_color=COLOR_ACCENT,
            ).pack(side="left", padx=2, pady=6)

    def _build_costs_header(self, parent) -> None:
        """Tworzy stały nagłówek tabeli kosztów poza obszarem scrollowania."""
        ctk.CTkLabel(parent, text="Typ", font=FONT_BODY_BOLD, width=72).pack(side="left", padx=(10, 4), pady=8)
        ctk.CTkLabel(parent, text="Cecha / Umiejętność", font=FONT_BODY_BOLD, width=250).pack(side="left", padx=(0, 6), pady=8)
        ctk.CTkLabel(parent, text="Aktualna wartość", font=FONT_BODY_BOLD, width=96).pack(side="left", padx=(0, 6), pady=8)
        ctk.CTkLabel(parent, text="5 rozw.", font=FONT_BODY_BOLD, width=78).pack(side="left", padx=2, pady=8)
        ctk.CTkLabel(parent, text="10 rozw.", font=FONT_BODY_BOLD, width=78).pack(side="left", padx=2, pady=8)
        ctk.CTkLabel(parent, text="15 rozw.", font=FONT_BODY_BOLD, width=78).pack(side="left", padx=2, pady=8)
        ctk.CTkLabel(parent, text="20 rozw.", font=FONT_BODY_BOLD, width=78).pack(side="left", padx=2, pady=8)

    def _get_cost_table_skill_entries(self, filter_text: str) -> List[Tuple[str, Dict]]:
        """Zwraca odfiltrowane umiejętności do tabeli kosztów."""
        entries = []
        for skill_name, skill_data in self.data_manager.skills.items():
            skill_filter_blob = f"{skill_name} {get_skill_category(skill_name)}"
            if filter_text and filter_text not in normalize_search_text(skill_filter_blob):
                continue
            entries.append((skill_name, skill_data))
        return entries

    def refresh_costs_display(self) -> None:
        """Odświeża tabelę kosztów."""
        self.refresh_cost_options()
        filter_text = normalize_search_text(self.cost_filter_var.get())
        self.cost_layout_mode = max(self.costs_frame.winfo_width(), self.winfo_width()) >= 1500

        # Wyczyść ramkę
        for widget in self.costs_frame.winfo_children():
            widget.destroy()

        # Koszty dla cech
        ctk.CTkLabel(self.costs_frame, text="─ CECHY ─", font=FONT_SECTION, text_color=COLOR_SUCCESS).pack(anchor="w", padx=6, pady=(8, 5))

        for attr_name in self.data_manager.attributes.keys():
            searchable_name = f"{attr_name} {ATTRIBUTE_DETAILS[attr_name]}"
            if filter_text and filter_text not in normalize_search_text(searchable_name):
                continue
            current_adv = self.data_manager.attributes[attr_name]["advanced"]
            profession_suffix = " +" if self.data_manager.attributes[attr_name].get("profession_available") else ""
            label = f"{ATTRIBUTE_DETAILS[attr_name]} ({attr_name}){profession_suffix}"
            current_value = self.data_manager.attributes[attr_name]["initial"] + current_adv
            self._build_cost_row(self.costs_frame, label, current_adv, "cecha", "Cecha", current_value)

        # Koszty dla umiejętności
        ctk.CTkLabel(self.costs_frame, text="─ UMIEJĘTNOŚCI ─", font=FONT_SECTION, text_color=COLOR_SUCCESS).pack(anchor="w", padx=6, pady=(12, 5))

        skill_entries = self._get_cost_table_skill_entries(filter_text)
        is_two_column = self.cost_layout_mode and len(skill_entries) > 8

        grouped_skill_entries = {
            SKILL_CATEGORY_BASIC: [],
            SKILL_CATEGORY_ADVANCED: [],
        }
        for skill_name, skill_data in skill_entries:
            grouped_skill_entries[get_skill_category(skill_name)].append((skill_name, skill_data))

        if is_two_column:
            columns = ctk.CTkFrame(self.costs_frame, fg_color="transparent")
            columns.pack(fill="both", expand=True)
            left_column = ctk.CTkFrame(columns, fg_color="transparent")
            right_column = ctk.CTkFrame(columns, fg_color="transparent")
            left_column.pack(side="left", fill="both", expand=True, padx=(0, 8))
            separator = ctk.CTkFrame(columns, fg_color=COLOR_BORDER, width=2)
            separator.pack(side="left", fill="y", padx=4)
            right_column.pack(side="left", fill="both", expand=True, padx=(8, 0))

            ordered_skill_entries = (
                grouped_skill_entries[SKILL_CATEGORY_BASIC]
                + grouped_skill_entries[SKILL_CATEGORY_ADVANCED]
            )
            midpoint = (len(ordered_skill_entries) + 1) // 2
            column_data = [ordered_skill_entries[:midpoint], ordered_skill_entries[midpoint:]]
            column_frames = [left_column, right_column]
        else:
            single_column = ctk.CTkFrame(self.costs_frame, fg_color="transparent")
            single_column.pack(fill="both", expand=True)
            column_data = [None]
            column_frames = [single_column]

        if is_two_column:
            for column_frame, entries in zip(column_frames, column_data):
                for skill_name, skill_data in entries:
                    self._build_cost_row(
                        column_frame,
                        self._format_skill_display_name(skill_name),
                        skill_data["advanced"],
                        "umiejetnosc",
                        get_skill_category_short(skill_name),
                        skill_data["initial"] + skill_data["advanced"],
                    )
        else:
            for category_name, title in (
                (SKILL_CATEGORY_BASIC, "Podstawowe"),
                (SKILL_CATEGORY_ADVANCED, "Zaawansowane / grupowe"),
            ):
                entries = grouped_skill_entries[category_name]
                if not entries:
                    continue
                ctk.CTkLabel(
                    single_column,
                    text=title,
                    font=FONT_BODY_BOLD,
                    text_color=COLOR_TEXT_MUTED,
                ).pack(anchor="w", padx=6, pady=(6, 2))
                for skill_name, skill_data in entries:
                    self._build_cost_row(
                        single_column,
                        self._format_skill_display_name(skill_name),
                        skill_data["advanced"],
                        "umiejetnosc",
                        get_skill_category_short(skill_name),
                        skill_data["initial"] + skill_data["advanced"],
                    )

    def on_add_skill_dialog(self) -> None:
        """Okno dodawania nowej umiejętności (przeniesione z osobnej zakładki)."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Dodaj umiejętność")
        dialog.geometry("520x360")
        dialog.transient(self)
        dialog.grab_set()

        ctk.CTkLabel(dialog, text="Nazwa umiejętności:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(16, 2)
        )
        name_entry = ctk.CTkEntry(
            dialog, width=480, height=34,
            placeholder_text="np. Broń Biała (Kawaleryjska)",
        )
        name_entry.pack(padx=16, pady=(0, 8))

        ctk.CTkLabel(dialog, text="Atrybut:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(0, 2)
        )
        attr_combo = ctk.CTkComboBox(dialog, values=ATTRIBUTES, width=480)
        attr_combo.set("")
        attr_combo.pack(padx=16, pady=(0, 8))

        ctk.CTkLabel(dialog, text="Rozwinięcia:", font=FONT_BODY_BOLD).pack(
            anchor="w", padx=16, pady=(0, 2)
        )
        advanced_entry = ctk.CTkEntry(dialog, width=120, height=34)
        advanced_entry.insert(0, "0")
        advanced_entry.pack(anchor="w", padx=16, pady=(0, 8))

        ctk.CTkLabel(
            dialog,
            text="Wartość początkowa będzie równa wartości przypisanej cechy.",
            text_color=COLOR_TEXT_MUTED,
            font=FONT_SMALL,
            wraplength=480,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 8))

        def confirm() -> None:
            skill_name = name_entry.get().strip()
            attribute = attr_combo.get()
            validation_error = self._validate_skill_form(skill_name, attribute)
            if validation_error:
                messagebox.showerror("Błąd", validation_error, parent=dialog)
                return
            try:
                advanced = int(advanced_entry.get())
            except ValueError:
                messagebox.showerror("Błąd", "Wartość numeryczna musi być liczbą.", parent=dialog)
                return
            if advanced < 0:
                messagebox.showerror("Błąd", "Liczba rozwinięć nie może być ujemna.", parent=dialog)
                return
            dialog.destroy()
            self._create_skill(skill_name, attribute, advanced)

        button_row = ctk.CTkFrame(dialog, fg_color="transparent")
        button_row.pack(fill="x", padx=16, pady=(8, 16))
        ctk.CTkButton(
            button_row, text="Dodaj", command=confirm, fg_color=COLOR_SUCCESS, width=120
        ).pack(side="right", padx=(8, 0))
        ctk.CTkButton(
            button_row, text="Anuluj", command=dialog.destroy,
            fg_color=COLOR_SURFACE_SOFT, hover_color="#48505e", width=120,
        ).pack(side="right")

    def _create_skill(self, skill_name: str, attribute: str, advanced: int) -> None:
        """Dodaje umiejętność do postaci i odświeża widoki."""
        initial = self.data_manager.attributes[attribute]["initial"]
        self.data_manager.add_skill(skill_name, attribute, initial, advanced, is_new=True)
        self.pending_changes["new_skills"].add(skill_name)

        self.history_manager.add_entry(
            "Dodano umiejętność",
            f"{skill_name} ({attribute}) - Pocz: {initial}, Rozw: {advanced}",
        )
        messagebox.showinfo("Sukces", f"Umiejętność '{skill_name}' dodana!")

        self.initialized_skills = False
        self.cost_options_dirty = True
        self.initialize_skills_display()
        self.refresh_history_display()
        self.refresh_costs_if_visible(force=True)
    def create_experience_tab(self) -> None:
        """Zakładka: Zarządzanie doświadczeniem."""
        _, wrapper = self._create_tab_shell(
            "Doświadczenie",
            "Zarządzanie doświadczeniem",
            "Dodawaj pulę PD ręcznie albo szybkim przyciskiem. Wszystkie zmiany od razu wpływają na maksymalną liczbę możliwych rozwinięć.",
        )

        frame = ctk.CTkFrame(
            wrapper,
            corner_radius=18,
            fg_color=COLOR_SURFACE_ALT,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        frame.pack(fill="both", expand=True)

        ctk.CTkLabel(frame, text="Panel doświadczenia", font=FONT_TITLE).pack(anchor="w", padx=18, pady=(18, 6))
        self.exp_total_label = ctk.CTkLabel(
            frame,
            text="Łączna pula postaci: 0 PD",
            font=FONT_BODY,
            text_color=COLOR_TEXT_MUTED,
        )
        self.exp_total_label.pack(anchor="w", padx=18, pady=(0, 10))

        # Aktualne doświadczenie
        ctk.CTkLabel(frame, text="Dostępne PD:", font=FONT_BODY_BOLD).pack(anchor="w", padx=18, pady=(10, 0))
        exp_value_frame = ctk.CTkFrame(frame, fg_color="transparent")
        exp_value_frame.pack(fill="x", padx=18, pady=5)

        self.exp_entry = ctk.CTkEntry(exp_value_frame)
        self.exp_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.exp_entry.bind("<Return>", lambda _event: self.on_set_experience_value())

        ctk.CTkButton(
            exp_value_frame,
            text="Ustaw wartość",
            command=self.on_set_experience_value,
            width=120,
            height=36,
            fg_color=COLOR_SURFACE_SOFT,
            hover_color="#48505e",
            font=FONT_BODY_BOLD,
        ).pack(side="left")

        # Dodaj doświadczenie
        ctk.CTkLabel(frame, text="Dodaj doświadczenie:", font=FONT_BODY_BOLD).pack(anchor="w", padx=18, pady=(10, 0))
        add_frame = ctk.CTkFrame(frame)
        add_frame.pack(fill="x", padx=18, pady=5)

        self.exp_add_entry = ctk.CTkEntry(add_frame, placeholder_text="Ilość PD")
        self.exp_add_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))

        ctk.CTkButton(
            add_frame, text="Dodaj", command=self.on_add_experience, width=100
        ).pack(side="left")

        # Przyciski szybkie
        quick_frame = ctk.CTkFrame(frame)
        quick_frame.pack(fill="x", padx=18, pady=(10, 18))

        for amount in [10, 25, 50, 100]:
            ctk.CTkButton(
                quick_frame, text=f"+{amount}", width=80,
                command=lambda a=amount: self.on_quick_experience(a)
            ).pack(side="left", padx=2)

    def create_history_tab(self) -> None:
        """Zakładka: Historia działań."""
        _, wrapper = self._create_tab_shell(
            "Historia",
            "Historia działań",
            "Tutaj zobaczysz chronologiczny zapis ważnych operacji: wczytań, zatwierdzeń, cofnięć, dodawania PD i zmian na umiejętnościach.",
        )

        header = ctk.CTkFrame(wrapper, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))
        self.status_chip = self._create_info_chip(
            header,
            "Brak oczekujących zmian",
            COLOR_SUCCESS,
        )
        self.source_chip = self._create_info_chip(
            header,
            "Źródło: Excel",
            COLOR_SURFACE_SOFT,
        )

        self.history_text = ctk.CTkTextbox(
            wrapper,
            fg_color=COLOR_SURFACE,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        self.history_text.pack(fill="both", expand=True)
        self.history_text.configure(state="disabled")

    # ========================================================================
    # EVENT HANDLERS
    # ========================================================================

    def initialize_attributes_display(self) -> None:
        """Tworzy wiersze cech jeden raz (bez destroy). Wywoływane przy ładowaniu danych."""
        if self.initialized_attrs:
            self._update_all_attribute_labels()
            return
            
        # Wyczyść starą cache jeśli jest
        for widget in self.attributes_frame.winfo_children():
            widget.destroy()
        self.attribute_rows.clear()

        for attr_name, attr_data in self.data_manager.attributes.items():
            self._create_attribute_row_cached(attr_name, attr_data)

        self.initialized_attrs = True
        self.apply_attribute_filter()

    def _create_attribute_row_cached(self, attr_name: str, attr_data: Dict) -> None:
        """Tworzy wiersz cechy i przechowuje referencje do widgetów."""
        row = ctk.CTkFrame(
            self.attributes_frame,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=14,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        row.pack(fill="x", pady=5, padx=10)

        labels = []
        buttons = []

        # Nazwa cechy
        lbl_name = ctk.CTkLabel(
            row, text=self._format_attribute_display_name(attr_name), font=FONT_BODY_BOLD, width=80, anchor="w"
        )
        lbl_name.pack(side="left", padx=10, pady=10)
        labels.append(("name", lbl_name))

        # Wartość początkowa
        lbl_initial = ctk.CTkLabel(
            row, text=f"Pocz: {attr_data['initial']}", width=82, font=FONT_BODY
        )
        lbl_initial.pack(side="left", padx=2)
        labels.append(("initial", lbl_initial))

        # Rozwinięcia
        current_adv = attr_data["advanced"]
        lbl_adv = ctk.CTkLabel(
            row, text=f"Rozw: {current_adv}", width=82, font=FONT_BODY_BOLD, text_color=COLOR_ACCENT
        )
        lbl_adv.pack(side="left", padx=2)
        labels.append(("adv", lbl_adv))

        # Bieżąca wartość
        lbl_current = ctk.CTkLabel(
            row, text=f"Wartość: {attr_data['initial'] + current_adv}", width=110, font=FONT_BODY
        )
        lbl_current.pack(side="left", padx=2)
        labels.append(("current", lbl_current))

        # Maksymalne rozwinięcia
        max_adv = self.get_max_advancements("cecha", attr_name)
        lbl_max = ctk.CTkLabel(
            row, text=f"Maks: {max_adv}", width=82, font=FONT_BODY_BOLD, text_color=COLOR_WARNING
        )
        lbl_max.pack(side="left", padx=2)
        labels.append(("max", lbl_max))

        # Przyciski
        btn_plus1 = ctk.CTkButton(
            row, text="+1", width=46, height=34,
            command=lambda: self.on_increase_attribute(attr_name, 1),
            fg_color=COLOR_SUCCESS if self.can_afford_attribute(attr_name) else COLOR_ERROR
        )
        btn_plus1.pack(side="left", padx=1)
        buttons.append(("plus1", btn_plus1))

        btn_plus5 = ctk.CTkButton(
            row, text="+5", width=46, height=34,
            command=lambda: self.on_increase_attribute(attr_name, 5),
            fg_color=COLOR_SUCCESS if max_adv >= 5 else COLOR_ERROR
        )
        btn_plus5.pack(side="left", padx=1)
        buttons.append(("plus5", btn_plus5))

        btn_minus1 = ctk.CTkButton(
            row, text="-1", width=46, height=34,
            command=lambda: self.on_decrease_attribute(attr_name, 1),
            fg_color=COLOR_WARNING if current_adv > attr_data["base_advanced"] else "#555555"
        )
        btn_minus1.pack(side="left", padx=1)
        buttons.append(("minus1", btn_minus1))

        btn_minus5 = ctk.CTkButton(
            row, text="-5", width=46, height=34,
            command=lambda: self.on_decrease_attribute(attr_name, 5),
            fg_color=COLOR_WARNING if current_adv > attr_data["base_advanced"] else "#555555"
        )
        btn_minus5.pack(side="left", padx=1)
        buttons.append(("minus5", btn_minus5))

        self.attribute_rows[attr_name] = {
            "row": row,
            "labels": labels,
            "buttons": buttons,
        }
        self._apply_row_developable_style(row, self._attr_developable(attr_name))

    def _update_all_attribute_labels(self) -> None:
        """Aktualizuje tylko tekst labelek bez tworzenia nowych widgetów. SZYBKIE."""
        for attr_name, attr_data in self.data_manager.attributes.items():
            if attr_name in self.attribute_rows:
                row_data = self.attribute_rows[attr_name]
                labels = dict(row_data["labels"])
                
                current_adv = attr_data["advanced"]
                max_adv = self.get_max_advancements("cecha", attr_name)
                
                # Aktualizuj wartości
                labels["name"].configure(text=self._format_attribute_display_name(attr_name))
                labels["adv"].configure(text=f"Rozw: {current_adv}")
                labels["current"].configure(text=f"Wartość: {attr_data['initial'] + current_adv}")
                labels["max"].configure(text=f"Maks: {max_adv}")
                
                # Aktualizuj kolory przycisków
                buttons = dict(row_data["buttons"])
                buttons["plus1"].configure(fg_color=COLOR_SUCCESS if self.can_afford_attribute(attr_name) else COLOR_ERROR)
                buttons["plus5"].configure(fg_color=COLOR_SUCCESS if max_adv >= 5 else COLOR_ERROR)
                buttons["minus1"].configure(fg_color=COLOR_WARNING if current_adv > attr_data["base_advanced"] else "#555555")
                buttons["minus5"].configure(fg_color=COLOR_WARNING if current_adv > attr_data["base_advanced"] else "#555555")
                self._apply_row_developable_style(row_data["row"], self._attr_developable(attr_name))

    def refresh_attributes_display(self) -> None:
        """Compatibility wrapper - teraz tylko aktualizuje, nie niszczy widgety."""
        if self.initialized_attrs:
            self._update_all_attribute_labels()
            self.apply_attribute_filter()
        else:
            self.initialize_attributes_display()

    def apply_attribute_filter(self) -> None:
        """Pokazuje/ukrywa cechy zależnie od filtra 'tylko rozwijalne (+)'."""
        if not hasattr(self, "attributes_frame"):
            return
        profession_only = bool(self.attr_profession_filter_var.get())
        for attr_name, row_data in self.attribute_rows.items():
            row_data["row"].pack_forget()
            if not profession_only or self._attr_developable(attr_name):
                row_data["row"].pack(fill="x", pady=5, padx=10)

    def initialize_skills_display(self) -> None:
        """Tworzy wiersze umiejętności jeden raz (bez destroy). Wywoływane przy ładowaniu danych."""
        if self.initialized_skills:
            self._update_all_skill_labels()
            return
            
        # Wyczyść starą cache jeśli jest
        self._build_skill_sections()
        self.skill_rows.clear()

        for skill_name, skill_data in self.data_manager.skills.items():
            self._create_skill_row_cached(skill_name, skill_data)

        self.initialized_skills = True
        self.apply_skill_filter()

    def _create_skill_row_cached(self, skill_name: str, skill_data: Dict) -> None:
        """Tworzy wiersz umiejętności i przechowuje referencje do widgetów."""
        category = get_skill_category(skill_name)
        parent_frame = self.skills_group_frames[category]["content"]
        row = ctk.CTkFrame(
            parent_frame,
            fg_color=COLOR_SURFACE_ALT,
            corner_radius=14,
            border_width=1,
            border_color=COLOR_BORDER,
        )
        row.pack(fill="x", pady=5, padx=10)

        labels = []
        buttons = []
        minimum_advancement = self._get_minimum_skill_advancement(skill_data)

        # Nazwa umiejętności
        lbl_name = ctk.CTkLabel(
            row, text=self._format_skill_display_name(skill_name), font=FONT_BODY_BOLD, width=240, anchor="w"
        )
        lbl_name.pack(side="left", padx=(10, 4), pady=10)
        labels.append(("name", lbl_name))

        # Atrybut
        lbl_attr = ctk.CTkLabel(row, text=f"({skill_data['attribute']})", width=52, font=FONT_BODY)
        lbl_attr.pack(side="left", padx=2)
        labels.append(("attr", lbl_attr))

        lbl_category = ctk.CTkLabel(
            row,
            text=get_skill_category_short(skill_name),
            width=72,
            font=FONT_BODY_BOLD,
            text_color=COLOR_TEXT_MUTED,
        )
        lbl_category.pack(side="left", padx=(2, 6))
        labels.append(("category", lbl_category))

        # Wartość początkowa
        lbl_initial = ctk.CTkLabel(
            row, text=f"Pocz: {skill_data['initial']}", width=82, font=FONT_BODY
        )
        lbl_initial.pack(side="left", padx=2)
        labels.append(("initial", lbl_initial))

        # Rozwinięcia
        current_adv = skill_data["advanced"]
        lbl_adv = ctk.CTkLabel(
            row, text=f"Rozw: {current_adv}", width=82, font=FONT_BODY_BOLD, text_color=COLOR_ACCENT
        )
        lbl_adv.pack(side="left", padx=2)
        labels.append(("adv", lbl_adv))

        # Bieżąca wartość
        lbl_current = ctk.CTkLabel(
            row, text=f"Suma: {skill_data['initial'] + current_adv}", width=82, font=FONT_BODY
        )
        lbl_current.pack(side="left", padx=2)
        labels.append(("current", lbl_current))

        # Maksymalne rozwinięcia
        max_adv = self.get_max_advancements("umiejetnosc", skill_name)
        lbl_max = ctk.CTkLabel(
            row, text=f"Maks: {max_adv}", width=82, font=FONT_BODY_BOLD, text_color=COLOR_WARNING
        )
        lbl_max.pack(side="left", padx=2)
        labels.append(("max", lbl_max))

        # Przyciski
        btn_plus1 = ctk.CTkButton(
            row, text="+1", width=46, height=34,
            command=lambda: self.on_increase_skill(skill_name, 1),
            fg_color=COLOR_SUCCESS if self.can_afford_skill(skill_name) else COLOR_ERROR
        )
        btn_plus1.pack(side="left", padx=1)
        buttons.append(("plus1", btn_plus1))

        btn_plus5 = ctk.CTkButton(
            row, text="+5", width=46, height=34,
            command=lambda: self.on_increase_skill(skill_name, 5),
            fg_color=COLOR_SUCCESS if max_adv >= 5 else COLOR_ERROR
        )
        btn_plus5.pack(side="left", padx=1)
        buttons.append(("plus5", btn_plus5))

        btn_minus1 = ctk.CTkButton(
            row, text="-1", width=46, height=34,
            command=lambda: self.on_decrease_skill(skill_name, 1),
            fg_color=COLOR_WARNING if current_adv > minimum_advancement else "#555555"
        )
        btn_minus1.pack(side="left", padx=1)
        buttons.append(("minus1", btn_minus1))

        btn_minus5 = ctk.CTkButton(
            row, text="-5", width=46, height=34,
            command=lambda: self.on_decrease_skill(skill_name, 5),
            fg_color=COLOR_WARNING if current_adv > minimum_advancement else "#555555"
        )
        btn_minus5.pack(side="left", padx=1)
        buttons.append(("minus5", btn_minus5))
        
        # Przycisk usuń - tylko dla umiejętności dodanych
        if skill_data.get("is_new", False):
            btn_delete = ctk.CTkButton(
                row, text="Usuń", width=62, height=34,
                command=lambda: self.on_delete_skill(skill_name),
                fg_color=COLOR_ERROR
            )
            btn_delete.pack(side="left", padx=2)
            buttons.append(("delete", btn_delete))

        self.skill_rows[skill_name] = {
            "row": row,
            "group": category,
            "labels": labels,
            "buttons": buttons,
        }
        self._apply_row_developable_style(row, self._skill_developable(skill_name))

    def _update_all_skill_labels(self) -> None:
        """Aktualizuje tylko tekst labelek bez tworzenia nowych widgetów. SZYBKIE."""
        for skill_name, skill_data in self.data_manager.skills.items():
            if skill_name in self.skill_rows:
                row_data = self.skill_rows[skill_name]
                labels = dict(row_data["labels"])
                
                current_adv = skill_data["advanced"]
                max_adv = self.get_max_advancements("umiejetnosc", skill_name)
                minimum_advancement = self._get_minimum_skill_advancement(skill_data)
                
                # Aktualizuj wartości
                labels["name"].configure(text=self._format_skill_display_name(skill_name))
                labels["adv"].configure(text=f"Rozw: {current_adv}")
                labels["current"].configure(text=f"Suma: {skill_data['initial'] + current_adv}")
                labels["max"].configure(text=f"Maks: {max_adv}")
                labels["category"].configure(text=get_skill_category_short(skill_name))
                
                # Aktualizuj kolory przycisków
                buttons = dict(row_data["buttons"])
                buttons["plus1"].configure(fg_color=COLOR_SUCCESS if self.can_afford_skill(skill_name) else COLOR_ERROR)
                buttons["plus5"].configure(fg_color=COLOR_SUCCESS if max_adv >= 5 else COLOR_ERROR)
                buttons["minus1"].configure(fg_color=COLOR_WARNING if current_adv > minimum_advancement else "#555555")
                buttons["minus5"].configure(fg_color=COLOR_WARNING if current_adv > minimum_advancement else "#555555")
                self._apply_row_developable_style(row_data["row"], self._skill_developable(skill_name))

    def _get_minimum_skill_advancement(self, skill_data: Dict) -> int:
        """Zwraca minimalny dozwolony poziom rozwinięcia dla umiejętności."""
        if skill_data.get("is_new", False):
            return 0
        return skill_data["base_advanced"]

    def refresh_skills_display(self) -> None:
        """Compatibility wrapper - teraz tylko aktualizuje, nie niszczy widgety."""
        if self.initialized_skills:
            self._update_all_skill_labels()
            if self._has_active_skill_filters():
                self.apply_skill_filter()
            else:
                self.skill_summary_var.set(
                    f"Wyświetlane umiejętności: {len(self.skill_rows)} / {len(self.skill_rows)}"
                )
        else:
            self.initialize_skills_display()

    def clear_skill_filter(self) -> None:
        """Czyści aktywny filtr listy umiejętności."""
        self.skill_filter_var.set("")
        self.skill_attribute_filter_var.set(ATTRIBUTE_FILTER_ALL)
        self.skill_profession_filter_var.set(False)
        self.apply_skill_filter()

    def clear_cost_filter(self) -> None:
        """Czyści filtr tabeli kosztów."""
        self.cost_filter_var.set("")
        self.refresh_costs_display()

    def apply_skill_filter(self) -> None:
        """Filtruje widoczne umiejętności po nazwie lub atrybucie."""
        if not hasattr(self, "skills_frame"):
            return

        query = normalize_search_text(self.skill_filter_var.get())
        selected_attribute = self.skill_attribute_filter_var.get()
        profession_only = self.skill_profession_filter_var.get()
        exact_attribute = None
        if selected_attribute and selected_attribute != ATTRIBUTE_FILTER_ALL:
            exact_attribute = get_attribute_code_from_filter_label(selected_attribute)

        suggested_attribute = self._get_attribute_suggestion(query) if query else None
        visible_count = 0
        total_count = len(self.skill_rows)
        visible_names_by_group = {
            SKILL_CATEGORY_BASIC: [],
            SKILL_CATEGORY_ADVANCED: [],
        }

        for skill_name, row_data in self.skill_rows.items():
            skill_data = self.data_manager.skills.get(skill_name, {})
            searchable_text = normalize_search_text(
                f"{skill_name} {skill_data.get('attribute', '')} {get_skill_category(skill_name)} {get_skill_category_short(skill_name)}"
            )
            matches_query = not query or query in searchable_text
            matches_attribute = not exact_attribute or skill_data.get("attribute") == exact_attribute
            matches_profession = not profession_only or self._skill_developable(skill_name)
            should_show = matches_query and matches_attribute and matches_profession
            row_data["row"].pack_forget()
            if should_show:
                visible_count += 1
                visible_names_by_group[row_data["group"]].append(skill_name)

        for group_name, group_data in self.skills_group_frames.items():
            group_data["section"].pack_forget()
            if visible_names_by_group[group_name]:
                group_data["section"].pack(fill="x", padx=10, pady=(8, 6))
                for skill_name in visible_names_by_group[group_name]:
                    self.skill_rows[skill_name]["row"].pack(fill="x", pady=5, padx=10)

        if suggested_attribute and exact_attribute != suggested_attribute:
            self.skill_filter_hint_button.configure(
                text=f"Filtruj dokładnie po: {get_attribute_filter_label(suggested_attribute)}",
            )
            self.skill_filter_hint_button.pack(side="left", padx=(0, 12), pady=14)
        else:
            self.skill_filter_hint_button.pack_forget()

        self.skill_summary_var.set(
            f"Wyświetlane umiejętności: {visible_count} / {total_count}"
        )

    def _get_attribute_suggestion(self, query: str) -> Optional[str]:
        """Zwraca sugerowany atrybut dla wpisanego tekstu filtra."""
        if not query:
            return None

        for attr_code, attr_name in ATTRIBUTE_DETAILS.items():
            variants = {
                normalize_search_text(attr_code),
                normalize_search_text(attr_name),
            }
            if any(value.startswith(query) or query == value for value in variants):
                return attr_code
        return None

    def apply_suggested_attribute_filter(self) -> None:
        """Ustawia dokładny filtr atrybutu na podstawie bieżącej podpowiedzi."""
        suggested_attribute = self._get_attribute_suggestion(
            normalize_search_text(self.skill_filter_var.get())
        )
        if not suggested_attribute:
            return
        self.skill_attribute_filter_var.set(get_attribute_filter_label(suggested_attribute))
        self.apply_skill_filter()

    def on_increase_attribute(self, attr_name: str, amount: int = 1) -> None:
        """Zwiększa rozwinięcie cechy (pokazowe, do zatwierdzenia)."""
        self._increase_advancement("cecha", attr_name, amount)

    def on_increase_skill(self, skill_name: str, amount: int = 1) -> None:
        """Zwiększa rozwinięcie umiejętności (pokazowe, do zatwierdzenia)."""
        self._increase_advancement("umiejetnosc", skill_name, amount)

    def on_decrease_attribute(self, attr_name: str, amount: int = 1) -> None:
        """Zmniejsza rozwinięcie cechy (cofanie zmian)."""
        self._decrease_advancement("cecha", attr_name, amount)

    def on_decrease_skill(self, skill_name: str, amount: int = 1) -> None:
        """Zmniejsza rozwinięcie umiejętności (cofanie zmian)."""
        self._decrease_advancement("umiejetnosc", skill_name, amount)

    def _increase_advancement(self, advancement_type: str, item_name: str, amount: int) -> None:
        """Helper do zwiększania rozwiniec."""
        if advancement_type == "cecha":
            data = self.data_manager.attributes[item_name]
            pending_key = "attribute_changes"
        else:
            data = self.data_manager.skills[item_name]
            pending_key = "skill_changes"
        
        out_of_profession, gm_approved = self._advancement_cost_mode(advancement_type, item_name)
        cost = calculate_advancement_cost(
            advancement_type, data["advanced"], amount, out_of_profession, gm_approved
        )

        if self.data_manager.experience["available"] < cost:
            messagebox.showerror("Za mało PD", f"Brak wystarczającego doświadczenia. Potrzebujesz {cost} PD, masz {self.data_manager.experience['available']} PD.")
            return

        if item_name not in self.pending_changes[pending_key]:
            self.pending_changes[pending_key][item_name] = 0

        self.pending_changes[pending_key][item_name] += amount
        data["advanced"] += amount
        self.data_manager.experience["available"] -= cost

        self.refresh_attributes_display() if advancement_type == "cecha" else self.refresh_skills_display()
        self.update_experience_display()
        self.refresh_costs_if_visible()

    def _decrease_advancement(self, advancement_type: str, item_name: str, amount: int) -> None:
        """Helper do zmniejszania rozwiniec."""
        if advancement_type == "cecha":
            data = self.data_manager.attributes[item_name]
            pending_key = "attribute_changes"
            minimum_advancement = data["base_advanced"]
        else:
            data = self.data_manager.skills[item_name]
            pending_key = "skill_changes"
            minimum_advancement = self._get_minimum_skill_advancement(data)

        current_adv = data["advanced"]

        if current_adv - amount < minimum_advancement:
            play_blocked_sound()
            return

        out_of_profession, gm_approved = self._advancement_cost_mode(advancement_type, item_name)
        refund = calculate_advancement_cost(
            advancement_type, current_adv - amount, amount, out_of_profession, gm_approved
        )

        if item_name not in self.pending_changes[pending_key]:
            self.pending_changes[pending_key][item_name] = 0

        self.pending_changes[pending_key][item_name] -= amount
        data["advanced"] -= amount
        self.data_manager.experience["available"] += refund

        self.refresh_attributes_display() if advancement_type == "cecha" else self.refresh_skills_display()
        self.update_experience_display()
        self.refresh_costs_if_visible()

    def can_afford_attribute(self, attr_name: str) -> bool:
        """Sprawdza, czy stać postać na rozwinięcie cechy."""
        current_adv = self.data_manager.attributes[attr_name]["advanced"]
        out_of_profession, gm_approved = self._advancement_cost_mode("cecha", attr_name)
        cost = calculate_advancement_cost("cecha", current_adv, 1, out_of_profession, gm_approved)
        return self.data_manager.experience["available"] >= cost

    def can_afford_skill(self, skill_name: str) -> bool:
        """Sprawdza, czy stać postać na rozwinięcie umiejętności."""
        current_adv = self.data_manager.skills[skill_name]["advanced"]
        out_of_profession, gm_approved = self._advancement_cost_mode("umiejetnosc", skill_name)
        cost = calculate_advancement_cost("umiejetnosc", current_adv, 1, out_of_profession, gm_approved)
        return self.data_manager.experience["available"] >= cost

    def get_max_advancements(self, advancement_type: str, attr_name_or_skill: str) -> int:
        """Oblicza maksymalną liczbę rozwinięć możliwych przy aktualnym PD."""
        if advancement_type == "cecha":
            current_adv = self.data_manager.attributes[attr_name_or_skill]["advanced"]
        else:
            current_adv = self.data_manager.skills[attr_name_or_skill]["advanced"]

        out_of_profession, gm_approved = self._advancement_cost_mode(advancement_type, attr_name_or_skill)
        max_adv = 0
        available = self.data_manager.experience["available"]
        
        while True:
            cost = calculate_advancement_cost(
                advancement_type, current_adv, max_adv + 1, out_of_profession, gm_approved
            )
            if available >= cost:
                max_adv += 1
            else:
                break
        
        return max_adv
    
    def _get_button_color_for_increase(self, advancement_type: str, item_name: str, amount: int) -> str:
        """Zwraca kolor przycisku +X na podstawie możliwości."""
        if advancement_type == "cecha":
            can_afford = self.can_afford_attribute(item_name)
        else:
            can_afford = self.can_afford_skill(item_name)
        
        if not can_afford:
            return COLOR_ERROR
        
        max_adv = self.get_max_advancements(advancement_type, item_name)
        return COLOR_SUCCESS if max_adv >= amount else COLOR_ERROR
    
    def _get_button_color_for_decrease(self, advancement_type: str, item_name: str) -> str:
        """Zwraca kolor przycisku -X na podstawie możliwości cofnięcia."""
        if advancement_type == "cecha":
            current_adv = self.data_manager.attributes[item_name]["advanced"]
            base_adv = self.data_manager.attributes[item_name]["base_advanced"]
        else:
            current_adv = self.data_manager.skills[item_name]["advanced"]
            base_adv = self._get_minimum_skill_advancement(
                self.data_manager.skills[item_name]
            )
        
        return COLOR_WARNING if current_adv > base_adv else "#555555"

    def on_confirm_changes(self) -> None:
        """Potwierdza wszystkie oczekujące zmiany."""
        if not self.has_pending_changes():
            messagebox.showinfo("Brak zmian", "Nie ma żadnych oczekujących zmian do zatwierdzenia.")
            return

        # Kalkulacja całkowitego kosztu
        total_cost = 0
        details = []

        for attr_name, change_count in self.pending_changes["attribute_changes"].items():
            if change_count > 0:
                current_adv = (
                    self.data_manager.attributes[attr_name]["advanced"] - change_count
                )
                oop, gm = self._advancement_cost_mode("cecha", attr_name)
                cost = calculate_advancement_cost("cecha", current_adv, change_count, oop, gm)
                total_cost += cost
                note = " [spoza profesji ×2]" if (oop and not gm) else (" [spoza profesji, zgoda MG]" if oop else "")
                details.append(f"Cecha {attr_name}: +{change_count} rozwinięć ({cost} PD){note}")

        for skill_name, change_count in self.pending_changes["skill_changes"].items():
            if change_count > 0:
                current_adv = (
                    self.data_manager.skills[skill_name]["advanced"] - change_count
                )
                oop, gm = self._advancement_cost_mode("umiejetnosc", skill_name)
                cost = calculate_advancement_cost("umiejetnosc", current_adv, change_count, oop, gm)
                total_cost += cost
                note = " [spoza profesji ×2]" if (oop and not gm) else (" [spoza profesji, zgoda MG]" if oop else "")
                details.append(f"Umiejętność {skill_name}: +{change_count} rozwinięć ({cost} PD){note}")

        out_of_profession, gm_approved = self._talent_cost_mode()
        for talent_name, change_count in self.pending_changes["talent_changes"].items():
            if change_count > 0:
                cost = calculate_talent_cost(change_count, out_of_profession, gm_approved)
                total_cost += cost
                details.append(
                    f"Talent {talent_name}: +{change_count} wykupień ({cost} PD)"
                )

        for skill_name in sorted(self.pending_changes["new_skills"], key=str.casefold):
            skill_data = self.data_manager.skills.get(skill_name)
            if skill_data:
                details.append(
                    f"Nowa umiejętność {skill_name} ({skill_data['attribute']})"
                )

        for talent_name in sorted(self.pending_changes["new_talents"], key=str.casefold):
            if talent_name in self.data_manager.talents:
                details.append(f"Nowy talent {talent_name}")

        experience_delta = self.pending_changes["experience_delta"]
        if experience_delta:
            direction = "+" if experience_delta > 0 else ""
            details.append(f"Dostępne PD ustawione / zmienione: {direction}{experience_delta} PD")

        # Potwierdzenie
        details_text = "\n".join(details)
        result = messagebox.askyesno(
            "Potwierdzenie zmian",
            f"Koszt zmian: {total_cost} PD\n\n{details_text}\n\nCzy zatwierdzić zmiany?"
        )

        if result:
            self.data_manager.experience["spent"] += total_cost
            self.data_manager.experience["total"] = (
                self.data_manager.experience["available"]
                + self.data_manager.experience["spent"]
            )

            # Reset oczekujących zmian
            for attr_name in self.pending_changes["attribute_changes"]:
                self.data_manager.attributes[attr_name]["base_advanced"] = (
                    self.data_manager.attributes[attr_name]["advanced"]
                )

            for skill_name in self.pending_changes["skill_changes"]:
                self.data_manager.skills[skill_name]["base_advanced"] = (
                    self.data_manager.skills[skill_name]["advanced"]
                )

            for skill_name in self.pending_changes["new_skills"]:
                if skill_name in self.data_manager.skills:
                    self.data_manager.skills[skill_name]["base_advanced"] = (
                        self.data_manager.skills[skill_name]["advanced"]
                    )
                    self.data_manager.skills[skill_name]["is_new"] = False

            for talent_name in self.pending_changes["talent_changes"]:
                if talent_name in self.data_manager.talents:
                    self.data_manager.talents[talent_name]["base_advances"] = (
                        self.data_manager.talents[talent_name]["advances"]
                    )

            for talent_name in self.pending_changes["new_talents"]:
                if talent_name in self.data_manager.talents:
                    self.data_manager.talents[talent_name]["base_advances"] = (
                        self.data_manager.talents[talent_name]["advances"]
                    )
                    self.data_manager.talents[talent_name]["is_new"] = False

            self.pending_changes = self._create_empty_pending_changes()

            # Zapisz do aktualnego źródła
            if self.data_manager.file_path:
                if self.data_manager.source_type == FILE_TYPE_PDF:
                    self.data_manager.save_to_pdf(self.data_manager.file_path)
                else:
                    self.data_manager.save_to_excel(self.data_manager.file_path)

            self.history_manager.add_entry("Zatwierdzono zmiany", f"Koszt: {total_cost} PD")

            messagebox.showinfo("Sukces", "Zmiany zatwierdzone!")
            self.refresh_attributes_display()
            self.refresh_skills_display()
            self.initialized_talents = False
            self.initialize_talents_display()
            self.update_experience_display()
            self.refresh_history_display()
            self.refresh_costs_if_visible(force=True)

    def on_revert_changes(self) -> None:
        """Cofa wszystkie oczekujące zmiany."""
        if not self.has_pending_changes():
            messagebox.showinfo("Brak zmian", "Nie ma żadnych zmian do cofnięcia.")
            return

        result = messagebox.askyesno(
            "Potwierdzenie cofnięcia",
            "Czy na pewno chcesz cofnąć wszystkie oczekujące zmiany?"
        )

        if result:
            # Oblicz całkowity koszt do przywrócenia
            total_refund = 0
            
            # Przywróć cechy i oblicz refund
            for attr_name in self.pending_changes["attribute_changes"]:
                change_count = self.pending_changes["attribute_changes"][attr_name]
                if change_count > 0:
                    current_adv = self.data_manager.attributes[attr_name]["advanced"]
                    oop, gm = self._advancement_cost_mode("cecha", attr_name)
                    refund = calculate_advancement_cost("cecha", current_adv - change_count, change_count, oop, gm)
                    total_refund += refund
                
                self.data_manager.attributes[attr_name]["advanced"] = (
                    self.data_manager.attributes[attr_name]["base_advanced"]
                )

            # Przywróć umiejętności i oblicz refund
            for skill_name in self.pending_changes["skill_changes"]:
                change_count = self.pending_changes["skill_changes"][skill_name]
                if change_count > 0:
                    current_adv = self.data_manager.skills[skill_name]["advanced"]
                    oop, gm = self._advancement_cost_mode("umiejetnosc", skill_name)
                    refund = calculate_advancement_cost("umiejetnosc", current_adv - change_count, change_count, oop, gm)
                    total_refund += refund
                
                self.data_manager.skills[skill_name]["advanced"] = (
                    self.data_manager.skills[skill_name]["base_advanced"]
                )
            
            # Usuń umiejętności, które były dodane (nie mają base_advanced)
            skills_to_remove = list(self.pending_changes["new_skills"])
            for skill_name in skills_to_remove:
                if skill_name in self.skill_rows:
                    self.skill_rows[skill_name]["row"].destroy()
                    del self.skill_rows[skill_name]
            
            for skill_name in skills_to_remove:
                if skill_name in self.data_manager.skills:
                    del self.data_manager.skills[skill_name]

            # Przywróć talenty i oblicz refund
            out_of_profession, gm_approved = self._talent_cost_mode()
            for talent_name, change_count in self.pending_changes["talent_changes"].items():
                if change_count > 0:
                    total_refund += calculate_talent_cost(
                        change_count, out_of_profession, gm_approved
                    )
                if talent_name in self.data_manager.talents:
                    self.data_manager.talents[talent_name]["advances"] = (
                        self.data_manager.talents[talent_name].get("base_advances", 0)
                    )

            # Usuń talenty dodane w tej sesji
            talents_to_remove = list(self.pending_changes["new_talents"])
            for talent_name in talents_to_remove:
                if talent_name in self.talent_rows:
                    self.talent_rows[talent_name]["row"].destroy()
                    del self.talent_rows[talent_name]
                if talent_name in self.data_manager.talents:
                    del self.data_manager.talents[talent_name]

            # Przywróć doświadczenie
            self.data_manager.experience["available"] += total_refund
            self.data_manager.experience["available"] -= self.pending_changes["experience_delta"]
            self.data_manager.experience["total"] -= self.pending_changes["experience_delta"]
            
            self.pending_changes = self._create_empty_pending_changes()

            self.history_manager.add_entry("Cofnięto zmiany", f"Przywrócono {total_refund} PD")

            messagebox.showinfo("Sukces", "Zmiany cofnięte!")
            self.refresh_attributes_display()
            self.refresh_skills_display()
            self.refresh_talents_display()
            self.update_experience_display()
            self.refresh_history_display()
            self.refresh_costs_if_visible(force=True)

    def on_delete_skill(self, skill_name: str) -> None:
        """Usuwa umiejętność (tylko dodane)."""
        if skill_name not in self.data_manager.skills:
            return
        
        skill_data = self.data_manager.skills[skill_name]
        if not skill_data.get("is_new", False):
            messagebox.showerror("Błąd", "Można usuwać tylko umiejętności dodane przez gracza.")
            return
        
        result = messagebox.askyesno(
            "Potwierdzenie usunięcia",
            f"Czy chcesz usunąć umiejętność '{skill_name}'?"
        )
        
        if result:
            del self.data_manager.skills[skill_name]
            self.pending_changes["new_skills"].discard(skill_name)
            self.pending_changes["skill_changes"].pop(skill_name, None)
            if skill_name in self.skill_rows:
                self.skill_rows[skill_name]["row"].destroy()
                del self.skill_rows[skill_name]
            
            self.history_manager.add_entry(
                "Usunięto umiejętność",
                skill_name
            )
            messagebox.showinfo("Sukces", f"Umiejętność '{skill_name}' usunięta!")
            self.refresh_skills_display()
            self.refresh_history_display()
            self.cost_options_dirty = True
            self.refresh_costs_if_visible(force=True)

    def _validate_skill_form(self, skill_name: str, attribute: str) -> Optional[str]:
        """Waliduje formularz dodania umiejętności. Zwraca None jeśli OK, inaczej komunikat błędu."""
        if not skill_name:
            return "Wpisz nazwę umiejętności."
        
        if not attribute:
            return "Wybierz atrybut."

        normalized_new_name = normalize_skill_name(skill_name)
        for existing_name in self.data_manager.skills:
            if normalize_skill_name(existing_name) == normalized_new_name:
                return f"Umiejętność o podobnej nazwie już istnieje: {existing_name}."
        
        return None

    def on_add_experience(self) -> None:
        """Dodaje doświadczenie."""
        try:
            amount = int(self.exp_add_entry.get())
        except ValueError:
            messagebox.showerror("Błąd", "Wpisz poprawną liczbę.")
            return

        if amount <= 0:
            messagebox.showerror("Błąd", "Ilość musi być większa od 0.")
            return

        self.data_manager.experience["available"] += amount
        self.data_manager.experience["total"] += amount
        self.pending_changes["experience_delta"] += amount

        self.exp_add_entry.delete(0, "end")

        self.history_manager.add_entry("Dodano doświadczenie", f"+{amount} PD")

        messagebox.showinfo("Sukces", f"Dodano {amount} PD!")
        self.update_experience_display()
        self.refresh_attributes_display()
        self.refresh_skills_display()
        self.refresh_history_display()
        self.refresh_costs_if_visible()

    def on_quick_experience(self, amount: int) -> None:
        """Szybkie dodanie doświadczenia."""
        self.data_manager.experience["available"] += amount
        self.data_manager.experience["total"] += amount
        self.pending_changes["experience_delta"] += amount

        self.history_manager.add_entry("Dodano doświadczenie", f"+{amount} PD")

        self.update_experience_display()
        self.refresh_attributes_display()
        self.refresh_skills_display()
        self.refresh_history_display()
        self.refresh_costs_if_visible()

    def on_set_experience_value(self) -> None:
        """Ustawia dokładną wartość dostępnego doświadczenia."""
        try:
            target_value = int(self.exp_entry.get())
        except ValueError:
            messagebox.showerror("Błąd", "Wpisz poprawną liczbę PD.")
            self.update_experience_display()
            return

        if target_value < 0:
            messagebox.showerror("Błąd", "Dostępne PD nie może być ujemne.")
            self.update_experience_display()
            return

        current_value = self.data_manager.experience["available"]
        delta = target_value - current_value
        if delta == 0:
            return

        self.data_manager.experience["available"] = target_value
        self.data_manager.experience["total"] += delta
        self.pending_changes["experience_delta"] += delta

        direction = "Ustawiono dostępne doświadczenie"
        self.history_manager.add_entry(direction, f"{current_value} -> {target_value} PD")

        self.update_experience_display()
        self.refresh_attributes_display()
        self.refresh_skills_display()
        self.refresh_history_display()
        self.refresh_costs_if_visible()

    def on_calculate_cost(self) -> None:
        """Oblicza koszt rozwinięć dla wybranej cechy/umiejętności."""
        selection = self.cost_combo.get().strip()
        try:
            advancements = int(self.cost_spinbox.get())
        except ValueError:
            self.cost_type_label.configure(text="Typ: -")
            self.cost_current_label.configure(text="Aktualne rozwinięcia: -")
            self.cost_result_label.configure(text="Koszt: -")
            return

        if not selection or advancements <= 0:
            self.cost_type_label.configure(text="Typ: -")
            self.cost_current_label.configure(text="Aktualne rozwinięcia: -")
            self.cost_result_label.configure(text="Koszt: -")
            return

        preview = self.get_cost_preview(selection, advancements)
        if not preview:
            self.cost_type_label.configure(text="Typ: -")
            self.cost_current_label.configure(text="Aktualne rozwinięcia: -")
            self.cost_result_label.configure(text="Koszt: -")
            return

        self.cost_type_label.configure(text=f"Typ: {preview['type']}")
        self.cost_current_label.configure(
            text=f"Aktualne rozwinięcia: {preview['current_adv']}"
        )
        self.cost_result_label.configure(text=f"Koszt: {preview['cost']} PD")

    def on_load_character(self) -> None:
        """Ładuje postać z pliku Excel."""
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Obsługiwane pliki", "*.xlsx *.pdf"),
                ("Excel files", "*.xlsx"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ]
        )

        if not file_path:
            return

        extension = Path(file_path).suffix.lower()
        if extension == ".pdf":
            was_loaded = self.data_manager.load_from_pdf(file_path)
        else:
            was_loaded = self.data_manager.load_from_excel(file_path)

        if was_loaded:
            self.pending_changes = self._create_empty_pending_changes()
            # Reset flag aby reinicjalizować widgety z nowymi danymi
            self.initialized_attrs = False
            self.initialized_skills = False
            self.initialized_talents = False
            self.history_manager.set_current_character(self.data_manager.character_name)
            self._reconcile_career_path()
            self.history_manager.add_entry("Wczytano postać", Path(file_path).name)
            messagebox.showinfo("Sukces", "Postać wczytana!")
            self.initialize_attributes_display()
            self.initialize_skills_display()
            self.initialize_talents_display()
            self.refresh_profession_display()
            self.update_experience_display()
            self.refresh_history_display()
            self.update_character_info()
            self.cost_options_dirty = True
            self.refresh_costs_if_visible(force=True)
        else:
            messagebox.showerror("Błąd", "Nie można wczytać pliku.")

    def _reconcile_career_path(self) -> None:
        """Porównuje ścieżkę kariery z PDF ze ścieżką zapisaną w historii."""
        name = self.data_manager.character_name
        if not name:
            return
        stored = self.history_manager.get_career_path(name)
        current = self.data_manager.career_path

        def signature(path):
            return [
                (
                    (step.get("title") or step.get("profession") or "").strip().lower(),
                    step.get("level"),
                )
                for step in path
            ]

        if not stored:
            # Brak zapisu – zapisz ścieżkę z PDF.
            self._persist_career_path()
            return

        if signature(stored) == signature(current):
            return

        # Niezgodność – pozwól użytkownikowi wybrać źródło.
        use_pdf = messagebox.askyesno(
            "Niezgodność ścieżki kariery",
            "Ścieżka kariery z pliku różni się od zapisanej w historii.\n\n"
            f"Z pliku: {self._format_path_signature(current)}\n"
            f"Z historii: {self._format_path_signature(stored)}\n\n"
            "Tak = użyj ścieżki z pliku (i nadpisz historię)\n"
            "Nie = użyj ścieżki z historii",
        )
        if use_pdf:
            self._persist_career_path()
        else:
            self.data_manager.career_path = list(stored)
            if stored:
                last = stored[-1]
                self.data_manager.current_career = (
                    last.get("profession") or last.get("title") or self.data_manager.current_career
                )
                self.data_manager.current_career_level = last.get("level") or 1
                resolved_class = game_data.class_of_career(self.data_manager.current_career)
                if resolved_class:
                    self.data_manager.character_class = resolved_class

    @staticmethod
    def _format_path_signature(path) -> str:
        """Czytelny opis ścieżki kariery do komunikatów."""
        if not path:
            return "(pusta)"
        return " → ".join(
            f"{(step.get('title') or step.get('profession') or '?')}({step.get('level') or '?'})"
            for step in path
        )

    def on_save_character(self) -> None:
        """Zapisuje postać do pliku Excel."""
        if self.has_pending_changes():
            messagebox.showwarning(
                "Oczekujące zmiany",
                "Najpierw zatwierdź albo cofnij oczekujące zmiany, a dopiero potem zapisz plik.",
            )
            return

        file_path = self.data_manager.file_path
        default_extension = ".pdf" if self.data_manager.source_type == FILE_TYPE_PDF else ".xlsx"
        dialog_options = {
            "defaultextension": default_extension,
            "filetypes": [
                ("PDF files", "*.pdf"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
        }

        if self.data_manager.file_path:
            current_path = Path(self.data_manager.file_path)
            save_current = messagebox.askyesnocancel(
                "Zapis postaci",
                f"Czy zapisać do bieżącego pliku?\n\n{current_path.name}\n\nWybierz 'Nie', aby wskazać nową lokalizację.",
            )
            if save_current is None:
                return
            if not save_current:
                dialog_options["initialdir"] = str(current_path.parent)
                dialog_options["initialfile"] = current_path.name
                file_path = filedialog.asksaveasfilename(**dialog_options)
                if not file_path:
                    return
        else:
            file_path = filedialog.asksaveasfilename(**dialog_options)
            if not file_path:
                return

        extension = Path(file_path).suffix.lower()
        if extension == ".pdf" or self.data_manager.source_type == FILE_TYPE_PDF:
            save_ok = self.data_manager.save_to_pdf(file_path)
        else:
            save_ok = self.data_manager.save_to_excel(file_path)

        if save_ok:
            self.history_manager.add_entry("Zapisano postać", Path(file_path).name)
            messagebox.showinfo("Sukces", "Postać zapisana!")
            self.refresh_history_display()
        else:
            messagebox.showerror("Błąd", "Nie można zapisać pliku.")

    def on_new_character(self) -> None:
        """Tworzy nową postać."""
        result = messagebox.askyesno(
            "Nowa postać",
            "Czy chcesz stworzyć nową postać? Niezapisane zmiany zostaną utracone."
        )

        if result:
            self.data_manager.create_new_character()
            self.pending_changes = self._create_empty_pending_changes()
            self.history_manager.set_current_character(self.data_manager.character_name)
            self.history_manager.add_entry("Utworzono nową postać", "")
            messagebox.showinfo("Sukces", "Nowa postać gotowa!")
            self.initialized_attrs = False
            self.initialized_skills = False
            self.initialized_talents = False
            self.initialize_attributes_display()
            self.initialize_skills_display()
            self.initialize_talents_display()
            self.refresh_profession_display()
            self.update_experience_display()
            self.refresh_history_display()
            self.update_character_info()
            self.cost_options_dirty = True
            self.refresh_costs_if_visible(force=True)

    def update_experience_display(self) -> None:
        """Aktualizuje wyświetlanie doświadczenia."""
        available = self.data_manager.experience["available"]

        self.exp_entry.delete(0, "end")
        self.exp_entry.insert(0, str(available))
        self.refresh_header_summary()

    def update_character_info(self) -> None:
        """Aktualizuje informacje o postaci."""
        self.refresh_header_summary()

    def refresh_history_display(self) -> None:
        """Odświeża wyświetlanie historii."""
        self.history_text.configure(state="normal")
        self.history_text.delete("1.0", "end")
        self.history_text.insert("1.0", self.history_manager.get_history_text())
        self.history_text.configure(state="disabled")
        self.refresh_header_summary()


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    app = CharacterSheetApp()
    app.mainloop()
